import html
import json
import re
import subprocess
import sys
import threading
import time
import traceback
try:
    from importlib.metadata import PackageNotFoundError, version as get_package_version
except Exception:
    class PackageNotFoundError(Exception):
        pass

    def get_package_version(_package_name):
        raise PackageNotFoundError

from itertools import product
from pathlib import Path

DEPENDENCY_IMPORT_ERRORS = {}

try:
    import matplotlib.pyplot as plt
except Exception as exc:
    plt = None
    DEPENDENCY_IMPORT_ERRORS["matplotlib"] = exc

try:
    import pandas as pd
except Exception as exc:
    pd = None
    DEPENDENCY_IMPORT_ERRORS["pandas"] = exc

try:
    import questionary
    from questionary import Choice
except Exception as exc:
    questionary = None
    Choice = None
    DEPENDENCY_IMPORT_ERRORS["questionary"] = exc

try:
    import requests
except Exception as exc:
    requests = None
    DEPENDENCY_IMPORT_ERRORS["requests"] = exc

try:
    from openai import OpenAI
except Exception as exc:
    OpenAI = None
    DEPENDENCY_IMPORT_ERRORS["openai"] = exc

try:
    import msvcrt
except ImportError:
    msvcrt = None


NVIDIA_SMI_QUERY = [
    "nvidia-smi",
    "--query-gpu=index,name,memory.used,memory.total",
    "--format=csv,noheader,nounits",
]


PARAM_INFO = {
    "temperature": {
        "label": "溫度 (Temperature)",
        "range": "0.0 - 2.0",
        "desc": "調高更有創意，調低更穩定；Qwen 類模型常用 0.1 或 1.0。",
        "default": "0.1, 0.8",
        "backends": ["ollama", "llama.cpp"],
        "backend_keys": {"ollama": "temperature", "llama.cpp": "temperature"},
    },
    "num_ctx": {
        "label": "上下文長度 (Num_Ctx)",
        "range": "128 - 262144",
        "desc": "可測長上下文與 TTFT 影響；此項為 Ollama 請求級參數。",
        "default": "4096, 8192",
        "backends": ["ollama"],
        "backend_keys": {"ollama": "num_ctx"},
    },
    "num_predict": {
        "label": "最大生成 Token",
        "range": "-1, 1 - 4096",
        "desc": "控制回覆長度；llama.cpp 會映射為 n_predict。",
        "default": "256, 512",
        "backends": ["ollama", "llama.cpp"],
        "backend_keys": {"ollama": "num_predict", "llama.cpp": "n_predict"},
    },
    "top_p": {
        "label": "核心採樣 (Top_P)",
        "range": "0.0 - 1.0",
        "desc": "限制候選詞機率總和，數值越低越保守。",
        "default": "0.8, 0.95",
        "backends": ["ollama", "llama.cpp"],
        "backend_keys": {"ollama": "top_p", "llama.cpp": "top_p"},
    },
    "top_k": {
        "label": "Top-K 候選數",
        "range": "0 - 200",
        "desc": "限制只從前 K 個高機率 token 中取樣；llama.cpp 常用 40 左右。",
        "default": "20, 40",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "top_k"},
    },
    "min_p": {
        "label": "最小概率 (Min_P)",
        "range": "0.0 - 1.0",
        "desc": "過濾低機率雜訊詞；常見平衡點在 0.05 左右。",
        "default": "0.02, 0.05",
        "backends": ["ollama", "llama.cpp"],
        "backend_keys": {"ollama": "min_p", "llama.cpp": "min_p"},
    },
    "typical_p": {
        "label": "局部典型採樣 (Typical_P)",
        "range": "0.0 - 1.0",
        "desc": "llama.cpp 的 locally typical sampling；1.0 代表關閉。",
        "default": "1.0, 0.95",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "typical_p"},
    },
    "dynatemp_range": {
        "label": "動態溫度範圍 (Dynatemp Range)",
        "range": "0.0 - 2.0",
        "desc": "讓實際溫度在 temperature 上下浮動；0.0 代表關閉。",
        "default": "0.0, 0.5",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "dynatemp_range"},
    },
    "dynatemp_exponent": {
        "label": "動態溫度指數 (Dynatemp Exponent)",
        "range": "0.0 - 5.0",
        "desc": "調整 dynatemp 的變化曲線；通常搭配 dynatemp_range 一起測。",
        "default": "1.0, 2.0",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "dynatemp_exponent"},
    },
    "repeat_penalty": {
        "label": "重複懲罰 (Repeat Penalty)",
        "range": "1.0 - 2.0",
        "desc": "降低重複句與繞圈輸出。",
        "default": "1.05, 1.15",
        "backends": ["ollama", "llama.cpp"],
        "backend_keys": {"ollama": "repeat_penalty", "llama.cpp": "repeat_penalty"},
    },
    "repeat_last_n": {
        "label": "重複檢查視窗 (Repeat Last N)",
        "range": "-1, 0 - 4096",
        "desc": "llama.cpp 重複懲罰要回看多少 token；-1 代表使用 context size。",
        "default": "64, 256",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "repeat_last_n"},
    },
    "presence_penalty": {
        "label": "出現懲罰 (Presence Penalty)",
        "range": "-2.0 - 2.0",
        "desc": "降低已出現過主題再次被選中的機率；0.0 代表關閉。",
        "default": "0.0, 0.5",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "presence_penalty"},
    },
    "frequency_penalty": {
        "label": "頻率懲罰 (Frequency Penalty)",
        "range": "-2.0 - 2.0",
        "desc": "依出現次數加重懲罰，能更明顯壓制重複 token。",
        "default": "0.0, 0.5",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "frequency_penalty"},
    },
    "dry_multiplier": {
        "label": "DRY 倍率 (Dry Multiplier)",
        "range": "0.0 - 2.0",
        "desc": "Don't Repeat Yourself 懲罰強度；0.0 代表關閉。",
        "default": "0.0, 0.8",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "dry_multiplier"},
    },
    "dry_base": {
        "label": "DRY 基底 (Dry Base)",
        "range": "1.0 - 4.0",
        "desc": "DRY 懲罰成長基底，數值越大重複延伸時罰得越快。",
        "default": "1.75, 2.0",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "dry_base"},
    },
    "dry_allowed_length": {
        "label": "DRY 容許長度 (Dry Allowed Length)",
        "range": "0 - 32",
        "desc": "重複片段在多長之前不加重 DRY 懲罰。",
        "default": "2, 4",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "dry_allowed_length"},
    },
    "dry_penalty_last_n": {
        "label": "DRY 回看視窗 (Dry Penalty Last N)",
        "range": "-1, 0 - 4096",
        "desc": "DRY 要掃描多少 token；-1 代表使用 context size。",
        "default": "-1, 256",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "dry_penalty_last_n"},
    },
    "mirostat": {
        "label": "Mirostat 模式",
        "range": "0, 1, 2",
        "desc": "llama.cpp 的 Mirostat 採樣；0 關閉，1/2 為不同版本。",
        "default": "0, 2",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "mirostat"},
    },
    "mirostat_tau": {
        "label": "Mirostat 熵目標 (Tau)",
        "range": "0.0 - 10.0",
        "desc": "Mirostat 目標熵；越高通常越發散。",
        "default": "5.0, 8.0",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "mirostat_tau"},
    },
    "mirostat_eta": {
        "label": "Mirostat 學習率 (Eta)",
        "range": "0.01 - 1.0",
        "desc": "Mirostat 調整速度；越大反應越激進。",
        "default": "0.1, 0.3",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "mirostat_eta"},
    },
    "seed": {
        "label": "隨機種子 (Seed)",
        "range": "-1, 0 - 2147483647",
        "desc": "固定後可重現結果；-1 代表每次使用隨機 seed。",
        "default": "-1, 42",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "seed"},
    },
    "ignore_eos": {
        "label": "忽略 EOS (Ignore EOS)",
        "range": "enable | disable",
        "desc": "忽略結束 token 持續生成；通常只建議在特定壓測時使用。",
        "default": "disable, enable",
        "value_type": "boolean",
        "backends": ["llama.cpp"],
        "backend_keys": {"llama.cpp": "ignore_eos"},
    },
    "num_gpu": {
        "label": "GPU 層數 / 顯存卸載",
        "range": "0 - 100",
        "desc": "對 Ollama TPS 影響很大；llama.cpp 通常在 server 啟動時設定。",
        "default": "25, 50",
        "backends": ["ollama"],
        "backend_keys": {"ollama": "num_gpu"},
    },
    "enable_thinking": {
        "label": "Thinking / Reasoning 開關",
        "range": "enable | disable",
        "desc": "測試是否啟用模型的 thinking / reasoning 模式。Ollama 會送出 `think`，llama.cpp 會送入 `chat_template_kwargs.enable_thinking`。",
        "default": "disable, enable",
        "value_type": "boolean",
        "backends": ["ollama", "llama.cpp"],
        "backend_keys": {"ollama": "think", "llama.cpp": "enable_thinking"},
        "request_targets": {"ollama": "body", "llama.cpp": "chat_template_kwargs"},
    },
}

PARAM_GROUPS = {
    "🔥 生成核心": [
        "temperature",
        "num_ctx",
        "num_predict",
        "top_p",
        "top_k",
        "min_p",
        "typical_p",
        "dynatemp_range",
        "dynatemp_exponent",
    ],
    "⚖️ 採樣與懲罰": [
        "repeat_penalty",
        "repeat_last_n",
        "presence_penalty",
        "frequency_penalty",
        "dry_multiplier",
        "dry_base",
        "dry_allowed_length",
        "dry_penalty_last_n",
    ],
    "🌀 採樣策略與控制": [
        "mirostat",
        "mirostat_tau",
        "mirostat_eta",
        "seed",
        "ignore_eos",
    ],
    "🧠 Thinking / Reasoning": ["enable_thinking"],
    "🖥️ 硬體與部署": ["num_gpu"],
}


CAPABILITY_OPTIONS = {
    "chat": {
        "label": "聊天能力",
        "description": "一般對話輸出，沿用目前的文字串流 benchmark。",
        "default_prompt": (
            "解釋一下 3D 列印使用 PETG 時，長期受力下的潛變（creep）風險，"
            "以及有哪些實際的改善方式。"
        ),
    },
    "tools": {
        "label": "Tools 調用能力",
        "description": "要求模型先呼叫工具，檢查是否真的輸出 tool_calls。",
        "default_prompt": (
            "請查詢台北今天的天氣。若你支援 tools 或 function calling，"
            "請先呼叫 `lookup_weather` 工具，不要直接回答。"
        ),
    },
}

TOOL_BENCHMARK_TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "lookup_weather",
            "description": "Look up the current weather for a city.",
            "parameters": {
                "type": "object",
                "properties": {
                    "city": {
                        "type": "string",
                        "description": "City name in Chinese or English.",
                    },
                    "unit": {
                        "type": "string",
                        "enum": ["celsius", "fahrenheit"],
                        "description": "Preferred temperature unit.",
                    },
                },
                "required": ["city"],
            },
        },
    }
]


OLLAMA_HOST = "http://localhost:11434"
OLLAMA_BASE_URL = f"{OLLAMA_HOST}/v1"
DEFAULT_LLAMA_PORT = "8080"
BACKEND_CHECK_TIMEOUT_S = 3
BACKEND_LABELS = {
    "ollama": "Ollama",
    "llama.cpp": "llama.cpp (llama-server)",
}


def get_backend_display_name(backend):
    return BACKEND_LABELS.get(backend, backend)


def normalize_local_port(raw_port, default=DEFAULT_LLAMA_PORT):
    port_text = str(raw_port or default).strip() or default
    if not port_text.isdigit():
        return None, "Please enter a whole-number port between 1 and 65535. / 請輸入 1 到 65535 之間的整數端口。"

    port_number = int(port_text)
    if not 1 <= port_number <= 65535:
        return None, "The port must be between 1 and 65535. / 端口必須介於 1 到 65535 之間。"
    return str(port_number), None


def inspect_backend_readiness(backend, url):
    if backend == "ollama":
        tags_url = f"{OLLAMA_HOST}/api/tags"
        try:
            response = requests.get(tags_url, timeout=BACKEND_CHECK_TIMEOUT_S)
            response.raise_for_status()
            payload = response.json()
        except requests.RequestException as exc:
            return {
                "ok": False,
                "models": [],
                "message": (
                    f"Unable to connect to Ollama at {OLLAMA_HOST}. Please start `ollama serve` first, then retry. / "
                    f"無法連線到 {OLLAMA_HOST} 的 Ollama，請先啟動 `ollama serve`，再回來重試。"
                ),
                "detail": f"{type(exc).__name__}: {exc}",
            }
        except ValueError as exc:
            return {
                "ok": False,
                "models": [],
                "message": (
                    f"Ollama responded, but {tags_url} did not return valid JSON. / "
                    f"Ollama 有回應，但 {tags_url} 回傳的不是有效 JSON。"
                ),
                "detail": f"{type(exc).__name__}: {exc}",
            }

        if not isinstance(payload, dict):
            return {
                "ok": False,
                "models": [],
                "message": (
                    f"Ollama responded, but the payload from {tags_url} was not in the expected format. / "
                    f"Ollama 有回應，但 {tags_url} 的回傳格式不符合預期。"
                ),
                "detail": f"payload type: {type(payload).__name__}",
            }

        models = [model["name"] for model in payload.get("models", []) if model.get("name")]
        warning = None
        if not models:
            warning = (
                "Ollama is reachable, but `/api/tags` did not report any models. You can still type model names "
                "manually, but the benchmark will fail if those models are not installed locally. / "
                "Ollama 可正常連線，但 `/api/tags` 沒有回報任何模型。你仍可手動輸入模型名稱；若本機沒有"
                "安裝那些模型，benchmark 還是會失敗。"
            )
        return {"ok": True, "models": models, "warning": warning, "checked_url": tags_url}

    models_url = f"{url.rstrip('/')}/models"
    try:
        response = requests.get(models_url, timeout=BACKEND_CHECK_TIMEOUT_S)
        response.raise_for_status()
        payload = response.json()
    except requests.RequestException as exc:
        return {
            "ok": False,
            "models": [],
            "message": (
                f"Unable to connect to the llama.cpp OpenAI-compatible endpoint at {models_url}. "
                f"Please start `llama-server` on that port first, then retry. / 無法連線到 {models_url} 的 "
                f"llama.cpp OpenAI 相容端點，請先在該端口啟動 `llama-server`，再回來重試。"
            ),
            "detail": f"{type(exc).__name__}: {exc}",
        }
    except ValueError as exc:
        return {
            "ok": False,
            "models": [],
            "message": (
                f"The backend responded, but {models_url} did not return valid JSON. / "
                f"後端有回應，但 {models_url} 回傳的不是有效 JSON。"
            ),
            "detail": f"{type(exc).__name__}: {exc}",
        }

    if not isinstance(payload, dict):
        return {
            "ok": False,
            "models": [],
            "message": (
                f"The backend responded, but the `/v1/models` payload from {models_url} was not in the expected "
                f"format. / 後端有回應，但 {models_url} 的 `/v1/models` 回傳格式不符合預期。"
            ),
            "detail": f"payload type: {type(payload).__name__}",
        }

    data = payload.get("data", [])
    if not isinstance(data, list):
        return {
            "ok": False,
            "models": [],
            "message": (
                f"The backend responded, but the `/v1/models` payload from {models_url} was not in the expected "
                f"format. / 後端有回應，但 {models_url} 的 `/v1/models` 回傳格式不符合預期。"
            ),
            "detail": f"payload keys: {', '.join(sorted(payload.keys()))}",
        }

    models = []
    for item in data:
        if not isinstance(item, dict):
            continue
        model_id = str(item.get("id") or "").strip()
        if model_id:
            models.append(model_id)

    warning = None
    if not models:
        warning = (
            "The backend is reachable, but `/v1/models` did not report any model IDs. Please confirm the loaded "
            "model name before benchmarking. / 後端可正常連線，但 `/v1/models` 沒有回報任何模型 ID；"
            "開始 benchmark 前請先確認實際載入中的模型名稱。"
        )
    return {"ok": True, "models": models, "warning": warning, "checked_url": models_url}


def get_ollama_models():
    backend_status = inspect_backend_readiness("ollama", OLLAMA_BASE_URL)
    if not backend_status.get("ok"):
        return []
    return backend_status.get("models", [])


BOOLEAN_TRUE_ALIASES = {
    "1",
    "true",
    "on",
    "enable",
    "enabled",
    "yes",
    "y",
    "think",
    "thinking",
    "啟用",
    "開啟",
    "是",
}

BOOLEAN_FALSE_ALIASES = {
    "0",
    "false",
    "off",
    "disable",
    "disabled",
    "no",
    "n",
    "nothink",
    "關閉",
    "停用",
    "否",
}


def get_param_value_type(param_key):
    return PARAM_INFO.get(param_key, {}).get("value_type", "number")


def parse_boolean_value(raw_value):
    normalized = str(raw_value).strip().lower()
    if normalized in BOOLEAN_TRUE_ALIASES:
        return True
    if normalized in BOOLEAN_FALSE_ALIASES:
        return False
    raise ValueError(
        f"無法解析布林值：{raw_value}。請使用 enable/disable、true/false、on/off 或 1/0。"
    )


def format_param_value_for_display(param_key, value):
    if get_param_value_type(param_key) == "boolean":
        normalized_value = value
        if isinstance(value, str):
            normalized_value = parse_boolean_value(value)
        return "enable" if bool(normalized_value) else "disable"
    return str(value)


def format_param_values_for_display(param_key, values):
    return ", ".join(format_param_value_for_display(param_key, value) for value in values)


def resolve_thinking_mode(value):
    if value in (None, "", "N/A"):
        return "default"
    try:
        return "enable" if parse_boolean_value(value) else "disable"
    except ValueError:
        return str(value)


def get_thinking_mode_for_run(params):
    if not isinstance(params, dict) or "enable_thinking" not in params:
        return "default"
    return resolve_thinking_mode(params.get("enable_thinking"))


def get_param_request_target(param_key, backend):
    info = PARAM_INFO.get(param_key, {})
    request_targets = info.get("request_targets", {})
    if backend in request_targets:
        return request_targets[backend]
    return "options" if backend == "ollama" else "body"


def parse_csv_values(raw_text, param_key=None):
    value_type = get_param_value_type(param_key)
    values = []
    for item in raw_text.split(","):
        item = item.strip()
        if not item:
            continue
        if value_type == "boolean":
            values.append(parse_boolean_value(item))
            continue
        try:
            if "." in item or "e" in item.lower():
                value = float(item)
                values.append(int(value) if value.is_integer() else value)
            else:
                values.append(int(item))
        except ValueError as exc:
            raise ValueError(f"無法解析數值：{item}") from exc

    if not values:
        raise ValueError("至少需要一個測試值。")

    return values


def ask_param_values(param_key):
    info = PARAM_INFO[param_key]
    while True:
        raw_value = questionary.text(
            f"輸入 {info['label']} 測試值 (逗號隔開):",
            default=info["default"],
        ).ask()
        if raw_value is None:
            return None
        try:
            return parse_csv_values(raw_value, param_key=param_key)
        except ValueError as exc:
            print(f"⚠️ {exc} 請重新輸入。")


def format_param_dict(params):
    if not params:
        return "預設參數"
    joined = ", ".join(
        f"{key}={format_param_value_for_display(key, value)}" for key, value in params.items()
    )
    return "{" + joined + "}"


def build_backend_options(backend, params):
    backend_options = {}
    for key, value in params.items():
        backend_key = PARAM_INFO[key]["backend_keys"].get(backend)
        if backend_key:
            backend_options[backend_key] = value
    return backend_options


def build_backend_extra_body(backend, params):
    extra_body = {}
    options = {}
    chat_template_kwargs = {}

    for key, value in params.items():
        backend_key = PARAM_INFO[key]["backend_keys"].get(backend)
        if not backend_key:
            continue

        request_target = get_param_request_target(key, backend)
        if request_target == "options":
            options[backend_key] = value
        elif request_target == "chat_template_kwargs":
            chat_template_kwargs[backend_key] = value
        else:
            extra_body[backend_key] = value

    if backend == "ollama":
        if options:
            extra_body["options"] = options
        return extra_body

    if options:
        extra_body.update(options)
    if chat_template_kwargs:
        extra_body["chat_template_kwargs"] = chat_template_kwargs
    return extra_body


def build_ollama_modelfile_params(params):
    modelfile_params = {}
    for key, value in params.items():
        backend_key = PARAM_INFO[key]["backend_keys"].get("ollama")
        if not backend_key:
            continue
        if get_param_request_target(key, "ollama") != "options":
            continue
        modelfile_params[backend_key] = value
    return modelfile_params


SYSTEM_PROMPT_BLOCK_SEPARATOR = "---"
BACK_ACTION = "__back__"


def build_system_prompt_variants(system_prompts):
    prompts = [prompt.strip() for prompt in (system_prompts or []) if (prompt or "").strip()]
    if not prompts:
        return [{"label": "N/A", "text": ""}]
    return [{"label": f"SP{index}", "text": prompt} for index, prompt in enumerate(prompts, start=1)]


def build_benchmark_messages(capability, prompt, system_prompt_text=""):
    messages = []
    if (system_prompt_text or "").strip():
        messages.append({"role": "system", "content": system_prompt_text.strip()})
    if capability == "tools":
        messages.append(
            {
                "role": "system",
                "content": (
                    "You are being benchmarked for tool calling. "
                    "If a suitable tool is provided, call the tool before answering."
                ),
            }
        )
    messages.append({"role": "user", "content": prompt})
    return messages


def build_chat_request_payload(config, model, request_kwargs, system_prompt_text=""):
    capability = config.get("capability", "chat")
    payload = {
        "model": model,
        "messages": build_benchmark_messages(capability, config["prompt"], system_prompt_text),
        "stream": True,
        **request_kwargs,
    }
    if capability == "tools":
        payload["tools"] = TOOL_BENCHMARK_TOOLS
        payload["tool_choice"] = "auto"
    return payload


def parse_non_content_types(value):
    if value is None:
        return set()
    if isinstance(value, str):
        items = [item.strip() for item in value.split(",")]
    else:
        items = [str(item).strip() for item in value]
    return {item for item in items if item and item != "none"}


def adjust_classification_for_capability(classification, capability):
    if capability != "tools":
        return classification

    adjusted = classification.copy()
    non_content_types = parse_non_content_types(adjusted.get("Non_Content_Types"))
    if "tool_calls" in non_content_types:
        adjusted["Status"] = "ok"
        adjusted["Output_Category"] = "tool_call"
        finish_reason = adjusted.get("Finish_Reason") or "unknown"
        adjusted["Diagnosis"] = (
            "Received tool_calls payload during the tool benchmark "
            f"(finish_reason={finish_reason})."
        )
        return adjusted

    if adjusted["Status"] == "ok" and adjusted["Output_Category"] == "normal_content":
        adjusted["Status"] = "warning"
        adjusted["Output_Category"] = "text_reply_without_tool"
        adjusted["Diagnosis"] = (
            "Received textual content, but no tool_calls payload was emitted during the "
            "tool benchmark."
        )
    elif adjusted["Output_Category"] == "empty_reply":
        adjusted["Diagnosis"] = (
            "The tool benchmark finished without textual content or tool_calls payload."
        )
    elif adjusted["Output_Category"] == "non_content_stream":
        adjusted["Diagnosis"] = (
            "The tool benchmark returned non-content payloads, but none were tool_calls."
        )

    return adjusted


def query_nvidia_vram_snapshot():
    try:
        result = subprocess.run(
            NVIDIA_SMI_QUERY,
            capture_output=True,
            text=True,
            check=True,
            timeout=3,
        )
    except (FileNotFoundError, subprocess.SubprocessError):
        return None

    snapshot = []
    for line in result.stdout.splitlines():
        parts = [part.strip() for part in line.split(",", 3)]
        if len(parts) != 4:
            continue
        try:
            snapshot.append(
                {
                    "index": int(parts[0]),
                    "name": parts[1],
                    "memory_used_mib": int(parts[2]),
                    "memory_total_mib": int(parts[3]),
                }
            )
        except ValueError:
            continue

    return snapshot or None


def empty_vram_metrics():
    return {
        "VRAM_Base_MiB": None,
        "VRAM_Peak_MiB": None,
        "VRAM_Delta_MiB": None,
        "VRAM_Detail": "N/A",
    }


def format_mib_value(value):
    return "N/A" if value is None or pd.isna(value) else f"{int(value)} MiB"


def format_numeric_value(value, digits):
    if value is None or pd.isna(value):
        return "N/A"
    return f"{float(value):.{digits}f}"


def format_probability_value(value, digits=1):
    if value is None or pd.isna(value):
        return "N/A"
    return f"{float(value) * 100:.{digits}f}%"


def format_text_value(value, default="N/A"):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return default
    text = str(value).strip()
    return text or default


def calculate_efficiency_score(tps, vram_peak_mib):
    if tps is None or pd.isna(tps):
        return None
    if vram_peak_mib is None or pd.isna(vram_peak_mib) or vram_peak_mib <= 0:
        return None
    score = tps / (vram_peak_mib / 1024)
    return round(score, 3)


def calculate_text_tps(char_count, first_text_time, end_time):
    if char_count is None or pd.isna(char_count) or char_count <= 0:
        return None
    if first_text_time is None or end_time is None:
        return None

    generation_time = end_time - first_text_time
    if generation_time <= 0:
        return 0.0
    return round(char_count / generation_time, 2)


def calculate_tps_from_duration(value_count, duration_seconds):
    if value_count is None or pd.isna(value_count) or value_count <= 0:
        return None
    if duration_seconds is None:
        return None
    if duration_seconds <= 0:
        return 0.0
    return round(value_count / duration_seconds, 2)


def calculate_duration_seconds(start_time, end_time):
    if start_time is None or end_time is None:
        return None

    duration_seconds = end_time - start_time
    if duration_seconds <= 0:
        return 0.0
    return round(duration_seconds, 3)


def calculate_text_duration(char_count, first_text_time, end_time):
    if char_count is None or pd.isna(char_count) or char_count <= 0:
        return None
    if first_text_time is None or end_time is None:
        return None

    generation_time = end_time - first_text_time
    if generation_time <= 0:
        return 0.0
    return round(generation_time, 3)


def pick_earliest_time(*timestamps):
    valid_timestamps = [timestamp for timestamp in timestamps if timestamp is not None]
    if not valid_timestamps:
        return None
    return min(valid_timestamps)


def calculate_output_thinking_ratio(output_chars, thinking_chars):
    if thinking_chars is None or pd.isna(thinking_chars) or thinking_chars <= 0:
        return None
    if output_chars is None or pd.isna(output_chars):
        return None
    return round(output_chars / thinking_chars, 3)


def normalize_text_content(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        parts = []
        for item in value:
            if isinstance(item, dict):
                text = item.get("text")
                if text:
                    parts.append(str(text))
            elif item is not None:
                parts.append(str(item))
        return "".join(parts)
    return str(value)


# Model-agnostic token estimate used for relative throughput comparisons.
ESTIMATED_TOKEN_SEGMENT_PATTERN = re.compile(
    r"[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff\u3040-\u30ff\uac00-\ud7af]|"
    r"[A-Za-z0-9]+(?:[._:/+-][A-Za-z0-9]+)*|"
    r"[^\s]",
    re.UNICODE,
)
ASCII_TOKEN_RUN_PATTERN = re.compile(r"^[A-Za-z0-9]+(?:[._:/+-][A-Za-z0-9]+)*$")


def estimate_token_count(value):
    text = normalize_text_content(value)
    if not text:
        return 0

    token_count = 0
    for segment in ESTIMATED_TOKEN_SEGMENT_PATTERN.findall(text):
        if ASCII_TOKEN_RUN_PATTERN.fullmatch(segment):
            token_count += max(1, (len(segment) + 4) // 5)
        else:
            token_count += 1
    return token_count


def extract_delta_payload(delta):
    if delta is None:
        return {}

    if isinstance(delta, dict):
        raw_payload = delta
    elif hasattr(delta, "model_dump"):
        try:
            raw_payload = delta.model_dump(exclude_none=True)
        except TypeError:
            raw_payload = delta.model_dump()
    elif hasattr(delta, "dict"):
        try:
            raw_payload = delta.dict(exclude_none=True)
        except TypeError:
            raw_payload = delta.dict()
    else:
        raw_payload = vars(delta)

    payload = {}
    for key, value in raw_payload.items():
        if value is None:
            continue
        if isinstance(value, str) and value == "":
            continue
        if isinstance(value, (list, dict)) and not value:
            continue
        payload[key] = value

    return payload


def normalize_non_content_type(field_name):
    field_map = {
        "role": "role",
        "tool_calls": "tool_calls",
        "reasoning": "reasoning",
        "refusal": "refusal",
        "audio": "audio",
        "function_call": "tool_calls",
    }
    return field_map.get(field_name, "other")


def inspect_stream_chunk(chunk):
    chunk_info = {
        "content": "",
        "non_content_types": [],
        "finish_reason": None,
    }

    choices = getattr(chunk, "choices", None) or []
    if not choices:
        return chunk_info

    choice = choices[0]
    delta_payload = extract_delta_payload(getattr(choice, "delta", None))
    chunk_info["content"] = normalize_text_content(delta_payload.pop("content", None))
    chunk_info["non_content_types"] = sorted(
        {normalize_non_content_type(field_name) for field_name in delta_payload}
    )
    chunk_info["finish_reason"] = getattr(choice, "finish_reason", None) or None
    return chunk_info


def classify_stream_result(
    chunk_records,
    start_time,
    end_time,
    first_event_time=None,
    first_content_time=None,
    first_thinking_time=None,
    error_message=None,
):
    total_chunks = len(chunk_records)
    content_chunks = 0
    thinking_chunks = 0
    non_content_chunks = 0
    non_content_types = set()
    finish_reason = None
    output_chars = 0
    thinking_chars = 0

    for record in chunk_records:
        if record["content"]:
            content_chunks += 1
            output_chars += len(record["content"])
        if record.get("thinking"):
            thinking_chunks += 1
            thinking_chars += len(record["thinking"])
        if record["non_content_types"]:
            non_content_chunks += 1
            non_content_types.update(record["non_content_types"])
        if record["finish_reason"]:
            finish_reason = record["finish_reason"]

    first_event_seconds = round(first_event_time - start_time, 3) if first_event_time is not None else None
    stream_duration_seconds = round(end_time - start_time, 3)
    thinking_phase_end_time = (
        first_content_time
        if first_thinking_time is not None
        and first_content_time is not None
        and first_content_time >= first_thinking_time
        else end_time
    )
    thinking_time_s = calculate_duration_seconds(first_thinking_time, thinking_phase_end_time)
    output_time_s = calculate_duration_seconds(first_content_time, end_time)
    total_output_time_s = calculate_duration_seconds(
        pick_earliest_time(first_thinking_time, first_content_time, first_event_time),
        end_time,
    )

    if content_chunks > 0 and first_content_time is not None:
        ttft = round(first_content_time - start_time, 3)
        generation_time = end_time - first_content_time
        tps = round(content_chunks / generation_time, 2) if generation_time > 0 else 0.0
    else:
        ttft = None
        tps = None

    substantive_non_content_types = sorted(
        non_content_type for non_content_type in non_content_types if non_content_type != "role"
    )

    if content_chunks > 0:
        output_category = "normal_content"
        if error_message:
            status = "error"
            diagnosis = "Stream interrupted after textual content was received."
        elif not finish_reason:
            status = "warning"
            diagnosis = "Textual content was received, but the stream ended without a terminal finish_reason."
        else:
            status = "ok"
            diagnosis = f"Received textual content and completed with finish_reason={finish_reason}."
    elif error_message:
        status = "error"
        output_category = "early_stop"
        diagnosis = "Stream interrupted before any textual content was received."
    elif finish_reason:
        status = "warning"
        if substantive_non_content_types:
            output_category = "non_content_stream"
            diagnosis = (
                f"Completed with finish_reason={finish_reason} but only non-content payloads "
                f"were received: {', '.join(substantive_non_content_types)}."
            )
        else:
            output_category = "empty_reply"
            diagnosis = f"Completed with finish_reason={finish_reason} but no textual content was received."
    else:
        status = "warning"
        output_category = "early_stop"
        diagnosis = "Stream ended before any textual content or terminal finish_reason was received."

    return {
        "Status": status,
        "Output_Category": output_category,
        "Diagnosis": diagnosis,
        "Finish_Reason": finish_reason,
        "Total_Chunks": total_chunks,
        "Content_Chunks": content_chunks,
        "Thinking_Chunks": thinking_chunks,
        "Non_Content_Chunks": non_content_chunks,
        "Non_Content_Types": ", ".join(sorted(non_content_types)) if non_content_types else "none",
        "First_Event_s": first_event_seconds,
        "Stream_Duration_s": stream_duration_seconds,
        "TTFT": ttft,
        "TPS": tps,
        "Thinking_Chars": thinking_chars,
        "Output_Chars": output_chars,
        "Thinking_Time_s": thinking_time_s,
        "Output_Time_s": output_time_s,
        "Total_Output_Time_s": total_output_time_s,
        "Thinking_TPS": calculate_tps_from_duration(thinking_chars, thinking_time_s),
        "Output_TPS": calculate_tps_from_duration(output_chars, output_time_s),
        "Output_Thinking_Ratio": calculate_output_thinking_ratio(output_chars, thinking_chars),
    }


def build_result_row(
    run_id,
    config,
    model,
    param_set,
    applied_params,
    display_params,
    classification,
    vram_metrics,
    output_text,
    error_message,
):
    efficiency_score = calculate_efficiency_score(classification["TPS"], vram_metrics["VRAM_Peak_MiB"])
    return {
        "Run_ID": run_id,
        "Status": classification["Status"],
        "Capability": config.get("capability", "chat"),
        "Output_Category": classification["Output_Category"],
        "Diagnosis": classification["Diagnosis"],
        "Finish_Reason": classification["Finish_Reason"],
        "Backend": config["backend"],
        "Model": model,
        "Params": param_set.copy(),
        "Applied_Params": applied_params.copy(),
        "Config_Str": display_params,
        "TPS": classification["TPS"],
        "TTFT": classification["TTFT"],
        "First_Event_s": classification["First_Event_s"],
        "Stream_Duration_s": classification["Stream_Duration_s"],
        "Total_Chunks": classification["Total_Chunks"],
        "Content_Chunks": classification["Content_Chunks"],
        "Non_Content_Chunks": classification["Non_Content_Chunks"],
        "Non_Content_Types": classification["Non_Content_Types"],
        "VRAM_Base_MiB": vram_metrics["VRAM_Base_MiB"],
        "VRAM_Peak_MiB": vram_metrics["VRAM_Peak_MiB"],
        "VRAM_Delta_MiB": vram_metrics["VRAM_Delta_MiB"],
        "VRAM_Detail": vram_metrics["VRAM_Detail"],
        "Efficiency_Score": efficiency_score,
        "Output_Chars": len(output_text),
        "Output_Text": output_text,
        "Error": error_message or "",
    }


def filter_eligible_results(df, capability="chat"):
    success_categories = {"tool_call"} if capability == "tools" else {"normal_content"}
    return df[(df["Status"] == "ok") & (df["Output_Category"].isin(success_categories))].copy()


def build_outcome_summary_dataframe(df):
    outcome_df = (
        df["Output_Category"]
        .fillna("unknown")
        .value_counts(dropna=False)
        .rename_axis("Output Category")
        .reset_index(name="Count")
    )
    return outcome_df


def build_tool_call_success_summary_dataframe(df):
    if df.empty or "Model" not in df.columns:
        return pd.DataFrame(
            columns=[
                "Model",
                "Total Runs",
                "Tool Call Success Count",
                "Tool Call Success Probability",
            ]
        )

    summary_rows = []
    for model, group in df.groupby("Model", dropna=False, sort=True):
        total_runs = int(len(group))
        success_count = int(
            ((group["Status"] == "ok") & (group["Output_Category"] == "tool_call")).sum()
        )
        success_probability = success_count / total_runs if total_runs else None
        summary_rows.append(
            {
                "Model": model,
                "Total Runs": total_runs,
                "Tool Call Success Count": success_count,
                "Tool Call Success Probability": format_probability_value(success_probability),
            }
        )

    return pd.DataFrame(summary_rows)


def wrap_markdown_table_headers(df):
    header_map = {
        "System Prompt": "System Prompt<br>Variant",
        "Thinking Mode": "Thinking Mode<br>State",
        "Output Category": "Output<br>Category",
        "Finish Reason": "Finish<br>Reason",
        "Thinking Time (s)": "Thinking Time<br>(s)",
        "Output Time (s)": "Output Time<br>(s)",
        "Total Output (chars)": "Total Output<br>(chars)",
        "Total Output Time (s)": "Total Output Time<br>(s)",
        "TPS (chunk/s)": "TPS<br>(chunk/s)",
        "Thinking TPS (token/s)": "Thinking TPS<br>(token/s)",
        "Output TPS (token/s)": "Output TPS<br>(token/s)",
        "Output/Thinking Ratio": "Output/Thinking<br>Ratio",
        "TTFT (s)": "TTFT<br>(s)",
        "First Event (s)": "First Event<br>(s)",
        "VRAM Peak (MiB)": "VRAM Peak<br>(MiB)",
        "Efficiency Score (TPS/GiB Peak)": "Efficiency Score<br>(TPS/GiB Peak)",
        "Chunks (content/total)": "Chunks<br>(content/total)",
        "Total Runs": "Total<br>Runs",
        "Tool Call Success Count": "Tool Call Success<br>Count",
        "Tool Call Success Probability": "Tool Call Success<br>Probability",
    }
    return df.rename(columns=header_map)


def summarize_vram_samples(samples):
    if not samples:
        return empty_vram_metrics()

    baseline_snapshot = samples[0]
    total_base_mib = sum(item["memory_used_mib"] for item in baseline_snapshot)
    total_peak_mib = max(sum(item["memory_used_mib"] for item in snapshot) for snapshot in samples)

    per_gpu = {}
    for item in baseline_snapshot:
        per_gpu[item["index"]] = {
            "index": item["index"],
            "name": item["name"],
            "memory_total_mib": item["memory_total_mib"],
            "base_mib": item["memory_used_mib"],
            "peak_mib": item["memory_used_mib"],
        }

    for snapshot in samples:
        for item in snapshot:
            state = per_gpu.setdefault(
                item["index"],
                {
                    "index": item["index"],
                    "name": item["name"],
                    "memory_total_mib": item["memory_total_mib"],
                    "base_mib": item["memory_used_mib"],
                    "peak_mib": item["memory_used_mib"],
                },
            )
            state["peak_mib"] = max(state["peak_mib"], item["memory_used_mib"])
            state["memory_total_mib"] = item["memory_total_mib"]
            state["name"] = item["name"]

    detail_parts = []
    for gpu_index in sorted(per_gpu):
        gpu = per_gpu[gpu_index]
        delta_mib = gpu["peak_mib"] - gpu["base_mib"]
        detail_parts.append(
            (
                f"GPU {gpu['index']} {gpu['name']}: "
                f"{gpu['base_mib']} -> {gpu['peak_mib']} / {gpu['memory_total_mib']} MiB "
                f"(+{delta_mib} MiB)"
            )
        )

    return {
        "VRAM_Base_MiB": total_base_mib,
        "VRAM_Peak_MiB": total_peak_mib,
        "VRAM_Delta_MiB": total_peak_mib - total_base_mib,
        "VRAM_Detail": " | ".join(detail_parts) if detail_parts else "N/A",
    }


class NvidiaVRAMMonitor:
    def __init__(self, interval_seconds=0.2):
        self.interval_seconds = interval_seconds
        self.samples = []
        self._stop_event = threading.Event()
        self._thread = None
        self.enabled = False

    def start(self):
        baseline_snapshot = query_nvidia_vram_snapshot()
        if not baseline_snapshot:
            return False

        self.samples = [baseline_snapshot]
        self.enabled = True
        self._thread = threading.Thread(target=self._poll_loop, daemon=True)
        self._thread.start()
        return True

    def _poll_loop(self):
        while not self._stop_event.wait(self.interval_seconds):
            snapshot = query_nvidia_vram_snapshot()
            if snapshot:
                self.samples.append(snapshot)

    def stop(self):
        if not self.enabled:
            return empty_vram_metrics()

        self._stop_event.set()
        if self._thread is not None:
            self._thread.join(timeout=1)

        final_snapshot = query_nvidia_vram_snapshot()
        if final_snapshot:
            self.samples.append(final_snapshot)

        return summarize_vram_samples(self.samples)


def build_summary_dataframe(df):
    summary_source = df.copy()
    for column_name in (
        "Output_Chars",
        "Thinking_Time_s",
        "Output_Time_s",
        "Total_Output_Time_s",
        "Thinking_TPS",
        "Output_TPS",
        "Output_Thinking_Ratio",
    ):
        if column_name not in summary_source.columns:
            summary_source[column_name] = None

    summary_columns = [
        "Run_ID",
        "Status",
        "Output_Category",
        "Model",
        "Finish_Reason",
        "Config_Str",
        "Output_Chars",
        "Thinking_Time_s",
        "Output_Time_s",
        "Total_Output_Time_s",
        "TPS",
        "Thinking_TPS",
        "Output_TPS",
        "Output_Thinking_Ratio",
        "TTFT",
        "First_Event_s",
        "Content_Chunks",
        "Total_Chunks",
        "VRAM_Peak_MiB",
        "Efficiency_Score",
    ]
    if "Capability" in df.columns:
        summary_columns.insert(2, "Capability")
    if "System_Prompt_Label" in df.columns:
        summary_columns.insert(summary_columns.index("Model") + 1, "System_Prompt_Label")
    if "Thinking_Mode" in df.columns:
        thinking_mode_insert_at = summary_columns.index("Model") + 1
        if "System_Prompt_Label" in summary_columns:
            thinking_mode_insert_at = summary_columns.index("System_Prompt_Label") + 1
        summary_columns.insert(thinking_mode_insert_at, "Thinking_Mode")

    summary_df = summary_source[summary_columns].copy()
    summary_df["Finish_Reason"] = summary_df["Finish_Reason"].apply(format_text_value)
    if "Thinking_Mode" in summary_df.columns:
        summary_df["Thinking_Mode"] = summary_df["Thinking_Mode"].apply(resolve_thinking_mode)
    summary_df["TPS"] = summary_df["TPS"].apply(lambda value: format_numeric_value(value, 2))
    summary_df["Thinking_TPS"] = summary_df["Thinking_TPS"].apply(
        lambda value: format_numeric_value(value, 2)
    )
    summary_df["Output_TPS"] = summary_df["Output_TPS"].apply(
        lambda value: format_numeric_value(value, 2)
    )
    summary_df["Output_Thinking_Ratio"] = summary_df["Output_Thinking_Ratio"].apply(
        lambda value: format_numeric_value(value, 3)
    )
    summary_df["Output_Chars"] = summary_df["Output_Chars"].apply(
        lambda value: "N/A" if pd.isna(value) else int(value)
    )
    summary_df["Thinking_Time_s"] = summary_df["Thinking_Time_s"].apply(
        lambda value: format_numeric_value(value, 3)
    )
    summary_df["Output_Time_s"] = summary_df["Output_Time_s"].apply(
        lambda value: format_numeric_value(value, 3)
    )
    summary_df["Total_Output_Time_s"] = summary_df["Total_Output_Time_s"].apply(
        lambda value: format_numeric_value(value, 3)
    )
    summary_df["TTFT"] = summary_df["TTFT"].apply(lambda value: format_numeric_value(value, 3))
    summary_df["First_Event_s"] = summary_df["First_Event_s"].apply(
        lambda value: format_numeric_value(value, 3)
    )
    summary_df["VRAM_Peak_MiB"] = summary_df["VRAM_Peak_MiB"].apply(
        lambda value: "N/A" if pd.isna(value) else int(value)
    )
    summary_df["Efficiency_Score"] = summary_df["Efficiency_Score"].apply(
        lambda value: format_numeric_value(value, 3)
    )
    summary_df["Chunks (content/total)"] = summary_df.apply(
        lambda row: f"{int(row['Content_Chunks'])}/{int(row['Total_Chunks'])}",
        axis=1,
    )
    summary_df = summary_df.drop(columns=["Content_Chunks", "Total_Chunks"])
    return summary_df.rename(
        columns={
            "Run_ID": "Run",
            "Capability": "Capability",
            "Output_Category": "Output Category",
            "System_Prompt_Label": "System Prompt",
            "Thinking_Mode": "Thinking Mode",
            "Finish_Reason": "Finish Reason",
            "Config_Str": "Config",
            "Output_Chars": "Total Output (chars)",
            "Thinking_Time_s": "Thinking Time (s)",
            "Output_Time_s": "Output Time (s)",
            "Total_Output_Time_s": "Total Output Time (s)",
            "TPS": "TPS (chunk/s)",
            "Thinking_TPS": "Thinking TPS (token/s)",
            "Output_TPS": "Output TPS (token/s)",
            "Output_Thinking_Ratio": "Output/Thinking Ratio",
            "TTFT": "TTFT (s)",
            "First_Event_s": "First Event (s)",
            "VRAM_Peak_MiB": "VRAM Peak (MiB)",
            "Efficiency_Score": "Efficiency Score (TPS/GiB Peak)",
        }
    )


def dataframe_to_text_table(df):
    try:
        return df.to_markdown(index=False)
    except Exception:
        return df.to_string(index=False)


def dataframe_to_report_table(df):
    try:
        if any("<br>" in str(column_name) for column_name in df.columns):
            return df.to_html(index=False, escape=False, border=0)
        return df.to_markdown(index=False)
    except Exception:
        return df.to_string(index=False)


def html_escape_text(value):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except TypeError:
        pass
    if isinstance(value, (dict, list)):
        value = json.dumps(value, ensure_ascii=False)
    return html.escape(str(value))


def html_escape_header(value):
    return html.escape(str(value)).replace("&lt;br&gt;", "<br>")


def dataframe_to_html_table(df, table_class="report-table", empty_message="No data"):
    columns = [str(column_name) for column_name in getattr(df, "columns", [])]
    parts = [f'<table class="{table_class}">', "<thead><tr>"]

    if columns:
        for column_name in columns:
            parts.append(f"<th>{html_escape_header(column_name)}</th>")
    else:
        parts.append("<th>Value</th>")

    parts.append("</tr></thead><tbody>")

    if df is None or df.empty:
        parts.append(
            f'<tr><td class="empty-cell" colspan="{max(len(columns), 1)}">{html_escape_text(empty_message)}</td></tr>'
        )
    else:
        for _, row in df.iterrows():
            parts.append("<tr>")
            for column_name in columns:
                parts.append(f"<td>{html_escape_text(row[column_name])}</td>")
            parts.append("</tr>")

    parts.append("</tbody></table>")
    return "".join(parts)


def make_excel_friendly_dataframe(df):
    excel_df = df.copy()
    excel_df.columns = [re.sub(r"<br\s*/?>", "\n", str(column)) for column in excel_df.columns]
    return excel_df


def style_excel_worksheet(worksheet):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor="F6EFE2")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = header_fill

    for column_cells in worksheet.columns:
        lengths = []
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            lines = value.splitlines() or [value]
            lengths.append(max(len(line) for line in lines))
            if cell.row > 1:
                cell.alignment = Alignment(vertical="top")
        column_width = min(max(lengths) + 2 if lengths else 12, 42)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = max(column_width, 12)


def save_summary_excel_workbook(df, config, report_stem):
    capability = config.get("capability", "chat")
    workbook_path = Path(f"{report_stem}_summary.xlsx")
    summary_df = make_excel_friendly_dataframe(localize_report_dataframe(build_summary_dataframe(df)))
    outcome_summary_df = make_excel_friendly_dataframe(
        localize_report_dataframe(build_outcome_summary_dataframe(df))
    )
    tool_call_success_summary_df = (
        make_excel_friendly_dataframe(
            localize_report_dataframe(build_tool_call_success_summary_dataframe(df))
        )
        if capability == "tools"
        else pd.DataFrame()
    )

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        outcome_summary_df.to_excel(writer, sheet_name="Outcome Summary", index=False)
        if capability == "tools" and not tool_call_success_summary_df.empty:
            tool_call_success_summary_df.to_excel(writer, sheet_name="Tool Call Success", index=False)

        for worksheet in writer.book.worksheets:
            style_excel_worksheet(worksheet)

    return workbook_path


def build_download_button(href, label):
    if not href:
        return ""
    return (
        f'<a class="download-button" href="{html_escape_text(href)}" download>'
        f"{html_escape_text(label)}</a>"
    )


def build_run_filter_panel_html():
    filter_groups = [
        (
            bilingual_text("Tool Call", "工具呼叫"),
            "tool-call",
            [
                ("yes", bilingual_text("Has tool call", "有呼叫 tool")),
                ("no", bilingual_text("No tool call", "沒有呼叫 tool")),
            ],
        ),
        (
            bilingual_text("Thinking", "Thinking 內容"),
            "thinking",
            [
                ("yes", bilingual_text("Has thinking", "有 thinking")),
                ("no", bilingual_text("No thinking", "沒有 thinking")),
            ],
        ),
        (
            bilingual_text("Output", "輸出內容"),
            "output",
            [
                ("yes", bilingual_text("Has output", "有 output")),
                ("no", bilingual_text("No output", "沒有 output")),
            ],
        ),
        (
            bilingual_text("Thinking Mode", "思考模式"),
            "thinking-mode",
            [
                ("enable", bilingual_text("Enable", "啟用")),
                ("disable", bilingual_text("Disable", "停用")),
                ("default", bilingual_text("Default", "預設")),
            ],
        ),
        (
            bilingual_text("Status", "狀態"),
            "status",
            [
                ("ok", bilingual_text("ok", "正常")),
                ("warning", bilingual_text("warning", "警告")),
                ("error", bilingual_text("error", "錯誤")),
            ],
        ),
    ]

    parts = [
        '<div class="filter-panel" id="run-filter-panel">',
        '<div class="filter-toolbar">',
        '<p class="filter-summary" id="run-filter-summary"></p>',
        (
            '<button type="button" class="filter-reset-button" id="run-filter-reset">'
            'Reset Filters / 重設篩選</button>'
        ),
        "</div>",
        '<div class="filter-groups">',
    ]

    for legend, group_name, options in filter_groups:
        parts.append('<fieldset class="filter-group">')
        parts.append(f"<legend>{legend}</legend>")
        for value, label in options:
            parts.append(
                (
                    '<label class="filter-option">'
                    f'<input type="checkbox" class="run-filter-input" data-filter-group="{group_name}" '
                    f'value="{value}" checked> '
                    f"{label}</label>"
                )
            )
        parts.append("</fieldset>")

    parts.extend(
        [
            "</div>",
            (
                '<p class="empty-note" id="run-filter-empty" hidden>'
                'No runs match the current filters. / 目前篩選條件下沒有符合的結果。</p>'
            ),
            "</div>",
        ]
    )
    return "".join(parts)


def build_run_filter_script():
    return """<script>
(() => {
  const grid = document.getElementById('run-card-grid');
  if (!grid) {
    return;
  }

  const cards = Array.from(grid.querySelectorAll('.result-card'));
  const inputs = Array.from(document.querySelectorAll('.run-filter-input'));
  const summary = document.getElementById('run-filter-summary');
  const emptyNote = document.getElementById('run-filter-empty');
  const resetButton = document.getElementById('run-filter-reset');
  const groups = ['tool-call', 'thinking', 'output', 'thinking-mode', 'status'];

  function selectedValues(groupName) {
    return new Set(
      inputs
        .filter((input) => input.dataset.filterGroup === groupName && input.checked)
        .map((input) => input.value)
    );
  }

  function applyFilters() {
    const selectedByGroup = Object.fromEntries(
      groups.map((groupName) => [groupName, selectedValues(groupName)])
    );
    let visibleCount = 0;

    for (const card of cards) {
      const matches = groups.every((groupName) => {
        const selected = selectedByGroup[groupName];
        if (selected.size === 0) {
          return false;
        }
        return selected.has(card.getAttribute(`data-${groupName}`));
      });

      card.hidden = !matches;
      if (matches) {
        visibleCount += 1;
      }
    }

    if (summary) {
      summary.textContent = `${visibleCount} / ${cards.length} runs shown / 顯示 ${visibleCount} / ${cards.length} 筆`;
    }
    if (emptyNote) {
      emptyNote.hidden = visibleCount !== 0;
    }
  }

  for (const input of inputs) {
    input.addEventListener('change', applyFilters);
  }

  if (resetButton) {
    resetButton.addEventListener('click', () => {
      for (const input of inputs) {
        input.checked = true;
      }
      applyFilters();
    });
  }

  applyFilters();
})();
</script>"""


def key_value_rows_to_html_table(rows, table_class="kv-table"):
    frame = pd.DataFrame(rows, columns=["Field", "Value"])
    return dataframe_to_html_table(frame, table_class=table_class)


def bullet_list_to_html(items, list_class="note-list"):
    parts = [f'<ul class="{list_class}">']
    for item in items:
        parts.append(f"<li>{html_escape_text(item)}</li>")
    parts.append("</ul>")
    return "".join(parts)


def bilingual_text(english, chinese):
    return f"{english} / {chinese}"


STATUS_BILINGUAL_MAP = {
    "ok": bilingual_text("ok", "正常"),
    "warning": bilingual_text("warning", "警告"),
    "error": bilingual_text("error", "錯誤"),
}

CAPABILITY_BILINGUAL_MAP = {
    "chat": bilingual_text("chat", "對話"),
    "tools": bilingual_text("tools", "工具呼叫"),
}

OUTPUT_CATEGORY_BILINGUAL_MAP = {
    "normal_content": bilingual_text("normal_content", "正常文字輸出"),
    "tool_call": bilingual_text("tool_call", "工具呼叫"),
    "text_reply_without_tool": bilingual_text("text_reply_without_tool", "有文字回覆但未呼叫工具"),
    "empty_reply": bilingual_text("empty_reply", "空回覆"),
    "non_content_stream": bilingual_text("non_content_stream", "僅非文字串流"),
    "early_stop": bilingual_text("early_stop", "提早中斷"),
}

FINISH_REASON_BILINGUAL_MAP = {
    "stop": bilingual_text("stop", "正常結束"),
    "tool_calls": bilingual_text("tool_calls", "工具呼叫結束"),
    "length": bilingual_text("length", "長度上限"),
}

THINKING_MODE_BILINGUAL_MAP = {
    "enable": bilingual_text("enable", "啟用"),
    "disable": bilingual_text("disable", "停用"),
    "default": bilingual_text("default", "依後端預設"),
}

REPORT_HEADER_BILINGUAL_MAP = {
    "Run": "Run<br>執行編號",
    "Status": "Status<br>狀態",
    "Capability": "Capability<br>能力模式",
    "Output Category": "Output Category<br>輸出分類",
    "Model": "Model<br>模型",
    "System Prompt": "System Prompt<br>系統提示",
    "Thinking Mode": "Thinking Mode<br>思考模式",
    "Finish Reason": "Finish Reason<br>結束原因",
    "Config": "Config<br>測試設定",
    "Thinking Time (s)": "Thinking Time<br>(s)<br>思考時間",
    "Output Time (s)": "Output Time<br>(s)<br>回覆時間",
    "Total Output (chars)": "Total Output<br>(chars)<br>總輸出字數",
    "Total Output Time (s)": "Total Output Time<br>(s)<br>總輸出時間",
    "TPS (chunk/s)": "TPS<br>(chunk/s)<br>輸出速率",
    "Thinking TPS (token/s)": "Thinking TPS<br>(token/s)<br>思考速率",
    "Output TPS (token/s)": "Output TPS<br>(token/s)<br>回覆速率",
    "Output/Thinking Ratio": "Output/Thinking<br>Ratio<br>輸出思考比",
    "TTFT (s)": "TTFT<br>(s)<br>首字延遲",
    "First Event (s)": "First Event<br>(s)<br>首事件延遲",
    "VRAM Peak (MiB)": "VRAM Peak<br>(MiB)<br>顯存峰值",
    "Efficiency Score (TPS/GiB Peak)": "Efficiency Score<br>(TPS/GiB Peak)<br>效率分數",
    "Chunks (content/total)": "Chunks<br>(content/total)<br>片段數",
    "Count": "Count<br>次數",
    "Total Runs": "Total Runs<br>總測試次數",
    "Tool Call Success Count": "Tool Call Success<br>Count<br>工具成功次數",
    "Tool Call Success Probability": "Tool Call Success<br>Probability<br>工具成功率",
}


def localize_status_value(value):
    return STATUS_BILINGUAL_MAP.get(value, value)


def localize_capability_value(value):
    return CAPABILITY_BILINGUAL_MAP.get(value, value)


def localize_output_category_value(value):
    return OUTPUT_CATEGORY_BILINGUAL_MAP.get(value, value)


def localize_finish_reason_value(value):
    if value in (None, "", "N/A"):
        return bilingual_text("N/A", "無")
    return FINISH_REASON_BILINGUAL_MAP.get(value, value)


def localize_system_prompt_label(value):
    if value in (None, "", "N/A"):
        return bilingual_text("N/A", "未使用額外 system prompt")
    return bilingual_text(str(value), "系統提示變體")


def localize_thinking_mode_value(value):
    normalized = resolve_thinking_mode(value)
    return THINKING_MODE_BILINGUAL_MAP.get(normalized, value)


def localize_report_dataframe(df):
    localized_df = df.copy()
    if "Status" in localized_df.columns:
        localized_df["Status"] = localized_df["Status"].apply(localize_status_value)
    if "Capability" in localized_df.columns:
        localized_df["Capability"] = localized_df["Capability"].apply(localize_capability_value)
    if "Output Category" in localized_df.columns:
        localized_df["Output Category"] = localized_df["Output Category"].apply(
            localize_output_category_value
        )
    if "Finish Reason" in localized_df.columns:
        localized_df["Finish Reason"] = localized_df["Finish Reason"].apply(
            localize_finish_reason_value
        )
    if "System Prompt" in localized_df.columns:
        localized_df["System Prompt"] = localized_df["System Prompt"].apply(
            localize_system_prompt_label
        )
    if "Thinking Mode" in localized_df.columns:
        localized_df["Thinking Mode"] = localized_df["Thinking Mode"].apply(
            localize_thinking_mode_value
        )
    return localized_df.rename(columns=REPORT_HEADER_BILINGUAL_MAP)


def serialize_result_value(value):
    if isinstance(value, (dict, list)):
        return value
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass
    if hasattr(value, "item") and callable(value.item):
        try:
            return value.item()
        except (TypeError, ValueError):
            pass
    return value


def save_raw_outputs(df, report_stem):
    output_path = Path(f"{report_stem}_outputs.jsonl")
    with output_path.open("w", encoding="utf-8") as file:
        for _, row in df.iterrows():
            payload = {column: serialize_result_value(value) for column, value in row.items()}
            file.write(json.dumps(payload, ensure_ascii=False) + "\n")
    return output_path


def ensure_report_output_dir(base_dir="."):
    report_dir = Path(base_dir) / "Report"
    report_dir.mkdir(parents=True, exist_ok=True)
    return report_dir


def select_plot_dataframe(df, capability="chat"):
    if capability == "tools":
        return df.copy(), False

    eligible_df = filter_eligible_results(df, capability=capability)
    if eligible_df.empty:
        return df.copy(), True
    return eligible_df.copy(), False


def normalize_output_text(text):
    normalized = (text or "").replace("\r\n", "\n").strip()
    return normalized or "[No text returned]"


def pause_before_exit():
    if not sys.stdin.isatty():
        return

    try:
        if msvcrt is not None:
            print("\nPress any key to exit... / 按任意鍵結束...", end="", flush=True)
            msvcrt.getch()
            print()
        else:
            input("\nPress Enter to exit... / 按 Enter 結束...")
    except (EOFError, KeyboardInterrupt):
        pass


def get_installed_version(package_name):
    try:
        return get_package_version(package_name)
    except PackageNotFoundError:
        return None


def show_windows_error_dialog(title, message):
    if sys.platform != "win32":
        return False

    try:
        import ctypes

        ctypes.windll.user32.MessageBoxW(None, message, title, 0x10)
        return True
    except Exception:
        return False


def print_warning_box(message, detail=None):
    print("\n" + "=" * 62)
    print(f"⚠️ Warning / 警告\n{message}")
    if detail:
        print(f"Detail / 詳細資訊: {detail}")
    print("=" * 62)


def ask_backend_retry_or_back(message):
    return ask_select_with_back(
        message,
        choices=[
            Choice("Retry check / 重新檢查", value="retry"),
            Choice("Go back / 返回上一階段", value="back"),
        ],
        default="retry",
    )


def ensure_runtime_ready():
    if not DEPENDENCY_IMPORT_ERRORS:
        return

    guidance = {
        "openai": "請執行 `pip install -U openai`，並確認版本為 1.x 以上。",
        "questionary": "請執行 `pip install -U questionary prompt_toolkit`。",
        "matplotlib": "請執行 `pip install -U matplotlib`。",
        "pandas": "請執行 `pip install -U pandas`。",
        "requests": "請執行 `pip install -U requests`。",
    }
    lines = [
        "程式啟動前檢查失敗。",
        "這通常代表另一台電腦雖然有安裝套件，但版本不相容或安裝不完整。",
        "",
    ]
    for package_name, exc in DEPENDENCY_IMPORT_ERRORS.items():
        installed_version = get_installed_version(package_name)
        version_label = f"已安裝 {installed_version}" if installed_version else "未偵測到安裝版本"
        hint = guidance.get(package_name, f"請重新安裝 `{package_name}`。")
        lines.append(f"- {package_name} ({version_label}): {exc}. {hint}")

    raise RuntimeError("\n".join(lines))


def persist_crash_log(error_text):
    crash_log_path = Path("expert_LLM_benchmark_crash.log")
    crash_log_path.write_text(error_text, encoding="utf-8")
    return crash_log_path


def get_project_python_command_hint():
    if sys.platform == "win32":
        return r".\.venv\Scripts\python.exe expert_LLM_benchmark.py"

    return "./.venv/bin/python expert_LLM_benchmark.py"


def handle_fatal_error(exc):
    traceback_text = traceback.format_exc()
    if traceback_text.strip() == "NoneType: None":
        traceback_text = f"{type(exc).__name__}: {exc}"

    error_text = "\n".join(
        [
            "程式執行失敗。",
            f"建議在專案目錄用 `.venv` 內的 Python 執行，例如 `{get_project_python_command_hint()}`，比較容易看到完整訊息。",
            "",
            f"{type(exc).__name__}: {exc}",
            "",
            traceback_text,
        ]
    )
    crash_log_path = persist_crash_log(error_text)

    print(error_text, file=sys.stderr)
    print(f"\n錯誤記錄已寫入: {crash_log_path.resolve()}", file=sys.stderr)

    if not sys.stdin.isatty():
        dialog_message = "\n".join(
            [
                "程式執行失敗，錯誤記錄已保存。",
                str(crash_log_path.resolve()),
                "",
                f"{type(exc).__name__}: {exc}",
                "",
                "請改用 PowerShell / CMD 執行，或把 crash log 傳回來。",
            ]
        )
        show_windows_error_dialog("expert_LLM_benchmark 啟動失敗", dialog_message)


def select_models_and_url(backend, previous_url=None, previous_models=None):
    previous_models = previous_models or []

    if backend == "ollama":
        while True:
            backend_status = inspect_backend_readiness("ollama", OLLAMA_BASE_URL)
            if not backend_status.get("ok"):
                print_warning_box(
                    backend_status["message"],
                    detail=backend_status.get("detail"),
                )
                retry_action = ask_backend_retry_or_back(
                    "Ollama backend check failed. Retry after fixing it, or go back? / "
                    "Ollama 後端檢查失敗。修正後要重試，還是返回上一階段？"
                )
                if retry_action is None:
                    return None
                if retry_action in (BACK_ACTION, "back"):
                    return BACK_ACTION
                continue

            detected_models = backend_status.get("models", [])
            if backend_status.get("warning"):
                print_warning_box(backend_status["warning"])

            if detected_models:
                choice_list = [
                    Choice(model_name, value=model_name, checked=model_name in previous_models)
                    for model_name in detected_models
                ]
                selected_models = ask_checkbox_with_back(
                    "Select benchmark models / 選擇測試模型:",
                    choices=choice_list,
                )
                if selected_models is None:
                    return None
                if selected_models == BACK_ACTION:
                    return BACK_ACTION
                if selected_models:
                    return OLLAMA_BASE_URL, selected_models

            manual_input = ask_text_with_back(
                "Enter Ollama model names (comma separated) / 請輸入 Ollama 模型名稱（逗號分隔）:",
                default=",".join(previous_models) if previous_models else "qwen3.5:latest",
            )
            if manual_input is None:
                return None
            if manual_input == BACK_ACTION:
                return BACK_ACTION
            models = [name.strip() for name in (manual_input or "").split(",") if name.strip()]
            if models:
                return OLLAMA_BASE_URL, models
            print("Please enter at least one model name. / 請至少輸入一個模型名稱。")

    port_default = DEFAULT_LLAMA_PORT
    if previous_url and previous_url.startswith("http://localhost:") and previous_url.endswith("/v1"):
        port_default = previous_url.removeprefix("http://localhost:").removesuffix("/v1") or DEFAULT_LLAMA_PORT

    while True:
        raw_port = ask_text_with_back(
            "Enter llama-server port / 請輸入 llama-server 端口:",
            default=port_default,
        )
        if raw_port is None:
            return None
        if raw_port == BACK_ACTION:
            return BACK_ACTION

        port, port_error = normalize_local_port(raw_port, default=port_default)
        if port_error:
            print_warning_box(port_error)
            continue

        url = f"http://localhost:{port}/v1"
        backend_status = inspect_backend_readiness("llama.cpp", url)
        if not backend_status.get("ok"):
            print_warning_box(
                backend_status["message"],
                detail=backend_status.get("detail"),
            )
            retry_action = ask_backend_retry_or_back(
                "llama.cpp backend check failed. Retry after fixing it, or go back? / "
                "llama.cpp 後端檢查失敗。修正後要重試，還是返回上一階段？"
            )
            if retry_action is None:
                return None
            if retry_action in (BACK_ACTION, "back"):
                return BACK_ACTION
            port_default = port
            continue

        if backend_status.get("warning"):
            print_warning_box(backend_status["warning"])

        detected_models = backend_status.get("models", [])
        suggested_models = ",".join(previous_models) if previous_models else ",".join(detected_models)
        if not suggested_models:
            suggested_models = "llama.cpp-model"

        model_names = ask_text_with_back(
            "Enter loaded model names (comma separated, for labeling only) / "
            "請輸入載入中的模型名稱（逗號分隔，僅供辨識）:",
            default=suggested_models,
        )
        if model_names is None:
            return None
        if model_names == BACK_ACTION:
            continue

        models = [name.strip() for name in (model_names or "").split(",") if name.strip()]
        if models:
            return url, models
        print("Please enter at least one model name. / 請至少輸入一個模型名稱。")


def interactive_config():
    print("\n" + "═" * 62)
    print("🏆 LLM 專家參數 Benchmark 工具 V3")
    print("═" * 62)

    backend = questionary.select(
        "請選擇測試後端:",
        choices=[
            Choice("🦙 Ollama", value="ollama"),
            Choice("🏗️ llama.cpp (需先啟動 llama-server)", value="llama.cpp"),
        ],
    ).ask()
    if not backend:
        return None

    url, models = select_models_and_url(backend)
    if not models:
        print("⚠️ 沒有可用模型，已取消。")
        return None

    available_groups = {
        group_name: [key for key in param_keys if backend in PARAM_INFO[key]["backends"]]
        for group_name, param_keys in PARAM_GROUPS.items()
    }
    available_groups = {name: keys for name, keys in available_groups.items() if keys}

    selected_groups = questionary.checkbox(
        "選擇想測試的參數類別 (可不選，代表只比模型預設值):",
        choices=[
            Choice(f"{group_name} ({len(param_keys)} 項)", value=group_name)
            for group_name, param_keys in available_groups.items()
        ],
    ).ask() or []

    final_params = {}
    for group_name in selected_groups:
        param_keys = available_groups[group_name]
        selected_params = questionary.checkbox(
            f"勾選 {group_name} 要測試的參數:",
            choices=[
                Choice(
                    title=(
                        f"{PARAM_INFO[key]['label']} | 範圍: {PARAM_INFO[key]['range']} | "
                        f"{PARAM_INFO[key]['desc']}"
                    ),
                    value=key,
                )
                for key in param_keys
            ],
        ).ask() or []

        for key in selected_params:
            values = ask_param_values(key)
            if values is None:
                return None
            final_params[key] = values

    prompt = questionary.text(
        "測試 Prompt:",
        default="詳細解釋 3D 列印中，PETG 材質發生蠕變 (Creep) 的溫度臨界點。",
    ).ask()
    if prompt is None:
        return None

    return {
        "backend": backend,
        "url": url,
        "models": models,
        "params": final_params,
        "prompt": prompt,
    }


def run_bench(config):
    client = OpenAI(base_url=config["url"], api_key="sk-no-key-needed")
    param_keys = list(config["params"].keys())
    param_values = [config["params"][key] for key in param_keys]
    combos = [dict(zip(param_keys, combo)) for combo in product(*param_values)] if param_keys else [{}]

    results = []
    total_runs = len(config["models"]) * len(combos)
    vram_monitoring_enabled = query_nvidia_vram_snapshot() is not None
    config["vram_monitoring"] = "nvidia-smi" if vram_monitoring_enabled else "unavailable"
    print(f"\n⚡ 啟動測試，共 {total_runs} 組配置。")
    print("📌 TPS 以串流回傳片段估算，適合做相對比較。")
    if vram_monitoring_enabled:
        print("🧠 顯存監控: 已啟用 nvidia-smi 取樣。")
    else:
        print("🧠 顯存監控: 未偵測到 nvidia-smi，報告將顯示 N/A。")

    run_index = 0
    for model in config["models"]:
        for param_set in combos:
            run_index += 1
            applied_params = build_backend_options(config["backend"], param_set)
            display_params = format_param_dict(param_set)
            print(f"[{run_index}/{total_runs}] {model} | {display_params}")

            request_kwargs = {}
            if applied_params:
                request_kwargs["extra_body"] = (
                    {"options": applied_params}
                    if config["backend"] == "ollama"
                    else applied_params
                )

            start_time = time.time()
            first_event_time = None
            first_content_time = None
            output_parts = []
            chunk_records = []
            vram_monitor = NvidiaVRAMMonitor() if vram_monitoring_enabled else None
            vram_metrics = empty_vram_metrics()
            if vram_monitor is not None and not vram_monitor.start():
                vram_monitor = None
            try:
                stream = client.chat.completions.create(
                    model=model,
                    messages=[{"role": "user", "content": config["prompt"]}],
                    stream=True,
                    **request_kwargs,
                )

                for chunk in stream:
                    event_time = time.time()
                    if first_event_time is None:
                        first_event_time = event_time

                    chunk_info = inspect_stream_chunk(chunk)
                    chunk_records.append(chunk_info)

                    if chunk_info["content"] and first_content_time is None:
                        first_content_time = event_time
                    if chunk_info["content"]:
                        output_parts.append(chunk_info["content"])

                end_time = time.time()
                if vram_monitor is not None:
                    vram_metrics = vram_monitor.stop()
                output_text = "".join(output_parts)
                classification = classify_stream_result(
                    chunk_records=chunk_records,
                    start_time=start_time,
                    end_time=end_time,
                    first_event_time=first_event_time,
                    first_content_time=first_content_time,
                    error_message=None,
                )
                results.append(
                    build_result_row(
                        run_id=run_index,
                        config=config,
                        model=model,
                        param_set=param_set,
                        applied_params=applied_params,
                        display_params=display_params,
                        classification=classification,
                        vram_metrics=vram_metrics,
                        output_text=output_text,
                        error_message=None,
                    )
                )
            except Exception as exc:
                end_time = time.time()
                if vram_monitor is not None:
                    vram_metrics = vram_monitor.stop()
                output_text = "".join(output_parts)
                classification = classify_stream_result(
                    chunk_records=chunk_records,
                    start_time=start_time,
                    end_time=end_time,
                    first_event_time=first_event_time,
                    first_content_time=first_content_time,
                    error_message=str(exc),
                )
                print(f"❌ 失敗: {exc}")
                results.append(
                    build_result_row(
                        run_id=run_index,
                        config=config,
                        model=model,
                        param_set=param_set,
                        applied_params=applied_params,
                        display_params=display_params,
                        classification=classification,
                        vram_metrics=vram_metrics,
                        output_text=output_text,
                        error_message=str(exc),
                    )
                )

    return pd.DataFrame(results)


def plot_results(df, output_path, capability="chat"):
    if df.empty:
        return None

    plot_df, used_fallback = select_plot_dataframe(df, capability=capability)
    plot_df = plot_df.copy()
    multi_prompt_mode = (
        "System_Prompt_Label" in plot_df.columns
        and plot_df["System_Prompt_Label"].fillna("N/A").nunique(dropna=False) > 1
    )

    def build_label(row):
        system_prompt_label = row.get("System_Prompt_Label", "N/A")
        if multi_prompt_mode or (system_prompt_label not in ("", "N/A", None)):
            label = f"{row['Model']} | {system_prompt_label} | {row['Config_Str']}"
        else:
            label = f"{row['Model']} | {row['Config_Str']}"
        return label if len(label) <= 42 else label[:39] + "..."

    def build_run_color(row):
        if row["Output_Category"] == "tool_call":
            return "seagreen"
        if row["Status"] == "ok":
            return "skyblue"
        if row["Status"] == "warning":
            return "darkorange"
        return "indianred"

    plot_df["Plot_Label"] = plot_df.apply(build_label, axis=1)
    plot_df["Plot_Color"] = plot_df.apply(build_run_color, axis=1)

    if capability == "tools":
        first_event_values = plot_df["First_Event_s"].fillna(0)
        duration_values = plot_df["Stream_Duration_s"].fillna(0)
        outcome_df = build_outcome_summary_dataframe(plot_df)

        fig, axes = plt.subplots(3, 1, figsize=(14, 15))

        axes[0].bar(plot_df["Plot_Label"], first_event_values, color=plot_df["Plot_Color"])
        if plot_df["First_Event_s"].notna().any():
            axes[0].axhline(
                y=plot_df["First_Event_s"].min(),
                color="green",
                linestyle="--",
                alpha=0.3,
            )
        axes[0].set_title("Tools Benchmark First Event Comparison")
        axes[0].set_ylabel("First Event (s)")

        axes[1].bar(plot_df["Plot_Label"], duration_values, color=plot_df["Plot_Color"])
        if plot_df["Stream_Duration_s"].notna().any():
            axes[1].axhline(
                y=plot_df["Stream_Duration_s"].min(),
                color="green",
                linestyle="--",
                alpha=0.3,
            )
        axes[1].set_title("Tools Benchmark Stream Duration Comparison")
        axes[1].set_ylabel("Stream Duration (s)")

        axes[2].bar(outcome_df["Output Category"], outcome_df["Count"], color="steelblue")
        axes[2].set_title("Tools Outcome Category Counts")
        axes[2].set_ylabel("Runs")
        axes[2].set_xlabel("Output Category")

        axes[0].tick_params(axis="x", rotation=35)
        axes[1].tick_params(axis="x", rotation=35)
        axes[2].tick_params(axis="x", rotation=20)
    else:
        eligible_df = filter_eligible_results(plot_df, capability=capability)
        if not eligible_df.empty:
            max_tps = eligible_df["TPS"].max()
            min_ttft = eligible_df["TTFT"].min()
            colors = ["gold" if tps == max_tps else "skyblue" for tps in eligible_df["TPS"]]
            ttft_colors = [
                "lightgreen" if ttft == min_ttft else "salmon" for ttft in eligible_df["TTFT"]
            ]
            has_vram_data = eligible_df["VRAM_Peak_MiB"].notna().any()
            has_efficiency_data = eligible_df["Efficiency_Score"].notna().any()

            subplot_count = 4 if has_efficiency_data else 3 if has_vram_data else 2
            fig, axes = plt.subplots(subplot_count, 1, figsize=(13, 5 * subplot_count), sharex=True)
            if subplot_count == 1:
                axes = [axes]

            axes[0].bar(eligible_df["Plot_Label"], eligible_df["TPS"], color=colors)
            axes[0].axhline(y=max_tps, color="red", linestyle="--", alpha=0.3)
            axes[0].set_title("Throughput Comparison")
            axes[0].set_ylabel("TPS (chunk/s)")

            axes[1].bar(eligible_df["Plot_Label"], eligible_df["TTFT"], color=ttft_colors)
            axes[1].axhline(y=min_ttft, color="green", linestyle="--", alpha=0.3)
            axes[1].set_title("First Token Latency Comparison")
            axes[1].set_ylabel("TTFT (s)")

            if has_vram_data:
                min_vram_peak = eligible_df["VRAM_Peak_MiB"].min()
                vram_colors = [
                    "lightgreen" if peak == min_vram_peak else "mediumpurple"
                    for peak in eligible_df["VRAM_Peak_MiB"]
                ]
                axes[2].bar(eligible_df["Plot_Label"], eligible_df["VRAM_Peak_MiB"], color=vram_colors)
                axes[2].axhline(y=min_vram_peak, color="green", linestyle="--", alpha=0.3)
                axes[2].set_title("VRAM Peak Comparison")
                axes[2].set_ylabel("VRAM Peak (MiB)")
                if has_efficiency_data:
                    max_efficiency = eligible_df["Efficiency_Score"].max()
                    efficiency_colors = [
                        "gold" if score == max_efficiency else "steelblue"
                        for score in eligible_df["Efficiency_Score"]
                    ]
                    axes[3].bar(
                        eligible_df["Plot_Label"],
                        eligible_df["Efficiency_Score"],
                        color=efficiency_colors,
                    )
                    axes[3].axhline(y=max_efficiency, color="orange", linestyle="--", alpha=0.3)
                    axes[3].set_title("Efficiency Score Comparison")
                    axes[3].set_ylabel("TPS/GiB Peak")
                    axes[3].set_xlabel("Model | Config")
                else:
                    axes[2].set_xlabel("Model | Config")
            else:
                axes[1].set_xlabel("Model | Config")
        else:
            status_df = (
                plot_df["Status"]
                .fillna("unknown")
                .value_counts(dropna=False)
                .rename_axis("Status")
                .reset_index(name="Count")
            )
            outcome_df = build_outcome_summary_dataframe(plot_df)
            fig, axes = plt.subplots(2, 1, figsize=(13, 10))

            axes[0].bar(status_df["Status"], status_df["Count"], color="steelblue")
            axes[0].set_title("Run Status Counts")
            axes[0].set_ylabel("Runs")

            axes[1].bar(outcome_df["Output Category"], outcome_df["Count"], color="slategray")
            title = "Outcome Category Counts"
            if used_fallback:
                title += " (Fallback)"
            axes[1].set_title(title)
            axes[1].set_ylabel("Runs")
            axes[1].set_xlabel("Output Category")
            axes[0].tick_params(axis="x", rotation=20)
            axes[1].tick_params(axis="x", rotation=20)

    plt.xticks(rotation=35, ha="right")
    plt.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)

    return Path(output_path)


def save_markdown_report(df, config, report_stem):
    report_path = Path(f"{report_stem}.md")
    summary_df = build_summary_dataframe(df)
    outcome_summary_df = build_outcome_summary_dataframe(df)

    with report_path.open("w", encoding="utf-8") as file:
        file.write("# Benchmark Report\n\n")
        file.write(f"- Backend: {config['backend']}\n")
        file.write(f"- Base URL: {config['url']}\n")
        file.write(f"- Models: {', '.join(config['models'])}\n")
        file.write(f"- Prompt: {config['prompt']}\n")
        file.write("- Note: TPS is estimated from streaming content chunks for relative comparison.\n\n")
        file.write("## Environment Notes\n\n")
        file.write(f"- VRAM monitoring: {config.get('vram_monitoring', 'unavailable')}\n\n")
        file.write("## Metric Notes\n\n")
        file.write("- `TPS (chunk/s)`: 每秒收到的文字內容 chunk 數，代表輸出吞吐速度；數值越高通常越快。\n")
        file.write("- `TTFT (s)`: Time To First Token，從送出請求到收到第一段文字內容的秒數；數值越低通常越快。\n")
        file.write("- `First Event (s)`: 從送出請求到收到第一個串流事件的秒數，包含 role 或非文字事件。\n")
        file.write("- `VRAM Peak (MiB)`: 測試期間輪詢到的 NVIDIA GPU 總顯存最高占用，用來比較實際壓力。\n")
        file.write("- `Efficiency Score (TPS/GiB Peak)`: `TPS / (VRAM Peak in GiB)`，代表每 1 GiB 峰值顯存換到多少輸出速度；數值越高越划算。\n")
        file.write("- `TPS (chunk/s)` 與 `TTFT (s)` 在沒有任何文字內容輸出時會顯示 `N/A`。\n\n")
        file.write("## Output Diagnosis Notes\n\n")
        file.write("- `empty_reply`: 串流正常結束，但沒有任何文字內容。\n")
        file.write("- `non_content_stream`: 串流有事件，但只有非文字 payload，例如 `tool_calls` 或 `reasoning`。\n")
        file.write("- `early_stop`: 串流在產生任何文字前提早結束，或在前期就被異常中斷。\n\n")
        file.write("## Summary\n\n")
        file.write(summary_df.to_markdown(index=False))
        file.write("\n\n## Outcome Summary\n\n")
        file.write(outcome_summary_df.to_markdown(index=False))
        file.write("\n\n## Generated Outputs\n")

        for _, row in df.iterrows():
            params_json = json.dumps(row["Params"], ensure_ascii=False)
            applied_params_json = json.dumps(row["Applied_Params"], ensure_ascii=False)
            file.write(f"\n### Run {row['Run_ID']}\n\n")
            file.write(f"- Status: {row['Status']}\n")
            file.write(f"- Output Category: {row['Output_Category']}\n")
            file.write(f"- Diagnosis: {row['Diagnosis']}\n")
            file.write(f"- Finish Reason: {format_text_value(row['Finish_Reason'])}\n")
            file.write(f"- Backend: {row['Backend']}\n")
            file.write(f"- Model: {row['Model']}\n")
            file.write(f"- Params: `{params_json}`\n")
            file.write(f"- Applied Params: `{applied_params_json}`\n")
            file.write(f"- TPS: {format_numeric_value(row['TPS'], 2)} chunk/s\n")
            file.write(f"- TTFT: {format_numeric_value(row['TTFT'], 3)} s\n")
            file.write(f"- First Event: {format_numeric_value(row['First_Event_s'], 3)} s\n")
            file.write(f"- Stream Duration: {format_numeric_value(row['Stream_Duration_s'], 3)} s\n")
            file.write(f"- Total Chunks: {int(row['Total_Chunks'])}\n")
            file.write(f"- Content Chunks: {int(row['Content_Chunks'])}\n")
            file.write(f"- Non-Content Chunks: {int(row['Non_Content_Chunks'])}\n")
            file.write(f"- Non-Content Types: {row['Non_Content_Types']}\n")
            file.write(f"- VRAM Base: {format_mib_value(row['VRAM_Base_MiB'])}\n")
            file.write(f"- VRAM Peak: {format_mib_value(row['VRAM_Peak_MiB'])}\n")
            file.write(f"- VRAM Delta: {format_mib_value(row['VRAM_Delta_MiB'])}\n")
            file.write(f"- VRAM Detail: {row['VRAM_Detail']}\n")
            file.write(
                f"- Efficiency Score: "
                f"{format_numeric_value(row['Efficiency_Score'], 3)} TPS/GiB Peak\n"
            )
            file.write(f"- Output Chars: {row['Output_Chars']}\n")
            if row["Error"]:
                file.write(f"- Error: {row['Error']}\n")
            file.write("\n```text\n")
            file.write(normalize_output_text(row["Output_Text"]))
            file.write("\n```\n")

    return report_path


def export_best_config(df, config):
    eligible_df = filter_eligible_results(df)
    if eligible_df.empty:
        print("⚠️ 沒有正常文字輸出的結果可匯出最佳配置。")
        return None

    best_row = eligible_df.loc[eligible_df["TPS"].idxmax()]
    payload = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "backend": best_row["Backend"],
        "model": best_row["Model"],
        "params": best_row["Params"],
        "applied_params": best_row["Applied_Params"],
        "status": best_row["Status"],
        "output_category": best_row["Output_Category"],
        "finish_reason": best_row["Finish_Reason"],
        "diagnosis": best_row["Diagnosis"],
        "tps": best_row["TPS"],
        "ttft": best_row["TTFT"],
        "first_event_s": best_row["First_Event_s"],
        "stream_duration_s": best_row["Stream_Duration_s"],
        "vram_base_mib": best_row["VRAM_Base_MiB"],
        "vram_peak_mib": best_row["VRAM_Peak_MiB"],
        "vram_delta_mib": best_row["VRAM_Delta_MiB"],
        "vram_detail": best_row["VRAM_Detail"],
        "efficiency_score_tps_per_gib_peak": best_row["Efficiency_Score"],
        "prompt": config["prompt"],
    }

    with open("best_config.json", "w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=4)

    print("\n" + "⭐" * 18)
    print(f"🏆 性能冠軍: {best_row['Model']}")
    print(f"🚀 最高速度: {format_numeric_value(best_row['TPS'], 2)} TPS")
    print(f"⚙️ 最佳配置: {best_row['Config_Str']}")
    print(f"🧠 顯存峰值: {format_mib_value(best_row['VRAM_Peak_MiB'])}")
    if pd.notna(best_row["Efficiency_Score"]):
        print(f"📊 效率分數: {format_numeric_value(best_row['Efficiency_Score'], 3)} TPS/GiB Peak")
    print("⭐" * 18)
    print("✅ 已保存 best_config.json")

    if best_row["Backend"] == "ollama":
        with open("Ollama_Modelfile_Suggest", "w", encoding="utf-8") as file:
            file.write(f"FROM {best_row['Model']}\n")
            for key, value in best_row["Applied_Params"].items():
                file.write(f"PARAMETER {key} {value}\n")
        print("✅ 已生成 Ollama_Modelfile_Suggest")
    else:
        print("ℹ️ 本次後端不是 Ollama，略過 Modelfile 建議。")

    return payload


def plot_results(df, output_path):
    eligible_df = filter_eligible_results(df)
    if eligible_df.empty:
        return None

    def build_label(row):
        label = f"{row['Model']} | {row['Config_Str']}"
        return label if len(label) <= 42 else label[:39] + "..."

    eligible_df["Plot_Label"] = eligible_df.apply(build_label, axis=1)
    max_tps = eligible_df["TPS"].max()
    min_ttft = eligible_df["TTFT"].min()
    colors = ["gold" if tps == max_tps else "skyblue" for tps in eligible_df["TPS"]]
    ttft_colors = ["lightgreen" if ttft == min_ttft else "salmon" for ttft in eligible_df["TTFT"]]
    has_vram_data = eligible_df["VRAM_Peak_MiB"].notna().any()
    has_efficiency_data = eligible_df["Efficiency_Score"].notna().any()

    subplot_count = 4 if has_efficiency_data else 3 if has_vram_data else 2
    fig, axes = plt.subplots(subplot_count, 1, figsize=(13, 5 * subplot_count), sharex=True)
    if subplot_count == 1:
        axes = [axes]

    axes[0].bar(eligible_df["Plot_Label"], eligible_df["TPS"], color=colors)
    axes[0].axhline(y=max_tps, color="red", linestyle="--", alpha=0.3)
    axes[0].set_title("Throughput Comparison")
    axes[0].set_ylabel("TPS (chunk/s)")

    axes[1].bar(eligible_df["Plot_Label"], eligible_df["TTFT"], color=ttft_colors)
    axes[1].axhline(y=min_ttft, color="green", linestyle="--", alpha=0.3)
    axes[1].set_title("First Token Latency Comparison")
    axes[1].set_ylabel("TTFT (s)")

    if has_vram_data:
        min_vram_peak = eligible_df["VRAM_Peak_MiB"].min()
        vram_colors = [
            "lightgreen" if peak == min_vram_peak else "mediumpurple"
            for peak in eligible_df["VRAM_Peak_MiB"]
        ]
        axes[2].bar(eligible_df["Plot_Label"], eligible_df["VRAM_Peak_MiB"], color=vram_colors)
        axes[2].axhline(y=min_vram_peak, color="green", linestyle="--", alpha=0.3)
        axes[2].set_title("VRAM Peak Comparison")
        axes[2].set_ylabel("VRAM Peak (MiB)")
        if has_efficiency_data:
            max_efficiency = eligible_df["Efficiency_Score"].max()
            efficiency_colors = [
                "gold" if score == max_efficiency else "steelblue"
                for score in eligible_df["Efficiency_Score"]
            ]
            axes[3].bar(
                eligible_df["Plot_Label"],
                eligible_df["Efficiency_Score"],
                color=efficiency_colors,
            )
            axes[3].axhline(y=max_efficiency, color="orange", linestyle="--", alpha=0.3)
            axes[3].set_title("Efficiency Score Comparison")
            axes[3].set_ylabel("TPS/GiB Peak")
            axes[3].set_xlabel("Model | Config")
        else:
            axes[2].set_xlabel("Model | Config")
    else:
        axes[1].set_xlabel("Model | Config")

    plt.xticks(rotation=35, ha="right")
    plt.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)

    return Path(output_path)


def main():
    config = interactive_config()
    if not config:
        return

    results_df = run_bench(config)
    if results_df.empty:
        print("⚠️ 沒有任何測試結果。")
        return

    ok_count = int((results_df["Status"] == "ok").sum())
    warning_count = int((results_df["Status"] == "warning").sum())
    error_count = int((results_df["Status"] == "error").sum())

    print("\n" + "═" * 62)
    console_df = build_summary_dataframe(results_df)[
        [
            "Run",
            "Status",
            "Output Category",
            "Finish Reason",
            "Model",
            "TPS (chunk/s)",
            "TTFT (s)",
            "First Event (s)",
            "Chunks (content/total)",
            "Config",
        ]
    ]
    print(console_df.to_markdown(index=False))
    print(f"\n📌 結果統計: ok={ok_count}, warning={warning_count}, error={error_count}")

    report_stem = f"bench_{config['backend']}_{time.strftime('%Y%m%d_%H%M%S')}"
    report_path = save_markdown_report(results_df, config, report_stem)
    chart_path = plot_results(results_df, f"{report_stem}.png")
    export_best_config(results_df, config)

    print(f"\n✅ 報告: {report_path}")
    if chart_path:
        print(f"📈 圖表: {chart_path}")
    else:
        print("ℹ️ 沒有可供繪圖的正常文字輸出結果。")


def interactive_config():
    print("\n" + "=" * 62)
    print("LLM Benchmark V3")
    print("=" * 62)

    backend = questionary.select(
        "請選擇測試後端:",
        choices=[
            Choice("Ollama", value="ollama"),
            Choice("llama.cpp（llama-server）", value="llama.cpp"),
        ],
    ).ask()
    if not backend:
        return None

    capability = questionary.select(
        "請選擇要測試的能力:",
        choices=[
            Choice(
                f"{info['label']} | {info['description']}",
                value=capability_key,
            )
            for capability_key, info in CAPABILITY_OPTIONS.items()
        ],
    ).ask()
    if not capability:
        return None

    url, models = select_models_and_url(backend)
    if not models:
        print("⚠️ 沒有可用模型，已取消。")
        return None

    available_groups = {
        group_name: [key for key in param_keys if backend in PARAM_INFO[key]["backends"]]
        for group_name, param_keys in PARAM_GROUPS.items()
    }
    available_groups = {name: keys for name, keys in available_groups.items() if keys}

    selected_groups = questionary.checkbox(
        "選擇想測試的參數類別（可不選，代表只比模型預設值）:",
        choices=[
            Choice(f"{group_name} ({len(param_keys)} 個)", value=group_name)
            for group_name, param_keys in available_groups.items()
        ],
    ).ask() or []

    final_params = {}
    for group_name in selected_groups:
        param_keys = available_groups[group_name]
        selected_params = questionary.checkbox(
            f"選擇 {group_name} 內要測試的參數:",
            choices=[
                Choice(
                    title=(
                        f"{PARAM_INFO[key]['label']} | 範圍: {PARAM_INFO[key]['range']} | "
                        f"{PARAM_INFO[key]['desc']}"
                    ),
                    value=key,
                )
                for key in param_keys
            ],
        ).ask() or []

        for key in selected_params:
            values = ask_param_values(key)
            if values is None:
                return None
            final_params[key] = values

    prompt = questionary.text(
        "測試 Prompt:",
        default=CAPABILITY_OPTIONS[capability]["default_prompt"],
    ).ask()
    if prompt is None:
        return None

    return {
        "backend": backend,
        "capability": capability,
        "url": url,
        "models": models,
        "params": final_params,
        "prompt": prompt,
    }


def run_bench(config):
    client = OpenAI(base_url=config["url"], api_key="sk-no-key-needed")
    capability = config.get("capability", "chat")
    param_keys = list(config["params"].keys())
    param_values = [config["params"][key] for key in param_keys]
    combos = [dict(zip(param_keys, combo)) for combo in product(*param_values)] if param_keys else [{}]

    results = []
    total_runs = len(config["models"]) * len(combos)
    vram_monitoring_enabled = query_nvidia_vram_snapshot() is not None
    config["vram_monitoring"] = "nvidia-smi" if vram_monitoring_enabled else "unavailable"

    capability_label = CAPABILITY_OPTIONS.get(capability, {}).get("label", capability)
    print(f"\n開始測試，共 {total_runs} 組，模式: {capability_label}")
    print("TPS 與 TTFT 適用於文字輸出；tools 模式主要看 First Event 與 tool_call 成功率。")
    if vram_monitoring_enabled:
        print("VRAM 監控: 已啟用 nvidia-smi")
    else:
        print("VRAM 監控: 未偵測到 nvidia-smi，相關欄位將顯示 N/A")

    run_index = 0
    for model in config["models"]:
        for param_set in combos:
            run_index += 1
            applied_params = build_backend_options(config["backend"], param_set)
            display_params = format_param_dict(param_set)
            print(f"[{run_index}/{total_runs}] {model} | {display_params}")

            request_kwargs = {}
            if applied_params:
                request_kwargs["extra_body"] = (
                    {"options": applied_params}
                    if config["backend"] == "ollama"
                    else applied_params
                )

            start_time = time.time()
            first_event_time = None
            first_content_time = None
            output_parts = []
            chunk_records = []
            vram_monitor = NvidiaVRAMMonitor() if vram_monitoring_enabled else None
            vram_metrics = empty_vram_metrics()
            if vram_monitor is not None and not vram_monitor.start():
                vram_monitor = None

            try:
                request_payload = build_chat_request_payload(config, model, request_kwargs)
                stream = client.chat.completions.create(**request_payload)

                for chunk in stream:
                    event_time = time.time()
                    if first_event_time is None:
                        first_event_time = event_time

                    chunk_info = inspect_stream_chunk(chunk)
                    chunk_records.append(chunk_info)

                    if chunk_info["content"] and first_content_time is None:
                        first_content_time = event_time
                    if chunk_info["content"]:
                        output_parts.append(chunk_info["content"])

                end_time = time.time()
                if vram_monitor is not None:
                    vram_metrics = vram_monitor.stop()
                output_text = "".join(output_parts)
                classification = classify_stream_result(
                    chunk_records=chunk_records,
                    start_time=start_time,
                    end_time=end_time,
                    first_event_time=first_event_time,
                    first_content_time=first_content_time,
                    error_message=None,
                )
                classification = adjust_classification_for_capability(classification, capability)
                results.append(
                    build_result_row(
                        run_id=run_index,
                        config=config,
                        model=model,
                        param_set=param_set,
                        applied_params=applied_params,
                        display_params=display_params,
                        classification=classification,
                        vram_metrics=vram_metrics,
                        output_text=output_text,
                        error_message=None,
                    )
                )
            except Exception as exc:
                end_time = time.time()
                if vram_monitor is not None:
                    vram_metrics = vram_monitor.stop()
                output_text = "".join(output_parts)
                classification = classify_stream_result(
                    chunk_records=chunk_records,
                    start_time=start_time,
                    end_time=end_time,
                    first_event_time=first_event_time,
                    first_content_time=first_content_time,
                    error_message=str(exc),
                )
                classification = adjust_classification_for_capability(classification, capability)
                print(f"錯誤: {exc}")
                results.append(
                    build_result_row(
                        run_id=run_index,
                        config=config,
                        model=model,
                        param_set=param_set,
                        applied_params=applied_params,
                        display_params=display_params,
                        classification=classification,
                        vram_metrics=vram_metrics,
                        output_text=output_text,
                        error_message=str(exc),
                    )
                )

    return pd.DataFrame(results)


def save_markdown_report(df, config, report_stem):
    report_path = Path(f"{report_stem}.html")
    summary_df = build_summary_dataframe(df)
    outcome_summary_df = build_outcome_summary_dataframe(df)
    wrapped_summary_df = wrap_markdown_table_headers(summary_df)
    wrapped_outcome_summary_df = wrap_markdown_table_headers(outcome_summary_df)
    capability = config.get("capability", "chat")
    tool_call_success_summary_df = (
        build_tool_call_success_summary_dataframe(df) if capability == "tools" else pd.DataFrame()
    )
    wrapped_tool_call_success_summary_df = wrap_markdown_table_headers(tool_call_success_summary_df)

    with report_path.open("w", encoding="utf-8") as file:
        file.write("# Benchmark Report\n\n")
        file.write(f"- Backend: {config['backend']}\n")
        file.write(f"- Capability: {capability}\n")
        file.write(f"- Base URL: {config['url']}\n")
        file.write(f"- Models: {', '.join(config['models'])}\n")
        file.write(f"- Prompt: {config['prompt']}\n")
        file.write("- Note: TPS is estimated from streaming content chunks for relative comparison.\n")
        if capability == "tools":
            file.write(
                "- Tool mode note: successful tool-calling runs may not emit text tokens, "
                "so `TPS` and `TTFT` can be `N/A`; focus on `Output Category=tool_call` "
                "and `First Event (s)`.\n"
            )
        file.write("\n## Environment Notes\n\n")
        file.write(f"- VRAM monitoring: {config.get('vram_monitoring', 'unavailable')}\n\n")
        file.write("## Metric Notes\n\n")
        file.write("- `TPS (chunk/s)`: Estimated throughput from text-bearing streaming chunks.\n")
        file.write("- `TTFT (s)`: Time to first text chunk.\n")
        file.write("- `First Event (s)`: Time to the first streamed event of any kind.\n")
        file.write("- `VRAM Peak (MiB)`: Highest observed total NVIDIA GPU memory usage during a run.\n")
        file.write(
            "- `Efficiency Score (TPS/GiB Peak)`: `TPS / (VRAM Peak in GiB)` when both values are available.\n\n"
        )
        file.write("## Output Diagnosis Notes\n\n")
        file.write("- `normal_content`: Received text output as expected for chat benchmarking.\n")
        file.write("- `tool_call`: Received `tool_calls` payload as expected for tool benchmarking.\n")
        file.write("- `text_reply_without_tool`: Returned text, but did not emit any tool call in tool mode.\n")
        file.write("- `empty_reply`: Stream completed without text output.\n")
        file.write("- `non_content_stream`: Stream only carried non-text payloads.\n")
        file.write("- `early_stop`: Stream ended before a complete reply or tool call was received.\n\n")
        file.write("## Summary\n\n")
        file.write(summary_df.to_markdown(index=False))
        file.write("\n\n## Outcome Summary\n\n")
        file.write(outcome_summary_df.to_markdown(index=False))
        file.write("\n\n## Generated Outputs\n")

        for _, row in df.iterrows():
            params_json = json.dumps(row["Params"], ensure_ascii=False)
            applied_params_json = json.dumps(row["Applied_Params"], ensure_ascii=False)
            file.write(f"\n### Run {row['Run_ID']}\n\n")
            file.write(f"- Status: {row['Status']}\n")
            file.write(f"- Capability: {row.get('Capability', capability)}\n")
            file.write(f"- Output Category: {row['Output_Category']}\n")
            file.write(f"- Diagnosis: {row['Diagnosis']}\n")
            file.write(f"- Finish Reason: {format_text_value(row['Finish_Reason'])}\n")
            file.write(f"- Backend: {row['Backend']}\n")
            file.write(f"- Model: {row['Model']}\n")
            file.write(f"- Params: `{params_json}`\n")
            file.write(f"- Applied Params: `{applied_params_json}`\n")
            file.write(f"- TPS: {format_numeric_value(row['TPS'], 2)} chunk/s\n")
            file.write(f"- TTFT: {format_numeric_value(row['TTFT'], 3)} s\n")
            file.write(f"- First Event: {format_numeric_value(row['First_Event_s'], 3)} s\n")
            file.write(f"- Stream Duration: {format_numeric_value(row['Stream_Duration_s'], 3)} s\n")
            file.write(f"- Total Chunks: {int(row['Total_Chunks'])}\n")
            file.write(f"- Content Chunks: {int(row['Content_Chunks'])}\n")
            file.write(f"- Non-Content Chunks: {int(row['Non_Content_Chunks'])}\n")
            file.write(f"- Non-Content Types: {row['Non_Content_Types']}\n")
            file.write(f"- VRAM Base: {format_mib_value(row['VRAM_Base_MiB'])}\n")
            file.write(f"- VRAM Peak: {format_mib_value(row['VRAM_Peak_MiB'])}\n")
            file.write(f"- VRAM Delta: {format_mib_value(row['VRAM_Delta_MiB'])}\n")
            file.write(f"- VRAM Detail: {row['VRAM_Detail']}\n")
            file.write(
                f"- Efficiency Score: "
                f"{format_numeric_value(row['Efficiency_Score'], 3)} TPS/GiB Peak\n"
            )
            file.write(f"- Output Chars: {row['Output_Chars']}\n")
            if row["Error"]:
                file.write(f"- Error: {row['Error']}\n")
            file.write("\n```text\n")
            file.write(normalize_output_text(row["Output_Text"]))
            file.write("\n```\n")

    return report_path


def select_best_result(eligible_df, capability):
    if capability == "tools":
        if eligible_df["First_Event_s"].notna().any():
            best_row = eligible_df.loc[eligible_df["First_Event_s"].idxmin()]
            return best_row, "first_event_s"
        best_row = eligible_df.loc[eligible_df["Stream_Duration_s"].idxmin()]
        return best_row, "stream_duration_s"

    if eligible_df["TPS"].notna().any():
        best_row = eligible_df.loc[eligible_df["TPS"].idxmax()]
        return best_row, "tps"

    best_row = eligible_df.loc[eligible_df["TTFT"].idxmin()]
    return best_row, "ttft"


def export_best_config(df, config):
    capability = config.get("capability", "chat")
    eligible_df = filter_eligible_results(df, capability=capability)
    if eligible_df.empty:
        print("⚠️ 沒有符合本次測試模式的成功結果，因此不輸出 best_config.json。")
        return None

    best_row, selection_metric = select_best_result(eligible_df, capability)
    payload = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "backend": best_row["Backend"],
        "capability": capability,
        "selection_metric": selection_metric,
        "model": best_row["Model"],
        "system_prompt_label": best_row.get("System_Prompt_Label", "N/A"),
        "system_prompt_text": best_row.get("System_Prompt_Text", ""),
        "params": best_row["Params"],
        "applied_params": best_row["Applied_Params"],
        "status": best_row["Status"],
        "output_category": best_row["Output_Category"],
        "finish_reason": best_row["Finish_Reason"],
        "diagnosis": best_row["Diagnosis"],
        "tps": best_row["TPS"],
        "ttft": best_row["TTFT"],
        "first_event_s": best_row["First_Event_s"],
        "stream_duration_s": best_row["Stream_Duration_s"],
        "vram_base_mib": best_row["VRAM_Base_MiB"],
        "vram_peak_mib": best_row["VRAM_Peak_MiB"],
        "vram_delta_mib": best_row["VRAM_Delta_MiB"],
        "vram_detail": best_row["VRAM_Detail"],
        "efficiency_score_tps_per_gib_peak": best_row["Efficiency_Score"],
        "prompt": config["prompt"],
    }

    with open("best_config.json", "w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=4)

    print("\n" + "=" * 18)
    print(f"最佳模型: {best_row['Model']}")
    if capability == "tools":
        print(f"最佳首事件延遲: {format_numeric_value(best_row['First_Event_s'], 3)} s")
        print(f"輸出類型: {best_row['Output_Category']}")
    else:
        print(f"最高 TPS: {format_numeric_value(best_row['TPS'], 2)} TPS")
    print(f"最佳設定: {best_row['Config_Str']}")
    print(f"VRAM Peak: {format_mib_value(best_row['VRAM_Peak_MiB'])}")
    if pd.notna(best_row["Efficiency_Score"]):
        print(f"效率分數: {format_numeric_value(best_row['Efficiency_Score'], 3)} TPS/GiB Peak")
    print("=" * 18)
    print("已保存 best_config.json")

    if best_row["Backend"] == "ollama":
        with open("Ollama_Modelfile_Suggest", "w", encoding="utf-8") as file:
            file.write(f"FROM {best_row['Model']}\n")
            for key, value in best_row["Applied_Params"].items():
                file.write(f"PARAMETER {key} {value}\n")
        print("已輸出 Ollama_Modelfile_Suggest")
    else:
        print("本次後端不是 Ollama，略過 Modelfile 建議。")

    return payload


def plot_results(df, output_path, capability="chat"):
    eligible_df = filter_eligible_results(df, capability=capability)
    if eligible_df.empty:
        return None

    eligible_df = eligible_df.copy()

    def build_label(row):
        label = f"{row['Model']} | {row['Config_Str']}"
        return label if len(label) <= 42 else label[:39] + "..."

    eligible_df["Plot_Label"] = eligible_df.apply(build_label, axis=1)

    if capability == "tools":
        min_first_event = eligible_df["First_Event_s"].min()
        min_stream_duration = eligible_df["Stream_Duration_s"].min()
        fig, axes = plt.subplots(2, 1, figsize=(13, 10), sharex=True)
        first_event_colors = [
            "lightgreen" if value == min_first_event else "steelblue"
            for value in eligible_df["First_Event_s"]
        ]
        duration_colors = [
            "lightgreen" if value == min_stream_duration else "slategray"
            for value in eligible_df["Stream_Duration_s"]
        ]

        axes[0].bar(eligible_df["Plot_Label"], eligible_df["First_Event_s"], color=first_event_colors)
        axes[0].axhline(y=min_first_event, color="green", linestyle="--", alpha=0.3)
        axes[0].set_title("Tool Call First Event Comparison")
        axes[0].set_ylabel("First Event (s)")

        axes[1].bar(
            eligible_df["Plot_Label"],
            eligible_df["Stream_Duration_s"],
            color=duration_colors,
        )
        axes[1].axhline(y=min_stream_duration, color="green", linestyle="--", alpha=0.3)
        axes[1].set_title("Tool Call Stream Duration Comparison")
        axes[1].set_ylabel("Stream Duration (s)")
        axes[1].set_xlabel("Model | Config")
    else:
        max_tps = eligible_df["TPS"].max()
        min_ttft = eligible_df["TTFT"].min()
        colors = ["gold" if tps == max_tps else "skyblue" for tps in eligible_df["TPS"]]
        ttft_colors = ["lightgreen" if ttft == min_ttft else "salmon" for ttft in eligible_df["TTFT"]]
        has_vram_data = eligible_df["VRAM_Peak_MiB"].notna().any()
        has_efficiency_data = eligible_df["Efficiency_Score"].notna().any()

        subplot_count = 4 if has_efficiency_data else 3 if has_vram_data else 2
        fig, axes = plt.subplots(subplot_count, 1, figsize=(13, 5 * subplot_count), sharex=True)
        if subplot_count == 1:
            axes = [axes]

        axes[0].bar(eligible_df["Plot_Label"], eligible_df["TPS"], color=colors)
        axes[0].axhline(y=max_tps, color="red", linestyle="--", alpha=0.3)
        axes[0].set_title("Throughput Comparison")
        axes[0].set_ylabel("TPS (chunk/s)")

        axes[1].bar(eligible_df["Plot_Label"], eligible_df["TTFT"], color=ttft_colors)
        axes[1].axhline(y=min_ttft, color="green", linestyle="--", alpha=0.3)
        axes[1].set_title("First Token Latency Comparison")
        axes[1].set_ylabel("TTFT (s)")

        if has_vram_data:
            min_vram_peak = eligible_df["VRAM_Peak_MiB"].min()
            vram_colors = [
                "lightgreen" if peak == min_vram_peak else "mediumpurple"
                for peak in eligible_df["VRAM_Peak_MiB"]
            ]
            axes[2].bar(eligible_df["Plot_Label"], eligible_df["VRAM_Peak_MiB"], color=vram_colors)
            axes[2].axhline(y=min_vram_peak, color="green", linestyle="--", alpha=0.3)
            axes[2].set_title("VRAM Peak Comparison")
            axes[2].set_ylabel("VRAM Peak (MiB)")
            if has_efficiency_data:
                max_efficiency = eligible_df["Efficiency_Score"].max()
                efficiency_colors = [
                    "gold" if score == max_efficiency else "steelblue"
                    for score in eligible_df["Efficiency_Score"]
                ]
                axes[3].bar(
                    eligible_df["Plot_Label"],
                    eligible_df["Efficiency_Score"],
                    color=efficiency_colors,
                )
                axes[3].axhline(y=max_efficiency, color="orange", linestyle="--", alpha=0.3)
                axes[3].set_title("Efficiency Score Comparison")
                axes[3].set_ylabel("TPS/GiB Peak")
                axes[3].set_xlabel("Model | Config")
            else:
                axes[2].set_xlabel("Model | Config")
        else:
            axes[1].set_xlabel("Model | Config")

    plt.xticks(rotation=35, ha="right")
    plt.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)

    return Path(output_path)


def main():
    config = interactive_config()
    if not config:
        return

    results_df = run_bench(config)
    if results_df.empty:
        print("⚠️ 沒有產生任何 benchmark 結果。")
        return

    ok_count = int((results_df["Status"] == "ok").sum())
    warning_count = int((results_df["Status"] == "warning").sum())
    error_count = int((results_df["Status"] == "error").sum())

    print("\n" + "=" * 62)
    summary_df = build_summary_dataframe(results_df)
    console_columns = ["Run", "Status"]
    if "Capability" in summary_df.columns:
        console_columns.append("Capability")
    console_columns.extend(
        [
            "Output Category",
            "Finish Reason",
            "Model",
            "TPS (chunk/s)",
            "TTFT (s)",
            "First Event (s)",
            "Chunks (content/total)",
            "Config",
        ]
    )
    console_df = summary_df[console_columns]
    print(console_df.to_markdown(index=False))
    print(f"\n結果統計: ok={ok_count}, warning={warning_count}, error={error_count}")

    capability = config.get("capability", "chat")
    report_stem = f"bench_{config['backend']}_{capability}_{time.strftime('%Y%m%d_%H%M%S')}"
    report_path = save_markdown_report(results_df, config, report_stem)
    chart_path = plot_results(results_df, f"{report_stem}.png", capability=capability)
    export_best_config(results_df, config)

    print(f"\n報告已保存: {report_path}")
    if chart_path:
        print(f"圖表已保存: {chart_path}")
    else:
        print("沒有可繪圖的成功結果，略過圖表輸出。")


def interactive_config():
    capability_defaults = {
        "chat": "Explain the long-term creep risk of PETG in 3D printing and how to reduce it.",
        "tools": (
            "Check today's weather in Taipei. If you support tools or function calling, "
            "call the `lookup_weather` tool first instead of answering directly."
        ),
    }

    print("\n" + "=" * 62)
    print("LLM Benchmark V3")
    print("=" * 62)

    backend = questionary.select(
        "Select backend:",
        choices=[
            Choice("Ollama", value="ollama"),
            Choice("llama.cpp (llama-server)", value="llama.cpp"),
        ],
    ).ask()
    if not backend:
        return None

    capability = questionary.select(
        "Select benchmark mode:",
        choices=[
            Choice("Chat | Standard chat response benchmark", value="chat"),
            Choice("Tools | Check whether the model emits tool_calls", value="tools"),
        ],
    ).ask()
    if not capability:
        return None

    url, models = select_models_and_url(backend)
    if not models:
        print("No models available. Cancelled.")
        return None

    available_groups = {
        group_name: [key for key in param_keys if backend in PARAM_INFO[key]["backends"]]
        for group_name, param_keys in PARAM_GROUPS.items()
    }
    available_groups = {name: keys for name, keys in available_groups.items() if keys}

    selected_groups = questionary.checkbox(
        "Select parameter groups to benchmark (optional):",
        choices=[
            Choice(f"{group_name} ({len(param_keys)} params)", value=group_name)
            for group_name, param_keys in available_groups.items()
        ],
    ).ask() or []

    final_params = {}
    for group_name in selected_groups:
        param_keys = available_groups[group_name]
        selected_params = questionary.checkbox(
            f"Select params from {group_name}:",
            choices=[
                Choice(
                    title=(
                        f"{PARAM_INFO[key]['label']} | Range: {PARAM_INFO[key]['range']} | "
                        f"{PARAM_INFO[key]['desc']}"
                    ),
                    value=key,
                )
                for key in param_keys
            ],
        ).ask() or []

        for key in selected_params:
            values = ask_param_values(key)
            if values is None:
                return None
            final_params[key] = values

    prompt = questionary.text(
        "Benchmark prompt:",
        default=capability_defaults[capability],
    ).ask()
    if prompt is None:
        return None

    return {
        "backend": backend,
        "capability": capability,
        "url": url,
        "models": models,
        "params": final_params,
        "prompt": prompt,
    }


def run_bench(config):
    client = OpenAI(base_url=config["url"], api_key="sk-no-key-needed")
    capability = config.get("capability", "chat")
    capability_label = {"chat": "chat", "tools": "tools"}.get(capability, capability)
    param_keys = list(config["params"].keys())
    param_values = [config["params"][key] for key in param_keys]
    combos = [dict(zip(param_keys, combo)) for combo in product(*param_values)] if param_keys else [{}]

    results = []
    total_runs = len(config["models"]) * len(combos)
    vram_monitoring_enabled = query_nvidia_vram_snapshot() is not None
    config["vram_monitoring"] = "nvidia-smi" if vram_monitoring_enabled else "unavailable"

    print(f"\nStarting benchmark with {total_runs} runs. Mode: {capability_label}")
    print("Chat mode focuses on TPS/TTFT. Tools mode focuses on tool_call success and first event latency.")
    if vram_monitoring_enabled:
        print("VRAM monitoring: enabled via nvidia-smi")
    else:
        print("VRAM monitoring: nvidia-smi not available, VRAM fields will be N/A")

    run_index = 0
    for model in config["models"]:
        for param_set in combos:
            run_index += 1
            applied_params = build_backend_options(config["backend"], param_set)
            display_params = format_param_dict(param_set)
            print(f"[{run_index}/{total_runs}] {model} | {display_params}")

            request_kwargs = {}
            if applied_params:
                request_kwargs["extra_body"] = (
                    {"options": applied_params}
                    if config["backend"] == "ollama"
                    else applied_params
                )

            start_time = time.time()
            first_event_time = None
            first_content_time = None
            output_parts = []
            chunk_records = []
            vram_monitor = NvidiaVRAMMonitor() if vram_monitoring_enabled else None
            vram_metrics = empty_vram_metrics()
            if vram_monitor is not None and not vram_monitor.start():
                vram_monitor = None

            try:
                request_payload = build_chat_request_payload(config, model, request_kwargs)
                stream = client.chat.completions.create(**request_payload)

                for chunk in stream:
                    event_time = time.time()
                    if first_event_time is None:
                        first_event_time = event_time

                    chunk_info = inspect_stream_chunk(chunk)
                    chunk_records.append(chunk_info)

                    if chunk_info["content"] and first_content_time is None:
                        first_content_time = event_time
                    if chunk_info["content"]:
                        output_parts.append(chunk_info["content"])

                end_time = time.time()
                if vram_monitor is not None:
                    vram_metrics = vram_monitor.stop()
                output_text = "".join(output_parts)
                classification = classify_stream_result(
                    chunk_records=chunk_records,
                    start_time=start_time,
                    end_time=end_time,
                    first_event_time=first_event_time,
                    first_content_time=first_content_time,
                    error_message=None,
                )
                classification = adjust_classification_for_capability(classification, capability)
                results.append(
                    build_result_row(
                        run_id=run_index,
                        config=config,
                        model=model,
                        param_set=param_set,
                        applied_params=applied_params,
                        display_params=display_params,
                        classification=classification,
                        vram_metrics=vram_metrics,
                        output_text=output_text,
                        error_message=None,
                    )
                )
            except Exception as exc:
                end_time = time.time()
                if vram_monitor is not None:
                    vram_metrics = vram_monitor.stop()
                output_text = "".join(output_parts)
                classification = classify_stream_result(
                    chunk_records=chunk_records,
                    start_time=start_time,
                    end_time=end_time,
                    first_event_time=first_event_time,
                    first_content_time=first_content_time,
                    error_message=str(exc),
                )
                classification = adjust_classification_for_capability(classification, capability)
                print(f"Error: {exc}")
                results.append(
                    build_result_row(
                        run_id=run_index,
                        config=config,
                        model=model,
                        param_set=param_set,
                        applied_params=applied_params,
                        display_params=display_params,
                        classification=classification,
                        vram_metrics=vram_metrics,
                        output_text=output_text,
                        error_message=str(exc),
                    )
                )

    return pd.DataFrame(results)


def save_markdown_report(df, config, report_stem):
    report_path = Path(f"{report_stem}.html")
    summary_df = build_summary_dataframe(df)
    outcome_summary_df = build_outcome_summary_dataframe(df)
    capability = config.get("capability", "chat")
    wrapped_summary_df = wrap_markdown_table_headers(summary_df)
    wrapped_outcome_summary_df = wrap_markdown_table_headers(outcome_summary_df)
    tool_call_success_summary_df = (
        build_tool_call_success_summary_dataframe(df) if capability == "tools" else pd.DataFrame()
    )
    wrapped_tool_call_success_summary_df = wrap_markdown_table_headers(
        tool_call_success_summary_df
    )

    with report_path.open("w", encoding="utf-8") as file:
        file.write("# Benchmark Report\n\n")
        file.write(f"- Backend: {config['backend']}\n")
        file.write(f"- Capability: {capability}\n")
        file.write(f"- Base URL: {config['url']}\n")
        file.write(f"- Models: {', '.join(config['models'])}\n")
        file.write(f"- Prompt: {config['prompt']}\n")
        file.write("- Note: TPS is estimated from streaming content chunks for relative comparison.\n")
        if capability == "tools":
            file.write(
                "- Tool mode note: successful tool-calling runs may not emit text tokens, "
                "so `TPS` and `TTFT` can be `N/A`; focus on `Output Category=tool_call` "
                "and `First Event (s)`.\n"
            )
        file.write("\n## Environment Notes\n\n")
        file.write(f"- VRAM monitoring: {config.get('vram_monitoring', 'unavailable')}\n\n")
        file.write("## Metric Notes\n\n")
        file.write("- `TPS (chunk/s)`: Estimated throughput from text-bearing streaming chunks.\n")
        file.write("- `TTFT (s)`: Time to first text chunk.\n")
        file.write("- `First Event (s)`: Time to the first streamed event of any kind.\n")
        file.write("- `VRAM Peak (MiB)`: Highest observed total NVIDIA GPU memory usage during a run.\n")
        file.write(
            "- `Efficiency Score (TPS/GiB Peak)`: `TPS / (VRAM Peak in GiB)` when both values are available.\n\n"
        )
        file.write("## Output Diagnosis Notes\n\n")
        file.write("- `normal_content`: Received text output as expected for chat benchmarking.\n")
        file.write("- `tool_call`: Received `tool_calls` payload as expected for tool benchmarking.\n")
        file.write("- `text_reply_without_tool`: Returned text, but did not emit any tool call in tool mode.\n")
        file.write("- `empty_reply`: Stream completed without text output.\n")
        file.write("- `non_content_stream`: Stream only carried non-text payloads.\n")
        file.write("- `early_stop`: Stream ended before a complete reply or tool call was received.\n\n")
        file.write("## Summary\n\n")
        file.write(summary_df.to_markdown(index=False))
        file.write("\n\n## Outcome Summary\n\n")
        file.write(outcome_summary_df.to_markdown(index=False))
        file.write("\n\n## Generated Outputs\n")

        for _, row in df.iterrows():
            params_json = json.dumps(row["Params"], ensure_ascii=False)
            applied_params_json = json.dumps(row["Applied_Params"], ensure_ascii=False)
            file.write(f"\n### Run {row['Run_ID']}\n\n")
            file.write(f"- Status: {row['Status']}\n")
            file.write(f"- Capability: {row.get('Capability', capability)}\n")
            file.write(f"- Output Category: {row['Output_Category']}\n")
            file.write(f"- Diagnosis: {row['Diagnosis']}\n")
            file.write(f"- Finish Reason: {format_text_value(row['Finish_Reason'])}\n")
            file.write(f"- Backend: {row['Backend']}\n")
            file.write(f"- Model: {row['Model']}\n")
            file.write(f"- Params: `{params_json}`\n")
            file.write(f"- Applied Params: `{applied_params_json}`\n")
            file.write(f"- TPS: {format_numeric_value(row['TPS'], 2)} chunk/s\n")
            file.write(f"- TTFT: {format_numeric_value(row['TTFT'], 3)} s\n")
            file.write(f"- First Event: {format_numeric_value(row['First_Event_s'], 3)} s\n")
            file.write(f"- Stream Duration: {format_numeric_value(row['Stream_Duration_s'], 3)} s\n")
            file.write(f"- Total Chunks: {int(row['Total_Chunks'])}\n")
            file.write(f"- Content Chunks: {int(row['Content_Chunks'])}\n")
            file.write(f"- Non-Content Chunks: {int(row['Non_Content_Chunks'])}\n")
            file.write(f"- Non-Content Types: {row['Non_Content_Types']}\n")
            file.write(f"- VRAM Base: {format_mib_value(row['VRAM_Base_MiB'])}\n")
            file.write(f"- VRAM Peak: {format_mib_value(row['VRAM_Peak_MiB'])}\n")
            file.write(f"- VRAM Delta: {format_mib_value(row['VRAM_Delta_MiB'])}\n")
            file.write(f"- VRAM Detail: {row['VRAM_Detail']}\n")
            file.write(
                f"- Efficiency Score: "
                f"{format_numeric_value(row['Efficiency_Score'], 3)} TPS/GiB Peak\n"
            )
            file.write(f"- Output Chars: {row['Output_Chars']}\n")
            if row["Error"]:
                file.write(f"- Error: {row['Error']}\n")
            file.write("\n```text\n")
            file.write(normalize_output_text(row["Output_Text"]))
            file.write("\n```\n")

    return report_path


def select_best_result(eligible_df, capability):
    if capability == "tools":
        if eligible_df["First_Event_s"].notna().any():
            best_row = eligible_df.loc[eligible_df["First_Event_s"].idxmin()]
            return best_row, "first_event_s"
        best_row = eligible_df.loc[eligible_df["Stream_Duration_s"].idxmin()]
        return best_row, "stream_duration_s"

    if eligible_df["TPS"].notna().any():
        best_row = eligible_df.loc[eligible_df["TPS"].idxmax()]
        return best_row, "tps"

    best_row = eligible_df.loc[eligible_df["TTFT"].idxmin()]
    return best_row, "ttft"


def export_best_config(df, config, output_dir="."):
    capability = config.get("capability", "chat")
    eligible_df = filter_eligible_results(df, capability=capability)
    if eligible_df.empty:
        print("No successful result for this benchmark mode, so best_config.json was not written.")
        return None

    best_row, selection_metric = select_best_result(eligible_df, capability)
    payload = {
        "generated_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "backend": best_row["Backend"],
        "capability": capability,
        "selection_metric": selection_metric,
        "model": best_row["Model"],
        "params": best_row["Params"],
        "applied_params": best_row["Applied_Params"],
        "status": best_row["Status"],
        "output_category": best_row["Output_Category"],
        "finish_reason": best_row["Finish_Reason"],
        "diagnosis": best_row["Diagnosis"],
        "tps": best_row["TPS"],
        "ttft": best_row["TTFT"],
        "first_event_s": best_row["First_Event_s"],
        "stream_duration_s": best_row["Stream_Duration_s"],
        "vram_base_mib": best_row["VRAM_Base_MiB"],
        "vram_peak_mib": best_row["VRAM_Peak_MiB"],
        "vram_delta_mib": best_row["VRAM_Delta_MiB"],
        "vram_detail": best_row["VRAM_Detail"],
        "efficiency_score_tps_per_gib_peak": best_row["Efficiency_Score"],
        "prompt": config["prompt"],
        "system_prompts": config.get("system_prompts", []),
    }
    serializable_payload = {
        key: serialize_result_value(value) for key, value in payload.items()
    }

    output_dir_path = Path(output_dir)
    output_dir_path.mkdir(parents=True, exist_ok=True)
    best_config_path = output_dir_path / "best_config.json"
    with best_config_path.open("w", encoding="utf-8") as file:
        json.dump(serializable_payload, file, ensure_ascii=False, indent=4)

    print("\n" + "=" * 18)
    print(f"Best model: {best_row['Model']}")
    if capability == "tools":
        print(f"Best first-event latency: {format_numeric_value(best_row['First_Event_s'], 3)} s")
        print(f"Output category: {best_row['Output_Category']}")
    else:
        print(f"Highest TPS: {format_numeric_value(best_row['TPS'], 2)} TPS")
    print(f"Best config: {best_row['Config_Str']}")
    print(f"VRAM Peak: {format_mib_value(best_row['VRAM_Peak_MiB'])}")
    if pd.notna(best_row["Efficiency_Score"]):
        print(f"Efficiency score: {format_numeric_value(best_row['Efficiency_Score'], 3)} TPS/GiB Peak")
    print("=" * 18)
    print(f"Saved best_config.json: {best_config_path}")

    modelfile_path = None
    if best_row["Backend"] == "ollama":
        modelfile_path = output_dir_path / "Ollama_Modelfile_Suggest"
        modelfile_params = build_ollama_modelfile_params(best_row["Params"])
        with modelfile_path.open("w", encoding="utf-8") as file:
            file.write(f"FROM {best_row['Model']}\n")
            for key, value in modelfile_params.items():
                file.write(f"PARAMETER {key} {value}\n")
        print(f"Saved Ollama_Modelfile_Suggest: {modelfile_path}")
    else:
        print("Backend is not Ollama, so Modelfile output was skipped.")

    return {
        "payload": serializable_payload,
        "best_config_path": best_config_path,
        "modelfile_path": modelfile_path,
    }


def plot_results(df, output_path, capability="chat"):
    eligible_df = filter_eligible_results(df, capability=capability)
    if eligible_df.empty:
        return None

    eligible_df = eligible_df.copy()

    def build_label(row):
        label = f"{row['Model']} | {row['Config_Str']}"
        return label if len(label) <= 42 else label[:39] + "..."

    eligible_df["Plot_Label"] = eligible_df.apply(build_label, axis=1)

    if capability == "tools":
        min_first_event = eligible_df["First_Event_s"].min()
        min_stream_duration = eligible_df["Stream_Duration_s"].min()
        fig, axes = plt.subplots(2, 1, figsize=(13, 10), sharex=True)
        first_event_colors = [
            "lightgreen" if value == min_first_event else "steelblue"
            for value in eligible_df["First_Event_s"]
        ]
        duration_colors = [
            "lightgreen" if value == min_stream_duration else "slategray"
            for value in eligible_df["Stream_Duration_s"]
        ]

        axes[0].bar(eligible_df["Plot_Label"], eligible_df["First_Event_s"], color=first_event_colors)
        axes[0].axhline(y=min_first_event, color="green", linestyle="--", alpha=0.3)
        axes[0].set_title("Tool Call First Event Comparison")
        axes[0].set_ylabel("First Event (s)")

        axes[1].bar(
            eligible_df["Plot_Label"],
            eligible_df["Stream_Duration_s"],
            color=duration_colors,
        )
        axes[1].axhline(y=min_stream_duration, color="green", linestyle="--", alpha=0.3)
        axes[1].set_title("Tool Call Stream Duration Comparison")
        axes[1].set_ylabel("Stream Duration (s)")
        axes[1].set_xlabel("Model | Config")
    else:
        max_tps = eligible_df["TPS"].max()
        min_ttft = eligible_df["TTFT"].min()
        colors = ["gold" if tps == max_tps else "skyblue" for tps in eligible_df["TPS"]]
        ttft_colors = ["lightgreen" if ttft == min_ttft else "salmon" for ttft in eligible_df["TTFT"]]
        has_vram_data = eligible_df["VRAM_Peak_MiB"].notna().any()
        has_efficiency_data = eligible_df["Efficiency_Score"].notna().any()

        subplot_count = 4 if has_efficiency_data else 3 if has_vram_data else 2
        fig, axes = plt.subplots(subplot_count, 1, figsize=(13, 5 * subplot_count), sharex=True)
        if subplot_count == 1:
            axes = [axes]

        axes[0].bar(eligible_df["Plot_Label"], eligible_df["TPS"], color=colors)
        axes[0].axhline(y=max_tps, color="red", linestyle="--", alpha=0.3)
        axes[0].set_title("Throughput Comparison")
        axes[0].set_ylabel("TPS (chunk/s)")

        axes[1].bar(eligible_df["Plot_Label"], eligible_df["TTFT"], color=ttft_colors)
        axes[1].axhline(y=min_ttft, color="green", linestyle="--", alpha=0.3)
        axes[1].set_title("First Token Latency Comparison")
        axes[1].set_ylabel("TTFT (s)")

        if has_vram_data:
            min_vram_peak = eligible_df["VRAM_Peak_MiB"].min()
            vram_colors = [
                "lightgreen" if peak == min_vram_peak else "mediumpurple"
                for peak in eligible_df["VRAM_Peak_MiB"]
            ]
            axes[2].bar(eligible_df["Plot_Label"], eligible_df["VRAM_Peak_MiB"], color=vram_colors)
            axes[2].axhline(y=min_vram_peak, color="green", linestyle="--", alpha=0.3)
            axes[2].set_title("VRAM Peak Comparison")
            axes[2].set_ylabel("VRAM Peak (MiB)")
            if has_efficiency_data:
                max_efficiency = eligible_df["Efficiency_Score"].max()
                efficiency_colors = [
                    "gold" if score == max_efficiency else "steelblue"
                    for score in eligible_df["Efficiency_Score"]
                ]
                axes[3].bar(
                    eligible_df["Plot_Label"],
                    eligible_df["Efficiency_Score"],
                    color=efficiency_colors,
                )
                axes[3].axhline(y=max_efficiency, color="orange", linestyle="--", alpha=0.3)
                axes[3].set_title("Efficiency Score Comparison")
                axes[3].set_ylabel("TPS/GiB Peak")
                axes[3].set_xlabel("Model | Config")
            else:
                axes[2].set_xlabel("Model | Config")
        else:
            axes[1].set_xlabel("Model | Config")

    plt.xticks(rotation=35, ha="right")
    plt.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)

    return Path(output_path)


def main():
    config = interactive_config()
    if not config:
        return

    results_df = run_bench(config)
    if results_df.empty:
        print("No benchmark rows were produced. / 沒有產生任何 benchmark 結果列。")
        return

    ok_count = int((results_df["Status"] == "ok").sum())
    warning_count = int((results_df["Status"] == "warning").sum())
    error_count = int((results_df["Status"] == "error").sum())

    print("\n" + "=" * 62)
    summary_df = build_summary_dataframe(results_df)
    console_columns = ["Run", "Status"]
    if "Capability" in summary_df.columns:
        console_columns.append("Capability")
    console_columns.extend(
        [
            "Output Category",
            "Finish Reason",
            "Model",
            "TPS (chunk/s)",
            "TTFT (s)",
            "First Event (s)",
            "Chunks (content/total)",
            "Config",
        ]
    )
    console_df = summary_df[console_columns]
    print(console_df.to_markdown(index=False))
    print(f"\nResult counts: ok={ok_count}, warning={warning_count}, error={error_count}")

    capability = config.get("capability", "chat")
    report_stem = f"bench_{config['backend']}_{capability}_{time.strftime('%Y%m%d_%H%M%S')}"
    report_path = save_markdown_report(results_df, config, report_stem)
    chart_path = plot_results(results_df, f"{report_stem}.png", capability=capability)
    export_best_config(results_df, config)

    print(f"\nSaved report: {report_path}")
    if chart_path:
        print(f"Saved chart: {chart_path}")
    else:
        print("Skipped chart output because there were no eligible successful results.")


def save_markdown_report(df, config, report_stem):
    report_path = Path(f"{report_stem}.html")
    summary_df = build_summary_dataframe(df)
    outcome_summary_df = build_outcome_summary_dataframe(df)
    capability = config.get("capability", "chat")
    wrapped_summary_df = wrap_markdown_table_headers(summary_df)
    wrapped_outcome_summary_df = wrap_markdown_table_headers(outcome_summary_df)
    tool_call_success_summary_df = (
        build_tool_call_success_summary_dataframe(df) if capability == "tools" else pd.DataFrame()
    )
    wrapped_tool_call_success_summary_df = wrap_markdown_table_headers(
        tool_call_success_summary_df
    )

    with report_path.open("w", encoding="utf-8") as file:
        file.write("# Benchmark Report\n\n")
        file.write(f"- Backend: {config['backend']}\n")
        file.write(f"- Capability: {capability}\n")
        file.write(f"- Base URL: {config['url']}\n")
        file.write(f"- Models: {', '.join(config['models'])}\n")
        file.write(f"- Prompt: {config['prompt']}\n")
        file.write("- Note: TPS is estimated from streaming content chunks for relative comparison.\n")
        if capability == "tools":
            file.write(
                "- Tool mode note: successful tool-calling runs may not emit text tokens, "
                "so `TPS` and `TTFT` can be `N/A`; focus on `Output Category=tool_call` "
                "and `First Event (s)`.\n"
            )
        file.write("\n## Environment Notes\n\n")
        file.write(f"- VRAM monitoring: {config.get('vram_monitoring', 'unavailable')}\n\n")
        file.write("## Metric Notes\n\n")
        file.write("- `TPS (chunk/s)`: Estimated throughput from text-bearing streaming chunks.\n")
        file.write("- `TTFT (s)`: Time to first text chunk.\n")
        file.write("- `First Event (s)`: Time to the first streamed event of any kind.\n")
        file.write("- `VRAM Peak (MiB)`: Highest observed total NVIDIA GPU memory usage during a run.\n")
        file.write(
            "- `Efficiency Score (TPS/GiB Peak)`: `TPS / (VRAM Peak in GiB)` when both values are available.\n\n"
        )
        file.write("## Output Diagnosis Notes\n\n")
        file.write("- `normal_content`: Received text output as expected for chat benchmarking.\n")
        file.write("- `tool_call`: Received `tool_calls` payload as expected for tool benchmarking.\n")
        file.write("- `text_reply_without_tool`: Returned text, but did not emit any tool call in tool mode.\n")
        file.write("- `empty_reply`: Stream completed without text output.\n")
        file.write("- `non_content_stream`: Stream only carried non-text payloads.\n")
        file.write("- `early_stop`: Stream ended before a complete reply or tool call was received.\n\n")
        file.write("## Summary\n\n")
        file.write(dataframe_to_report_table(wrapped_summary_df))
        file.write("\n\n## Outcome Summary\n\n")
        file.write(dataframe_to_report_table(wrapped_outcome_summary_df))
        if capability == "tools" and not tool_call_success_summary_df.empty:
            file.write("\n\n## Tool Call Success by Model\n\n")
            file.write(dataframe_to_report_table(wrapped_tool_call_success_summary_df))
        file.write("\n\n## Generated Outputs\n")

        for _, row in df.iterrows():
            params_json = json.dumps(row["Params"], ensure_ascii=False)
            applied_params_json = json.dumps(row["Applied_Params"], ensure_ascii=False)
            file.write(f"\n### Run {row['Run_ID']}\n\n")
            file.write(f"- Status: {row['Status']}\n")
            file.write(f"- Capability: {row.get('Capability', capability)}\n")
            file.write(f"- Output Category: {row['Output_Category']}\n")
            file.write(f"- Diagnosis: {row['Diagnosis']}\n")
            file.write(f"- Finish Reason: {format_text_value(row['Finish_Reason'])}\n")
            file.write(f"- Backend: {row['Backend']}\n")
            file.write(f"- Model: {row['Model']}\n")
            file.write(f"- Params: `{params_json}`\n")
            file.write(f"- Applied Params: `{applied_params_json}`\n")
            file.write(f"- TPS: {format_numeric_value(row['TPS'], 2)} chunk/s\n")
            file.write(f"- TTFT: {format_numeric_value(row['TTFT'], 3)} s\n")
            file.write(f"- First Event: {format_numeric_value(row['First_Event_s'], 3)} s\n")
            file.write(f"- Stream Duration: {format_numeric_value(row['Stream_Duration_s'], 3)} s\n")
            file.write(f"- Total Chunks: {int(row['Total_Chunks'])}\n")
            file.write(f"- Content Chunks: {int(row['Content_Chunks'])}\n")
            file.write(f"- Non-Content Chunks: {int(row['Non_Content_Chunks'])}\n")
            file.write(f"- Non-Content Types: {row['Non_Content_Types']}\n")
            file.write(f"- VRAM Base: {format_mib_value(row['VRAM_Base_MiB'])}\n")
            file.write(f"- VRAM Peak: {format_mib_value(row['VRAM_Peak_MiB'])}\n")
            file.write(f"- VRAM Delta: {format_mib_value(row['VRAM_Delta_MiB'])}\n")
            file.write(f"- VRAM Detail: {row['VRAM_Detail']}\n")
            file.write(
                f"- Efficiency Score: "
                f"{format_numeric_value(row['Efficiency_Score'], 3)} TPS/GiB Peak\n"
            )
            file.write(f"- Output Chars: {row['Output_Chars']}\n")
            if row["Error"]:
                file.write(f"- Error: {row['Error']}\n")
            file.write("\n```text\n")
            file.write(normalize_output_text(row["Output_Text"]))
            file.write("\n```\n")

    return report_path


def main():
    config = interactive_config()
    if not config:
        return

    results_df = run_bench(config)
    if results_df.empty:
        print("No benchmark rows were produced.")
        return

    ok_count = int((results_df["Status"] == "ok").sum())
    warning_count = int((results_df["Status"] == "warning").sum())
    error_count = int((results_df["Status"] == "error").sum())
    capability = config.get("capability", "chat")
    report_dir = ensure_report_output_dir()
    report_stem = report_dir / f"bench_{config['backend']}_{capability}_{time.strftime('%Y%m%d_%H%M%S')}"

    raw_outputs_path = save_raw_outputs(results_df, report_stem)

    print("\n" + "=" * 62)
    summary_df = build_summary_dataframe(results_df)
    console_columns = ["Run", "Status"]
    if "Capability" in summary_df.columns:
        console_columns.append("Capability")
    console_columns.extend(
        [
            "Output Category",
            "Finish Reason",
            "Model",
            "TPS (chunk/s)",
            "TTFT (s)",
            "First Event (s)",
            "Chunks (content/total)",
            "Config",
        ]
    )
    console_df = summary_df[console_columns]
    print(dataframe_to_text_table(console_df))
    print(f"\nResult counts: ok={ok_count}, warning={warning_count}, error={error_count}")

    report_path = None
    chart_path = None
    summary_excel_path = None

    try:
        summary_excel_path = save_summary_excel_workbook(results_df, config, report_stem)
    except Exception as exc:
        print(f"Summary Excel export failed: {exc}")

    try:
        report_path = save_markdown_report(
            results_df,
            config,
            report_stem,
            summary_excel_path=summary_excel_path,
        )
    except Exception as exc:
        print(f"Report generation failed: {exc}")

    try:
        chart_path = plot_results(results_df, f"{report_stem}.png", capability=capability)
    except Exception as exc:
        print(f"Chart generation failed: {exc}")

    try:
        best_config_artifacts = export_best_config(results_df, config, output_dir=report_dir)
    except Exception as exc:
        best_config_artifacts = None
        print(f"best_config export failed: {exc}")

    print(f"\nSaved artifacts directory: {report_dir}")
    if report_path:
        print(f"\nSaved report: {report_path}")
    else:
        print("\nMarkdown report was not saved.")
    if summary_excel_path:
        print(f"Saved summary Excel: {summary_excel_path}")

    print(f"Saved raw outputs: {raw_outputs_path}")
    if chart_path:
        print(f"Saved chart: {chart_path}")
    else:
        print("Skipped chart output because there were no eligible successful results.")
    if best_config_artifacts:
        print(f"Saved best config: {best_config_artifacts['best_config_path']}")
        if best_config_artifacts["modelfile_path"]:
            print(f"Saved Modelfile suggestion: {best_config_artifacts['modelfile_path']}")


def plot_results(df, output_path, capability="chat"):
    if df.empty:
        return None

    plot_df, used_fallback = select_plot_dataframe(df, capability=capability)
    plot_df = plot_df.copy()

    def build_label(row):
        label = f"{row['Model']} | {row['Config_Str']}"
        return label if len(label) <= 42 else label[:39] + "..."

    def build_run_color(row):
        if row["Output_Category"] == "tool_call":
            return "seagreen"
        if row["Status"] == "ok":
            return "skyblue"
        if row["Status"] == "warning":
            return "darkorange"
        return "indianred"

    plot_df["Plot_Label"] = plot_df.apply(build_label, axis=1)
    plot_df["Plot_Color"] = plot_df.apply(build_run_color, axis=1)

    if capability == "tools":
        first_event_values = plot_df["First_Event_s"].fillna(0)
        duration_values = plot_df["Stream_Duration_s"].fillna(0)
        outcome_df = build_outcome_summary_dataframe(plot_df)

        fig, axes = plt.subplots(3, 1, figsize=(14, 15))

        axes[0].bar(plot_df["Plot_Label"], first_event_values, color=plot_df["Plot_Color"])
        if plot_df["First_Event_s"].notna().any():
            axes[0].axhline(
                y=plot_df["First_Event_s"].min(),
                color="green",
                linestyle="--",
                alpha=0.3,
            )
        axes[0].set_title("Tools Benchmark First Event Comparison")
        axes[0].set_ylabel("First Event (s)")

        axes[1].bar(plot_df["Plot_Label"], duration_values, color=plot_df["Plot_Color"])
        if plot_df["Stream_Duration_s"].notna().any():
            axes[1].axhline(
                y=plot_df["Stream_Duration_s"].min(),
                color="green",
                linestyle="--",
                alpha=0.3,
            )
        axes[1].set_title("Tools Benchmark Stream Duration Comparison")
        axes[1].set_ylabel("Stream Duration (s)")

        axes[2].bar(outcome_df["Output Category"], outcome_df["Count"], color="steelblue")
        axes[2].set_title("Tools Outcome Category Counts")
        axes[2].set_ylabel("Runs")
        axes[2].set_xlabel("Output Category")

        axes[0].tick_params(axis="x", rotation=35)
        axes[1].tick_params(axis="x", rotation=35)
        axes[2].tick_params(axis="x", rotation=20)
    else:
        eligible_df = filter_eligible_results(plot_df, capability=capability)
        if not eligible_df.empty:
            max_tps = eligible_df["TPS"].max()
            min_ttft = eligible_df["TTFT"].min()
            colors = ["gold" if tps == max_tps else "skyblue" for tps in eligible_df["TPS"]]
            ttft_colors = [
                "lightgreen" if ttft == min_ttft else "salmon" for ttft in eligible_df["TTFT"]
            ]
            has_vram_data = eligible_df["VRAM_Peak_MiB"].notna().any()
            has_efficiency_data = eligible_df["Efficiency_Score"].notna().any()

            subplot_count = 4 if has_efficiency_data else 3 if has_vram_data else 2
            fig, axes = plt.subplots(subplot_count, 1, figsize=(13, 5 * subplot_count), sharex=True)
            if subplot_count == 1:
                axes = [axes]

            axes[0].bar(eligible_df["Plot_Label"], eligible_df["TPS"], color=colors)
            axes[0].axhline(y=max_tps, color="red", linestyle="--", alpha=0.3)
            axes[0].set_title("Throughput Comparison")
            axes[0].set_ylabel("TPS (chunk/s)")

            axes[1].bar(eligible_df["Plot_Label"], eligible_df["TTFT"], color=ttft_colors)
            axes[1].axhline(y=min_ttft, color="green", linestyle="--", alpha=0.3)
            axes[1].set_title("First Token Latency Comparison")
            axes[1].set_ylabel("TTFT (s)")

            if has_vram_data:
                min_vram_peak = eligible_df["VRAM_Peak_MiB"].min()
                vram_colors = [
                    "lightgreen" if peak == min_vram_peak else "mediumpurple"
                    for peak in eligible_df["VRAM_Peak_MiB"]
                ]
                axes[2].bar(eligible_df["Plot_Label"], eligible_df["VRAM_Peak_MiB"], color=vram_colors)
                axes[2].axhline(y=min_vram_peak, color="green", linestyle="--", alpha=0.3)
                axes[2].set_title("VRAM Peak Comparison")
                axes[2].set_ylabel("VRAM Peak (MiB)")
                if has_efficiency_data:
                    max_efficiency = eligible_df["Efficiency_Score"].max()
                    efficiency_colors = [
                        "gold" if score == max_efficiency else "steelblue"
                        for score in eligible_df["Efficiency_Score"]
                    ]
                    axes[3].bar(
                        eligible_df["Plot_Label"],
                        eligible_df["Efficiency_Score"],
                        color=efficiency_colors,
                    )
                    axes[3].axhline(y=max_efficiency, color="orange", linestyle="--", alpha=0.3)
                    axes[3].set_title("Efficiency Score Comparison")
                    axes[3].set_ylabel("TPS/GiB Peak")
                    axes[3].set_xlabel("Model | Config")
                else:
                    axes[2].set_xlabel("Model | Config")
            else:
                axes[1].set_xlabel("Model | Config")
        else:
            status_df = (
                plot_df["Status"]
                .fillna("unknown")
                .value_counts(dropna=False)
                .rename_axis("Status")
                .reset_index(name="Count")
            )
            outcome_df = build_outcome_summary_dataframe(plot_df)
            fig, axes = plt.subplots(2, 1, figsize=(13, 10))

            axes[0].bar(status_df["Status"], status_df["Count"], color="steelblue")
            axes[0].set_title("Run Status Counts")
            axes[0].set_ylabel("Runs")

            axes[1].bar(outcome_df["Output Category"], outcome_df["Count"], color="slategray")
            title = "Outcome Category Counts"
            if used_fallback:
                title += " (Fallback)"
            axes[1].set_title(title)
            axes[1].set_ylabel("Runs")
            axes[1].set_xlabel("Output Category")
            axes[0].tick_params(axis="x", rotation=20)
            axes[1].tick_params(axis="x", rotation=20)

    plt.xticks(rotation=35, ha="right")
    plt.tight_layout()
    fig.savefig(output_path)
    plt.close(fig)

    return Path(output_path)


def normalize_reasoning_content(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, list):
        return "".join(normalize_reasoning_content(item) for item in value)
    if isinstance(value, dict):
        preferred_keys = (
            "text",
            "content",
            "reasoning",
            "reasoning_content",
            "thinking",
            "summary",
            "output_text",
        )
        parts = []
        for key in preferred_keys:
            if key in value:
                text = normalize_reasoning_content(value[key])
                if text:
                    parts.append(text)
        if parts:
            return "".join(parts)
        return "".join(normalize_reasoning_content(item) for item in value.values())
    return str(value)


def build_retained_sections(thinking_text, dialogue_output_text):
    retained_sections = []
    if (thinking_text or "").strip():
        retained_sections.append("thinking")
    if (dialogue_output_text or "").strip():
        retained_sections.append("dialogue_output")
    return retained_sections


def normalize_non_content_type(field_name):
    field_map = {
        "role": "role",
        "tool_calls": "tool_calls",
        "reasoning": "reasoning",
        "reasoning_content": "reasoning",
        "thinking": "reasoning",
        "refusal": "refusal",
        "audio": "audio",
        "function_call": "tool_calls",
    }
    return field_map.get(field_name, "other")


def inspect_stream_chunk(chunk):
    chunk_info = {
        "content": "",
        "thinking": "",
        "non_content_types": [],
        "finish_reason": None,
    }

    choices = getattr(chunk, "choices", None) or []
    if not choices:
        return chunk_info

    choice = choices[0]
    delta_payload = extract_delta_payload(getattr(choice, "delta", None))
    chunk_info["content"] = normalize_text_content(delta_payload.pop("content", None))

    reasoning_parts = []
    for reasoning_key in ("reasoning", "reasoning_content", "thinking"):
        reasoning_value = delta_payload.pop(reasoning_key, None)
        if reasoning_value is not None:
            reasoning_parts.append(normalize_reasoning_content(reasoning_value))
    chunk_info["thinking"] = "".join(reasoning_parts)

    chunk_info["non_content_types"] = sorted(
        {normalize_non_content_type(field_name) for field_name in delta_payload}
        | ({"reasoning"} if chunk_info["thinking"] else set())
    )
    chunk_info["finish_reason"] = getattr(choice, "finish_reason", None) or None
    return chunk_info


def build_result_row(
    run_id,
    config,
    model,
    param_set,
    applied_params,
    display_params,
    classification,
    vram_metrics,
    dialogue_output_text,
    thinking_text,
    error_message,
    system_prompt_label="N/A",
    system_prompt_text="",
):
    efficiency_score = calculate_efficiency_score(classification["TPS"], vram_metrics["VRAM_Peak_MiB"])
    retained_sections = build_retained_sections(thinking_text, dialogue_output_text)
    thinking_chars = classification.get("Thinking_Chars", len(thinking_text))
    output_chars = classification.get("Output_Chars", len(dialogue_output_text))
    thinking_tokens = estimate_token_count(thinking_text)
    output_tokens = estimate_token_count(dialogue_output_text)
    thinking_tps = calculate_tps_from_duration(thinking_tokens, classification.get("Thinking_Time_s"))
    output_tps = calculate_tps_from_duration(output_tokens, classification.get("Output_Time_s"))
    return {
        "Run_ID": run_id,
        "Status": classification["Status"],
        "Capability": config.get("capability", "chat"),
        "Output_Category": classification["Output_Category"],
        "Diagnosis": classification["Diagnosis"],
        "Finish_Reason": classification["Finish_Reason"],
        "Backend": config["backend"],
        "Model": model,
        "System_Prompt_Label": system_prompt_label,
        "System_Prompt_Text": system_prompt_text,
        "Thinking_Mode": get_thinking_mode_for_run(param_set),
        "Params": param_set.copy(),
        "Applied_Params": applied_params.copy(),
        "Config_Str": display_params,
        "TPS": classification["TPS"],
        "TTFT": classification["TTFT"],
        "First_Event_s": classification["First_Event_s"],
        "Stream_Duration_s": classification["Stream_Duration_s"],
        "Total_Chunks": classification["Total_Chunks"],
        "Content_Chunks": classification["Content_Chunks"],
        "Non_Content_Chunks": classification["Non_Content_Chunks"],
        "Non_Content_Types": classification["Non_Content_Types"],
        "VRAM_Base_MiB": vram_metrics["VRAM_Base_MiB"],
        "VRAM_Peak_MiB": vram_metrics["VRAM_Peak_MiB"],
        "VRAM_Delta_MiB": vram_metrics["VRAM_Delta_MiB"],
        "VRAM_Detail": vram_metrics["VRAM_Detail"],
        "Efficiency_Score": efficiency_score,
        "Thinking_Time_s": classification.get("Thinking_Time_s"),
        "Thinking_TPS": thinking_tps,
        "Output_Time_s": classification.get("Output_Time_s"),
        "Output_TPS": output_tps,
        "Output_Thinking_Ratio": classification.get("Output_Thinking_Ratio"),
        "Total_Output_Time_s": classification.get("Total_Output_Time_s"),
        "Retained_Sections": retained_sections,
        "Thinking_Chars": thinking_chars,
        "Thinking_Tokens": thinking_tokens,
        "Thinking_Text": thinking_text,
        "Dialogue_Output_Chars": output_chars,
        "Dialogue_Output_Tokens": output_tokens,
        "Dialogue_Output_Text": dialogue_output_text,
        "Output_Chars": output_chars,
        "Output_Tokens": output_tokens,
        "Output_Text": dialogue_output_text,
        "Error": error_message or "",
    }


def run_bench(config):
    client = OpenAI(base_url=config["url"], api_key="sk-no-key-needed")
    capability = config.get("capability", "chat")
    capability_label = {"chat": "chat", "tools": "tools"}.get(capability, capability)
    param_keys = list(config["params"].keys())
    param_values = [config["params"][key] for key in param_keys]
    combos = [dict(zip(param_keys, combo)) for combo in product(*param_values)] if param_keys else [{}]
    system_prompt_variants = build_system_prompt_variants(config.get("system_prompts", []))
    include_system_prompt_in_label = len(system_prompt_variants) > 1 or system_prompt_variants[0]["label"] != "N/A"

    results = []
    total_runs = len(config["models"]) * len(combos) * len(system_prompt_variants)
    vram_monitoring_enabled = query_nvidia_vram_snapshot() is not None
    config["vram_monitoring"] = "nvidia-smi" if vram_monitoring_enabled else "unavailable"

    print(f"\nStarting benchmark / 開始 benchmark，共 {total_runs} 次執行。Mode / 模式: {capability_label}")
    print(
        "Chat mode focuses on TPS/TTFT. Tools mode focuses on tool_call success and first event latency. / "
        "Chat 模式主要看 TPS 與 TTFT；Tools 模式主要看 tool_call 成功率與首事件延遲。"
    )
    print(f"System prompt variants / System prompt 變體數: {len(system_prompt_variants)}")
    if vram_monitoring_enabled:
        print("VRAM monitoring / VRAM 監控: enabled via nvidia-smi / 已透過 nvidia-smi 啟用")
    else:
        print("VRAM monitoring / VRAM 監控: nvidia-smi not available, VRAM fields will be N/A / 未偵測到 nvidia-smi，VRAM 欄位將顯示 N/A")

    run_index = 0
    for model in config["models"]:
        for system_prompt_variant in system_prompt_variants:
            for param_set in combos:
                run_index += 1
                applied_params = build_backend_options(config["backend"], param_set)
                display_params = format_param_dict(param_set)
                if include_system_prompt_in_label:
                    display_params = f"{display_params} | system_prompt={system_prompt_variant['label']}"
                print(
                    f"[{run_index}/{total_runs}] {model} | {display_params}"
                )

                request_kwargs = {}
                extra_body = build_backend_extra_body(config["backend"], param_set)
                if extra_body:
                    request_kwargs["extra_body"] = extra_body

                start_time = time.time()
                first_event_time = None
                first_content_time = None
                first_thinking_time = None
                dialogue_output_parts = []
                thinking_parts = []
                chunk_records = []
                vram_monitor = NvidiaVRAMMonitor() if vram_monitoring_enabled else None
                vram_metrics = empty_vram_metrics()
                if vram_monitor is not None and not vram_monitor.start():
                    vram_monitor = None

                try:
                    request_payload = build_chat_request_payload(
                        config,
                        model,
                        request_kwargs,
                        system_prompt_text=system_prompt_variant["text"],
                    )
                    stream = client.chat.completions.create(**request_payload)

                    for chunk in stream:
                        event_time = time.time()
                        if first_event_time is None:
                            first_event_time = event_time

                        chunk_info = inspect_stream_chunk(chunk)
                        chunk_records.append(chunk_info)

                        if chunk_info["content"] and first_content_time is None:
                            first_content_time = event_time
                        if chunk_info["thinking"] and first_thinking_time is None:
                            first_thinking_time = event_time
                        if chunk_info["content"]:
                            dialogue_output_parts.append(chunk_info["content"])
                        if chunk_info["thinking"]:
                            thinking_parts.append(chunk_info["thinking"])

                    end_time = time.time()
                    if vram_monitor is not None:
                        vram_metrics = vram_monitor.stop()
                    dialogue_output_text = "".join(dialogue_output_parts)
                    thinking_text = "".join(thinking_parts)
                    classification = classify_stream_result(
                        chunk_records=chunk_records,
                        start_time=start_time,
                        end_time=end_time,
                        first_event_time=first_event_time,
                        first_content_time=first_content_time,
                        first_thinking_time=first_thinking_time,
                        error_message=None,
                    )
                    classification = adjust_classification_for_capability(classification, capability)
                    results.append(
                        build_result_row(
                            run_id=run_index,
                            config=config,
                            model=model,
                            param_set=param_set,
                            applied_params=applied_params,
                            display_params=display_params,
                            system_prompt_label=system_prompt_variant["label"],
                            system_prompt_text=system_prompt_variant["text"],
                            classification=classification,
                            vram_metrics=vram_metrics,
                            dialogue_output_text=dialogue_output_text,
                            thinking_text=thinking_text,
                            error_message=None,
                        )
                    )
                except Exception as exc:
                    end_time = time.time()
                    if vram_monitor is not None:
                        vram_metrics = vram_monitor.stop()
                    dialogue_output_text = "".join(dialogue_output_parts)
                    thinking_text = "".join(thinking_parts)
                    classification = classify_stream_result(
                        chunk_records=chunk_records,
                        start_time=start_time,
                        end_time=end_time,
                        first_event_time=first_event_time,
                        first_content_time=first_content_time,
                        first_thinking_time=first_thinking_time,
                        error_message=str(exc),
                    )
                    classification = adjust_classification_for_capability(classification, capability)
                    print(f"Error: {exc}")
                    results.append(
                        build_result_row(
                            run_id=run_index,
                            config=config,
                            model=model,
                            param_set=param_set,
                            applied_params=applied_params,
                            display_params=display_params,
                            system_prompt_label=system_prompt_variant["label"],
                            system_prompt_text=system_prompt_variant["text"],
                            classification=classification,
                            vram_metrics=vram_metrics,
                            dialogue_output_text=dialogue_output_text,
                            thinking_text=thinking_text,
                            error_message=str(exc),
                        )
                    )

    return pd.DataFrame(results)


def save_markdown_report(df, config, report_stem, summary_excel_path=None):
    report_path = Path(f"{report_stem}.html")
    summary_df = build_summary_dataframe(df)
    outcome_summary_df = build_outcome_summary_dataframe(df)
    capability = config.get("capability", "chat")
    system_prompt_variants = build_system_prompt_variants(config.get("system_prompts", []))
    localized_summary_df = localize_report_dataframe(summary_df)
    localized_outcome_summary_df = localize_report_dataframe(outcome_summary_df)
    tool_call_success_summary_df = (
        build_tool_call_success_summary_dataframe(df) if capability == "tools" else pd.DataFrame()
    )
    localized_tool_call_success_summary_df = localize_report_dataframe(tool_call_success_summary_df)
    summary_download_href = Path(summary_excel_path).name if summary_excel_path else None

    matrix_rows = [
        (bilingual_text("Backend", "後端"), config["backend"]),
        (bilingual_text("Capability", "能力模式"), localize_capability_value(capability)),
        (bilingual_text("Base URL", "基礎網址"), config.get("url", "N/A")),
        (bilingual_text("Models", "模型"), ", ".join(config.get("models", []))),
        (bilingual_text("Prompt", "使用者提示"), config.get("prompt", "")),
        (
            bilingual_text("System Prompt Count", "系統提示數量"),
            len(system_prompt_variants),
        ),
        (
            bilingual_text("System Prompt Variants", "系統提示變體"),
            ", ".join(variant["label"] for variant in system_prompt_variants),
        ),
        (
            bilingual_text("Note", "備註"),
            "TPS is estimated from streaming content chunks, while Thinking TPS and Output TPS use estimated token counts from retained text. / "
            "TPS 以帶文字內容的串流片段估算；Thinking TPS 與 Output TPS 則以保留文字的估算 token 數計算。",
        ),
    ]
    if capability == "tools":
        matrix_rows.append(
            (
                bilingual_text("Tool Mode Note", "工具模式備註"),
                "Successful tool-calling runs may not emit text tokens, so `TPS` and `TTFT` can be `N/A`; "
                "focus on `Output Category=tool_call` and `First Event (s)`. / "
                "工具呼叫成功時可能不會輸出文字，因此 `TPS` 和 `TTFT` 可能是 `N/A`；"
                "請優先看 `Output Category=tool_call` 與 `First Event (s)`。",
            )
        )

    environment_notes = [
        f"VRAM monitoring: {config.get('vram_monitoring', 'unavailable')} / "
        f"顯存監控: {config.get('vram_monitoring', 'unavailable')}"
    ]
    metric_notes = [
        "`TPS (chunk/s)`: Estimated throughput from text-bearing streaming chunks. / 以含文字的串流片段估算輸出速度。",
        "`Total Output (chars)`: Total visible dialogue output character count retained for the run. / 本次保留的可見回覆總字數。",
        "`Thinking Time (s)`: Time from the first thinking payload to the first output chunk, or to stream end if no output chunk arrived. / 從第一段 thinking 到第一段 output 的時間；若沒有 output，則到串流結束。",
        "`Output Time (s)`: Time from the first output text chunk to stream end. / 從第一段輸出文字到串流結束的時間。",
        "`Total Output Time (s)`: Time from the earliest thinking/output payload, or first stream event fallback, to stream end. / 從最早的 thinking 或 output 開始計時；若兩者都沒有，則退回第一個串流事件到結束的時間。",
        "`Thinking TPS (token/s)`: Estimated thinking tokens divided by `Thinking Time (s)`. / 估算的 thinking token 數除以 `Thinking Time (s)`。",
        "`Output TPS (token/s)`: Estimated dialogue output tokens divided by `Output Time (s)`. / 估算的最終回覆 token 數除以 `Output Time (s)`。",
        "`Output/Thinking Ratio`: `Dialogue Output Chars / Thinking Chars`; higher means more visible answer text per retained thinking text. / 回覆字數除以 thinking 字數，越高代表可見答案佔比越高。",
        "`TTFT (s)`: Time to first text chunk. / 首段文字輸出的延遲。",
        "`First Event (s)`: Time to the first streamed event of any kind. / 第一個串流事件出現的延遲。",
        "`VRAM Peak (MiB)`: Highest observed total NVIDIA GPU memory usage during a run. / 單次測試觀測到的最高總顯存使用量。",
        "`Efficiency Score (TPS/GiB Peak)`: `TPS / (VRAM Peak in GiB)` when both values are available. / 兩者都有值時，以 `TPS / 顯存峰值 GiB` 計算效率。",
    ]
    retained_text_notes = [
        "`thinking`: Reasoning/thinking text captured from non-content reasoning payloads when the backend exposed them. / 後端有提供時，從 reasoning 類 payload 保留的思考文字。",
        "`dialogue_output`: Final conversational text emitted in normal content chunks. / 一般 content chunk 中輸出的最終對話文字。",
        "If neither exists for a run, the retained sections list will be `none`. / 若兩者都沒有，保留欄位會顯示 `none`。",
    ]
    output_diagnosis_notes = [
        "`normal_content`: Received text output as expected for chat benchmarking. / 收到正常文字輸出。",
        "`tool_call`: Received `tool_calls` payload as expected for tool benchmarking. / 收到工具呼叫 payload。",
        "`text_reply_without_tool`: Returned text, but did not emit any tool call in tool mode. / 工具模式下只回文字、沒有呼叫工具。",
        "`empty_reply`: Stream completed without text output. / 串流結束但沒有文字輸出。",
        "`non_content_stream`: Stream only carried non-text payloads. / 串流只有非文字 payload。",
        "`early_stop`: Stream ended before a complete reply or tool call was received. / 在完整回覆或工具呼叫前就中斷。",
    ]

    page_parts = [
        "<!DOCTYPE html>",
        '<html lang="zh-Hant">',
        "<head>",
        '<meta charset="utf-8">',
        '<meta name="viewport" content="width=device-width, initial-scale=1">',
        "<title>Benchmark Report / 基準測試報告</title>",
        """<style>
body {
    margin: 0;
    background: #f4efe6;
    color: #1f2933;
    font-family: "Segoe UI", "Noto Sans TC", sans-serif;
    line-height: 1.6;
}
.page {
    max-width: 1500px;
    margin: 0 auto;
    padding: 32px 24px 64px;
}
h1, h2, h3, h4 {
    margin: 0;
    color: #13212b;
}
h1 {
    font-size: 2.1rem;
    margin-bottom: 8px;
}
h2 {
    font-size: 1.3rem;
    margin-bottom: 14px;
}
h3 {
    font-size: 1.05rem;
    margin-bottom: 12px;
}
h4 {
    font-size: 0.98rem;
    margin: 14px 0 8px;
}
p {
    margin: 0;
}
.lead {
    color: #566370;
    margin-bottom: 24px;
}
.section {
    background: #fffdfa;
    border: 1px solid #dccfbb;
    border-radius: 18px;
    padding: 22px 24px;
    margin-top: 20px;
    box-shadow: 0 10px 28px rgba(76, 61, 36, 0.06);
}
.section-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 12px;
    flex-wrap: wrap;
    margin-bottom: 14px;
}
.table-wrap {
    overflow-x: auto;
}
table {
    width: 100%;
    border-collapse: collapse;
    table-layout: auto;
}
th,
td {
    padding: 10px 12px;
    border-bottom: 1px solid #e8dece;
    text-align: left;
    vertical-align: top;
}
th {
    background: #f6efe2;
    color: #263746;
    font-weight: 700;
    white-space: nowrap;
}
.matrix-table th:first-child,
.kv-table th:first-child,
.kv-table td:first-child {
    width: 220px;
}
.note-list {
    margin: 0;
    padding-left: 20px;
}
.note-list li + li {
    margin-top: 8px;
}
.run-grid {
    display: grid;
    gap: 16px;
}
.run-card {
    border: 1px solid #e5dac8;
    border-radius: 16px;
    padding: 18px;
    background: #fff;
}
.text-block {
    background: #17212b;
    color: #f8f4ed;
    padding: 16px;
    border-radius: 14px;
    overflow-x: auto;
    white-space: pre-wrap;
    word-break: break-word;
    font-family: "Cascadia Code", "Consolas", monospace;
    font-size: 0.92rem;
}
.empty-note {
    color: #6f7c88;
    font-style: italic;
}
.empty-cell {
    color: #6f7c88;
    text-align: center;
}
.download-button {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    min-height: 40px;
    padding: 0 14px;
    border-radius: 999px;
    border: 1px solid #b88a44;
    background: #d9a84c;
    color: #1f2933;
    text-decoration: none;
    font-weight: 700;
    white-space: nowrap;
}
.download-button:hover {
    background: #e4b45e;
}
.filter-panel {
    border: 1px solid #e5dac8;
    border-radius: 16px;
    padding: 16px 18px;
    background: #fff7eb;
    margin-bottom: 18px;
}
.filter-toolbar {
    display: flex;
    justify-content: space-between;
    align-items: center;
    gap: 12px;
    flex-wrap: wrap;
    margin-bottom: 14px;
}
.filter-summary {
    color: #566370;
    font-weight: 600;
}
.filter-groups {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 12px;
}
.filter-group {
    margin: 0;
    padding: 10px 12px 12px;
    border: 1px solid #dccfbb;
    border-radius: 14px;
    min-width: 0;
}
.filter-group legend {
    font-weight: 700;
    color: #263746;
    padding: 0 6px;
}
.filter-option {
    display: flex;
    align-items: center;
    gap: 8px;
    margin-top: 8px;
    font-size: 0.95rem;
}
.filter-reset-button {
    min-height: 36px;
    padding: 0 12px;
    border-radius: 999px;
    border: 1px solid #c6a56a;
    background: #fffdfa;
    color: #263746;
    font-weight: 700;
    cursor: pointer;
}
.filter-reset-button:hover {
    background: #f6efe2;
}
.run-badges {
    display: flex;
    flex-wrap: wrap;
    gap: 8px;
    margin-bottom: 14px;
}
.run-badge {
    display: inline-flex;
    align-items: center;
    min-height: 28px;
    padding: 0 10px;
    border-radius: 999px;
    background: #f6efe2;
    border: 1px solid #e5dac8;
    color: #495866;
    font-size: 0.88rem;
    font-weight: 600;
}
        </style>""",
        "</head>",
        "<body>",
        '<main class="page">',
        f"<h1>{bilingual_text('Benchmark Report', '基準測試報告')}</h1>",
        '<p class="lead">Same benchmark content, rendered as HTML for clearer sections and more stable table layout. / 保留原本 benchmark 內容，只將排版改成較穩定且較容易閱讀的 HTML 報告。</p>',
        '<section class="section">',
        f"<h2>{bilingual_text('Test Matrix', '測試矩陣')}</h2>",
        '<div class="table-wrap">',
        key_value_rows_to_html_table(matrix_rows, table_class="matrix-table"),
        "</div>",
        "</section>",
        '<section class="section">',
        f"<h2>{bilingual_text('System Prompt Variants', '系統提示變體')}</h2>",
    ]

    if system_prompt_variants[0]["label"] == "N/A" and len(system_prompt_variants) == 1:
        page_parts.append(
            '<p class="empty-note">No extra system prompt was used. / 本次未額外加入 system prompt。</p>'
        )
    else:
        page_parts.append('<div class="run-grid">')
        for variant in system_prompt_variants:
            page_parts.extend(
                [
                    '<article class="run-card">',
                    f"<h3>{bilingual_text(variant['label'], '系統提示變體')}</h3>",
                    f'<pre class="text-block">{html_escape_text(normalize_output_text(variant["text"]))}</pre>',
                    "</article>",
                ]
            )
        page_parts.append("</div>")

    page_parts.extend(
        [
        "</section>",
        '<section class="section">',
        f"<h2>{bilingual_text('Environment Notes', '環境說明')}</h2>",
        bullet_list_to_html(environment_notes),
        "</section>",
        '<section class="section">',
        f"<h2>{bilingual_text('Metric Notes', '指標說明')}</h2>",
        bullet_list_to_html(metric_notes),
        "</section>",
        '<section class="section">',
        f"<h2>{bilingual_text('Retained Text Notes', '保留文字說明')}</h2>",
        bullet_list_to_html(retained_text_notes),
        "</section>",
        '<section class="section">',
        f"<h2>{bilingual_text('Output Diagnosis Notes', '輸出診斷說明')}</h2>",
        bullet_list_to_html(output_diagnosis_notes),
        "</section>",
        '<section class="section">',
        '<div class="section-header">',
        f"<h2>{bilingual_text('Summary', '摘要')}</h2>",
        build_download_button(
            summary_download_href,
            "Download Excel / 下載 Excel 摘要",
        ),
        "</div>",
        '<div class="table-wrap">',
        dataframe_to_html_table(localized_summary_df),
        "</div>",
        "</section>",
        '<section class="section">',
        f"<h2>{bilingual_text('Outcome Summary', '結果摘要')}</h2>",
        '<div class="table-wrap">',
        dataframe_to_html_table(localized_outcome_summary_df),
        "</div>",
        "</section>",
    ]
    )

    if capability == "tools" and not tool_call_success_summary_df.empty:
        page_parts.extend(
            [
                '<section class="section">',
                f"<h2>{bilingual_text('Tool Call Success by Model', '各模型工具呼叫成功統計')}</h2>",
                '<div class="table-wrap">',
                dataframe_to_html_table(localized_tool_call_success_summary_df),
                "</div>",
                "</section>",
            ]
        )

    page_parts.extend(
        [
            '<section class="section">',
            f"<h2>{bilingual_text('Generated Outputs', '各次執行輸出')}</h2>",
            build_run_filter_panel_html(),
            '<div class="run-grid" id="run-card-grid">',
        ]
    )

    for _, row in df.iterrows():
        params_json = json.dumps(row["Params"], ensure_ascii=False)
        applied_params_json = json.dumps(row["Applied_Params"], ensure_ascii=False)
        retained_sections = row.get("Retained_Sections", []) or []
        retained_sections_label = (
            ", ".join(
                {
                    "thinking": bilingual_text("thinking", "思考內容"),
                    "dialogue_output": bilingual_text("dialogue_output", "對話輸出"),
                }.get(section, section)
                for section in retained_sections
            )
            if retained_sections
            else bilingual_text("none", "無")
        )
        thinking_text = row.get("Thinking_Text", "")
        dialogue_output_text = row.get("Dialogue_Output_Text", row.get("Output_Text", ""))
        system_prompt_text = row.get("System_Prompt_Text", "")
        thinking_mode = resolve_thinking_mode(row.get("Thinking_Mode", "default"))
        has_tool_call = (
            row.get("Output_Category") == "tool_call"
            or "tool_calls" in parse_non_content_types(row.get("Non_Content_Types"))
        )
        has_thinking = bool((thinking_text or "").strip())
        has_output = bool((dialogue_output_text or "").strip())
        run_badges = [
            bilingual_text("Tool Call", "工具呼叫")
            + ": "
            + bilingual_text("yes" if has_tool_call else "no", "有" if has_tool_call else "無"),
            bilingual_text("Thinking", "Thinking 內容")
            + ": "
            + bilingual_text("yes" if has_thinking else "no", "有" if has_thinking else "無"),
            bilingual_text("Output", "輸出內容")
            + ": "
            + bilingual_text("yes" if has_output else "no", "有" if has_output else "無"),
            bilingual_text("thinking_mode", "思考模式")
            + ": "
            + localize_thinking_mode_value(thinking_mode),
            bilingual_text("status", "狀態") + ": " + localize_status_value(row["Status"]),
        ]

        detail_rows = [
            (bilingual_text("Status", "狀態"), localize_status_value(row["Status"])),
            (bilingual_text("Capability", "能力模式"), localize_capability_value(row.get("Capability", capability))),
            (bilingual_text("Output Category", "輸出分類"), localize_output_category_value(row["Output_Category"])),
            (bilingual_text("Diagnosis", "診斷"), row["Diagnosis"]),
            (bilingual_text("Finish Reason", "結束原因"), localize_finish_reason_value(format_text_value(row["Finish_Reason"]))),
            (bilingual_text("Backend", "後端"), row["Backend"]),
            (bilingual_text("Model", "模型"), row["Model"]),
            (bilingual_text("System Prompt", "系統提示"), localize_system_prompt_label(row.get("System_Prompt_Label", "N/A"))),
            (
                bilingual_text("Thinking Mode", "思考模式"),
                localize_thinking_mode_value(row.get("Thinking_Mode", "default")),
            ),
            (bilingual_text("System Prompt Chars", "系統提示字數"), len(system_prompt_text)),
            (bilingual_text("Params", "參數"), params_json),
            (bilingual_text("Applied Params", "實際套用參數"), applied_params_json),
            (bilingual_text("Retained Sections", "保留區塊"), retained_sections_label),
            (bilingual_text("Thinking Chars", "思考字數"), int(row.get("Thinking_Chars", len(thinking_text)))),
            (bilingual_text("Thinking Tokens", "思考 token 數"), int(row.get("Thinking_Tokens", estimate_token_count(thinking_text)))),
            (
                bilingual_text("Dialogue Output Chars", "對話輸出字數"),
                int(row.get("Dialogue_Output_Chars", len(dialogue_output_text))),
            ),
            (
                bilingual_text("Dialogue Output Tokens", "對話輸出 token 數"),
                int(row.get("Dialogue_Output_Tokens", estimate_token_count(dialogue_output_text))),
            ),
            (bilingual_text("Thinking Time", "思考時間"), f"{format_numeric_value(row.get('Thinking_Time_s'), 3)} s"),
            (bilingual_text("Output Time", "回覆時間"), f"{format_numeric_value(row.get('Output_Time_s'), 3)} s"),
            (
                bilingual_text("Total Output Time", "總輸出時間"),
                f"{format_numeric_value(row.get('Total_Output_Time_s'), 3)} s",
            ),
            (bilingual_text("TPS", "輸出速率"), f"{format_numeric_value(row['TPS'], 2)} chunk/s"),
            (bilingual_text("Thinking TPS", "思考速率"), f"{format_numeric_value(row.get('Thinking_TPS'), 2)} token/s"),
            (bilingual_text("Output TPS", "回覆速率"), f"{format_numeric_value(row.get('Output_TPS'), 2)} token/s"),
            (bilingual_text("Output/Thinking Ratio", "輸出思考比"), format_numeric_value(row.get("Output_Thinking_Ratio"), 3)),
            (bilingual_text("TTFT", "首字延遲"), f"{format_numeric_value(row['TTFT'], 3)} s"),
            (bilingual_text("First Event", "首事件時間"), f"{format_numeric_value(row['First_Event_s'], 3)} s"),
            (bilingual_text("Stream Duration", "串流總時長"), f"{format_numeric_value(row['Stream_Duration_s'], 3)} s"),
            (bilingual_text("Total Chunks", "總片段數"), int(row["Total_Chunks"])),
            (bilingual_text("Content Chunks", "文字片段數"), int(row["Content_Chunks"])),
            (bilingual_text("Non-Content Chunks", "非文字片段數"), int(row["Non_Content_Chunks"])),
            (bilingual_text("Non-Content Types", "非文字類型"), row["Non_Content_Types"]),
            (bilingual_text("VRAM Base", "起始顯存"), format_mib_value(row["VRAM_Base_MiB"])),
            (bilingual_text("VRAM Peak", "顯存峰值"), format_mib_value(row["VRAM_Peak_MiB"])),
            (bilingual_text("VRAM Delta", "顯存增量"), format_mib_value(row["VRAM_Delta_MiB"])),
            (bilingual_text("VRAM Detail", "顯存細節"), row["VRAM_Detail"]),
            (
                bilingual_text("Efficiency Score", "效率分數"),
                f"{format_numeric_value(row['Efficiency_Score'], 3)} TPS/GiB Peak",
            ),
        ]
        if row["Error"]:
            detail_rows.append((bilingual_text("Error", "錯誤"), row["Error"]))

        page_parts.extend(
            [
                (
                    '<article class="run-card result-card" '
                    f'data-tool-call="{"yes" if has_tool_call else "no"}" '
                    f'data-thinking="{"yes" if has_thinking else "no"}" '
                    f'data-output="{"yes" if has_output else "no"}" '
                    f'data-thinking-mode="{thinking_mode}" '
                    f'data-status="{row["Status"]}">'
                ),
                f"<h3>{bilingual_text('Run ' + str(row['Run_ID']), f'第 {row['Run_ID']} 次執行')}</h3>",
                '<div class="run-badges">' + "".join(
                    f'<span class="run-badge">{html_escape_text(badge)}</span>' for badge in run_badges
                ) + "</div>",
                key_value_rows_to_html_table(detail_rows),
            ]
        )

        if thinking_text:
            page_parts.extend(
                [
                    f"<h4>{bilingual_text('thinking', '思考內容')}</h4>",
                    f'<pre class="text-block">{html_escape_text(normalize_output_text(thinking_text))}</pre>',
                ]
            )

        if dialogue_output_text:
            page_parts.extend(
                [
                    f"<h4>{bilingual_text('dialogue_output', '對話輸出')}</h4>",
                    f'<pre class="text-block">{html_escape_text(normalize_output_text(dialogue_output_text))}</pre>',
                ]
            )

        if not thinking_text and not dialogue_output_text:
            page_parts.append(
                '<p class="empty-note">No retained thinking or dialogue output text for this run. / 這次執行沒有保留 thinking 或 dialogue output 文字。</p>'
            )

        page_parts.append("</article>")

    page_parts.extend(["</div>", "</section>", build_run_filter_script(), "</main>", "</body>", "</html>"])

    with report_path.open("w", encoding="utf-8") as file:
        file.write("\n".join(page_parts))

    return report_path

import sys
from dataclasses import dataclass





CAPABILITY_DEFAULTS = {
    "chat": "Explain the long-term creep risk of PETG in 3D printing and how to reduce it.",
    "tools": (
        "Check today's weather in Taipei. If you support tools or function calling, "
        "call the `lookup_weather` tool first instead of answering directly."
    ),
}


TABLE_COLUMNS = (
    ("idx", 4),
    ("param_key", 18),
    ("state", 8),
    ("count", 7),
    ("values", 24),
    ("range_text", 16),
)


ALLOWED_VALUE_CHARS = set("0123456789,.-+eEabcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ_")


@dataclass
class ParamGridRow:
    key: str
    group: str
    label: str
    range_text: str
    desc: str
    supported: bool
    default_value: str
    enabled: bool = False
    raw_value: str = ""


def ordered_param_keys():
    ordered_keys = []
    for param_keys in PARAM_GROUPS.values():
        for key in param_keys:
            if key not in ordered_keys:
                ordered_keys.append(key)

    for key in PARAM_INFO:
        if key not in ordered_keys:
            ordered_keys.append(key)
    return ordered_keys


def build_param_rows(backend):
    group_by_key = {}
    for group_name, param_keys in PARAM_GROUPS.items():
        for key in param_keys:
            group_by_key[key] = group_name

    rows = []
    for key in ordered_param_keys():
        info = PARAM_INFO[key]
        supported = backend in info["backends"]
        if not supported:
            continue
        rows.append(
            ParamGridRow(
                key=key,
                group=group_by_key.get(key, "Other"),
                label=info["label"],
                range_text=info["range"],
                desc=info["desc"],
                supported=supported,
                default_value=str(info["default"]),
                raw_value=str(info["default"]),
            )
        )
    return rows


def truncate_text(text, width):
    text = str(text)
    if width <= 0:
        return ""
    if len(text) <= width:
        return text.ljust(width)
    if width == 1:
        return text[:1]
    return text[: width - 1] + "…"


def estimate_combo_count(params):
    combo_count = 1
    for values in params.values():
        combo_count *= len(values)
    return combo_count


def validate_param_rows(rows):
    final_params = {}
    for index, row in enumerate(rows):
        if not row.supported or not row.enabled:
            continue
        try:
            final_params[row.key] = parse_csv_values(row.raw_value, param_key=row.key)
        except ValueError as exc:
            return None, index, f"{row.label}: {exc}"
    return final_params, None, None


def row_value_count(row):
    if not row.supported:
        return "LOCK"
    if not row.enabled:
        return "-"

    try:
        return str(len(parse_csv_values(row.raw_value, param_key=row.key)))
    except ValueError:
        return "ERR"


def build_grid_fragments(rows, selected_row_index, selected_column_index, message, message_style, backend):
    from prompt_toolkit.formatted_text import to_formatted_text

    selected_row = rows[selected_row_index]
    preview_params, _, preview_error = validate_param_rows(rows)
    combo_count = "ERR" if preview_error else str(estimate_combo_count(preview_params or {}))
    selected_count = sum(1 for row in rows if row.supported and row.enabled)
    active_column_name = "state" if selected_column_index == 0 else "values"
    backend_label = get_backend_display_name(backend)

    fragments = []
    fragments.extend(
        to_formatted_text(
            [
                ("class:title", f"LLM Benchmark / LLM 基準測試 | {backend_label} Parameter Grid / {backend_label} 參數表\n"),
                (
                    "class:subtitle",
                    "Arrow keys move / 方向鍵移動 | Left/Right switch cell / 左右切換欄位 | "
                    "Space toggles N/A/TEST / Space 切換 N/A 或 TEST | "
                    "Type values in Values / 在 Values 欄輸入測試值 | "
                    "Backspace deletes or goes back from State / Backspace 刪字或在 State 欄返回上一階段 | "
                    "d restores default / d 恢復預設值 | Enter/Ctrl-S saves / Enter 或 Ctrl-S 儲存 | "
                    "Esc cancels / Esc 取消\n\n",
                ),
                ("class:panel.label", f"Backend View / 目前頁面: {backend_label} | Available Params / 可調參數數: {len(rows)}\n\n"),
            ]
        )
    )

    header_cells = {
        "idx": "#",
        "param_key": "Param / 參數",
        "state": "State / 狀態",
        "count": "Count / 數量",
        "values": "Values / 值",
        "range_text": "Range / 範圍",
    }
    for column_name, width in TABLE_COLUMNS:
        fragments.append(("class:table.header", truncate_text(header_cells[column_name], width)))
        fragments.append(("class:table.header", " "))
    fragments.append(("", "\n"))
    fragments.append(("class:table.rule", "-" * (sum(width for _, width in TABLE_COLUMNS) + len(TABLE_COLUMNS))))
    fragments.append(("", "\n"))

    for row_index, row in enumerate(rows):
        row_style = "class:table.row"
        if row_index == selected_row_index:
            row_style = "class:table.row.selected"

        state_text = "LOCK" if not row.supported else "TEST" if row.enabled else "N/A"
        value_text = "unsupported / 不支援" if not row.supported else row.raw_value if row.enabled else "N/A"
        cell_values = {
            "idx": f"{row_index + 1:02d}",
            "param_key": row.key,
            "state": state_text,
            "count": row_value_count(row),
            "values": value_text,
            "range_text": row.range_text,
        }

        for column_name, width in TABLE_COLUMNS:
            cell_style = row_style
            if row_index == selected_row_index and column_name == active_column_name:
                cell_style = "class:table.cell.current"
            fragments.append((cell_style, truncate_text(cell_values[column_name], width)))
            fragments.append((row_style, " "))
        fragments.append(("", "\n"))

    support_text = ", ".join(PARAM_INFO[selected_row.key]["backends"])
    status_style = "class:status.ok" if not preview_error else "class:status.error"
    fragments.extend(
        to_formatted_text(
            [
                ("", "\n"),
                ("class:panel.title", "Selected Parameter / 目前參數\n"),
                ("class:panel.label", f"Key / 參數鍵: {selected_row.key}\n"),
                ("class:panel.label", f"Label / 名稱: {selected_row.label}\n"),
                ("class:panel.label", f"Group / 群組: {selected_row.group}\n"),
                ("class:panel.label", f"Supported Backends / 支援後端: {support_text}\n"),
                ("class:panel.label", f"Default Values / 預設值: {selected_row.default_value}\n"),
                ("class:panel.label", f"Description / 說明: {selected_row.desc}\n\n"),
                ("class:panel.title", "Config Preview / 設定預覽\n"),
                ("class:panel.label", f"Selected Params / 已選參數數: {selected_count}\n"),
                ("class:panel.label", f"Combination Count / 組合數: {combo_count}\n"),
                (status_style, f"Validation / 驗證: {'OK / 正常' if not preview_error else preview_error}\n"),
            ]
        )
    )

    if message:
        fragments.append((message_style, f"\n{message}\n"))

    return fragments


def edit_param_grid(backend, initial_params=None):
    from prompt_toolkit.application import Application
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.keys import Keys
    from prompt_toolkit.layout import HSplit, Layout, Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style

    rows = build_param_rows(backend)
    for row in rows:
        if initial_params and row.key in initial_params:
            row.enabled = True
            row.raw_value = format_param_values_for_display(row.key, initial_params[row.key])
    state = {
        "row_index": 0,
        "column_index": 0,
        "message": "",
        "message_style": "class:hint",
    }

    def set_message(text, style="class:hint"):
        state["message"] = text
        state["message_style"] = style

    def refresh(event):
        if event is not None and getattr(event, "app", None):
            event.app.invalidate()

    def current_row():
        return rows[state["row_index"]]

    def move_row(delta):
        state["row_index"] = max(0, min(len(rows) - 1, state["row_index"] + delta))

    def move_column(delta):
        state["column_index"] = (state["column_index"] + delta) % 2

    def toggle_current_row():
        row = current_row()
        if not row.supported:
            set_message(
                f"{row.key} is not supported on {backend}. / {row.key} 不支援 {backend}。",
                "class:status.warning",
            )
            return

        row.enabled = not row.enabled
        if row.enabled and not row.raw_value.strip():
            row.raw_value = row.default_value
        set_message(
            f"{row.key} -> {'TEST' if row.enabled else 'N/A'} / {row.key} 已切換為 {'測試' if row.enabled else '不測試'}",
            "class:hint",
        )

    def restore_default():
        row = current_row()
        if not row.supported:
            set_message(
                f"{row.key} is locked for {backend}. / {row.key} 在 {backend} 上已鎖定。",
                "class:status.warning",
            )
            return

        row.enabled = True
        row.raw_value = row.default_value
        set_message(
            f"{row.key} restored to default values. / {row.key} 已恢復預設值。",
            "class:hint",
        )

    def append_value(char):
        row = current_row()
        if not row.supported:
            set_message(
                f"{row.key} is locked for {backend}. / {row.key} 在 {backend} 上已鎖定。",
                "class:status.warning",
            )
            return

        if char not in ALLOWED_VALUE_CHARS:
            set_message(f"Unsupported character: {char!r} / 不支援的字元：{char!r}", "class:status.warning")
            return

        if not row.enabled:
            row.enabled = True
            if row.raw_value == "N/A":
                row.raw_value = ""
        row.raw_value += char
        set_message(f"Editing {row.key} / 正在編輯 {row.key}", "class:hint")

    def backspace_value():
        row = current_row()
        if not row.supported:
            set_message(
                f"{row.key} is locked for {backend}. / {row.key} 在 {backend} 上已鎖定。",
                "class:status.warning",
            )
            return

        if not row.enabled:
            row.enabled = True
            row.raw_value = row.default_value
            set_message(
                f"{row.key} enabled with default values. / {row.key} 已啟用並帶入預設值。",
                "class:hint",
            )
            return

        row.raw_value = row.raw_value[:-1]
        set_message(f"Editing {row.key} / 正在編輯 {row.key}", "class:hint")

    def accept(event):
        params, error_row_index, error_message = validate_param_rows(rows)
        if error_message:
            state["row_index"] = error_row_index
            state["column_index"] = 1
            set_message(error_message, "class:status.error")
            refresh(event)
            return
        event.app.exit(result=params or {})

    table_control = FormattedTextControl(
        lambda: build_grid_fragments(
            rows=rows,
            selected_row_index=state["row_index"],
            selected_column_index=state["column_index"],
            message=state["message"],
            message_style=state["message_style"],
            backend=backend,
        ),
        focusable=True,
        show_cursor=False,
    )

    root_container = HSplit([Window(content=table_control, always_hide_cursor=True)])
    style = Style.from_dict(
        {
            "title": "bold ansicyan",
            "subtitle": "ansibrightblack",
            "table.header": "bold ansiyellow",
            "table.rule": "ansibrightblack",
            "table.row": "",
            "table.row.selected": "bg:ansiblue ansiwhite",
            "table.cell.current": "reverse",
            "panel.title": "bold ansigreen",
            "panel.label": "",
            "status.ok": "ansigreen",
            "status.warning": "ansiyellow",
            "status.error": "ansired",
            "hint": "ansicyan",
        }
    )

    kb = KeyBindings()

    @kb.add("up")
    def _(event):
        move_row(-1)
        refresh(event)

    @kb.add("down")
    def _(event):
        move_row(1)
        refresh(event)

    @kb.add("left")
    def _(event):
        move_column(-1)
        refresh(event)

    @kb.add("right")
    def _(event):
        move_column(1)
        refresh(event)

    @kb.add("tab")
    def _(event):
        move_column(1)
        refresh(event)

    @kb.add("s-tab")
    def _(event):
        move_column(-1)
        refresh(event)

    @kb.add("space")
    def _(event):
        if state["column_index"] == 0:
            toggle_current_row()
            refresh(event)

    @kb.add("backspace")
    @kb.add("c-h")
    def _(event):
        if state["column_index"] == 1:
            backspace_value()
            refresh(event)
        else:
            event.app.exit(result=BACK_ACTION)

    @kb.add("delete")
    def _(event):
        if state["column_index"] == 1:
            current_row().raw_value = ""
            current_row().enabled = True
            set_message(f"Cleared {current_row().key}. / 已清空 {current_row().key}。", "class:hint")
            refresh(event)

    @kb.add("d")
    def _(event):
        restore_default()
        refresh(event)

    @kb.add("enter")
    @kb.add("c-s")
    def _(event):
        accept(event)

    @kb.add("escape")
    @kb.add("c-c")
    def _(event):
        event.app.exit(result=None)

    @kb.add(Keys.Any)
    def _(event):
        if state["column_index"] != 1:
            return

        char = event.data
        if not char or char not in ALLOWED_VALUE_CHARS:
            return
        append_value(char)
        refresh(event)

    application = Application(
        layout=Layout(root_container),
        key_bindings=kb,
        full_screen=True,
        mouse_support=False,
        style=style,
    )
    return application.run()


def parse_system_prompt_blocks(raw_text, expected_count):
    normalized = (raw_text or "").replace("\r\n", "\n").strip()
    if expected_count <= 0:
        return []

    blocks = [
        block.strip()
        for block in re.split(r"(?m)^\s*---\s*$", normalized)
        if block.strip()
    ]
    if len(blocks) != expected_count:
        raise ValueError(
            f"Expected {expected_count} system prompt blocks, but found {len(blocks)}. / "
            f"預期應有 {expected_count} 段 system prompt，但目前找到 {len(blocks)} 段。"
            " Use a line containing only --- between prompts. / 請用單獨一行的 --- 分隔不同 prompt。"
        )
    return blocks


def edit_system_prompt_blocks(expected_count):
    from prompt_toolkit.application import Application
    from prompt_toolkit.key_binding import KeyBindings
    from prompt_toolkit.layout import HSplit, Layout, Window
    from prompt_toolkit.layout.controls import FormattedTextControl
    from prompt_toolkit.styles import Style
    from prompt_toolkit.widgets import Frame, TextArea

    text_area = TextArea(
        text="",
        multiline=True,
        scrollbar=True,
        line_numbers=True,
        wrap_lines=False,
        focus_on_click=True,
    )
    state = {
        "message": (
            f"Paste {expected_count} system prompt block(s). Use --- on its own line as a separator. "
            f"Ctrl-S saves, Esc cancels. / 請貼上 {expected_count} 段 system prompt，段落之間用單獨一行的 --- 分隔。"
            "Ctrl-S 儲存，Esc 取消。"
        ),
        "style": "class:hint",
    }

    def set_message(text, style):
        state["message"] = text
        state["style"] = style

    kb = KeyBindings()

    @kb.add("c-s")
    def save_editor(event):
        try:
            prompts = parse_system_prompt_blocks(text_area.text, expected_count)
        except ValueError as exc:
            set_message(str(exc), "class:status.error")
            return
        event.app.exit(result=prompts)

    @kb.add("escape")
    @kb.add("c-c")
    def cancel_editor(event):
        event.app.exit(result=None)

    @kb.add("backspace")
    @kb.add("c-h")
    def go_back(event):
        if text_area.text:
            return
        event.app.exit(result=BACK_ACTION)

    root_container = HSplit(
        [
            Window(
                height=4,
                content=FormattedTextControl(
                    lambda: [
                        ("class:title", "System Prompt Editor / 系統提示編輯器\n"),
                        (
                            "class:subtitle",
                            "Paste multiple system prompts here. Use --- on its own line as a separator.\n"
                            "可在這裡直接貼上多段 system prompt，段落之間用單獨一行的 --- 分隔。\n",
                        ),
                    ]
                ),
            ),
            Frame(text_area, title="System Prompt Blocks / 系統提示區塊"),
            Window(
                height=2,
                content=FormattedTextControl(lambda: [(state["style"], state["message"])]),
            ),
        ]
    )

    app = Application(
        layout=Layout(root_container, focused_element=text_area),
        key_bindings=kb,
        full_screen=True,
        mouse_support=True,
        style=Style.from_dict(
            {
                "title": "bold ansicyan",
                "subtitle": "ansibrightblack",
                "frame.label": "bold ansiyellow",
                "status.error": "bold ansired",
                "hint": "ansicyan",
            }
        ),
    )
    return app.run()


def merge_instruction_text(instruction, back_hint):
    if instruction:
        return f"{instruction} {back_hint}"
    return back_hint


def attach_backspace_binding(question, empty_text_only=False):
    from prompt_toolkit.filters import Condition
    from prompt_toolkit.key_binding import KeyBindings, merge_key_bindings
    from prompt_toolkit.keys import Keys

    bindings = KeyBindings()
    app = question.application

    if empty_text_only:
        active_filter = Condition(
            lambda: not getattr(getattr(app, "current_buffer", None), "text", "")
        )
    else:
        active_filter = True

    @bindings.add(Keys.Backspace, eager=True, filter=active_filter)
    @bindings.add(Keys.ControlH, eager=True, filter=active_filter)
    def go_back(event):
        event.app.exit(result=BACK_ACTION)

    app.key_bindings = merge_key_bindings([app.key_bindings, bindings])
    return question


def ask_select_with_back(message, choices, default=None, instruction=None):
    question = questionary.select(
        message,
        choices=choices,
        default=default,
        instruction=merge_instruction_text(
            instruction,
            "(Backspace: previous step / Backspace 返回上一階段)",
        ),
    )
    return attach_backspace_binding(question).ask()


def ask_checkbox_with_back(message, choices, instruction=None):
    question = questionary.checkbox(
        message,
        choices=choices,
        instruction=merge_instruction_text(
            instruction,
            "(Space: select | Backspace: previous step / Space 勾選 | Backspace 返回上一階段)",
        ),
    )
    return attach_backspace_binding(question).ask()


def ask_text_with_back(message, default="", instruction=None):
    question = questionary.text(
        message,
        default=default,
        instruction=merge_instruction_text(
            instruction,
            "(Backspace on empty input: previous step / 空白時按 Backspace 返回上一階段)",
        ),
    )
    return attach_backspace_binding(question, empty_text_only=True).ask()


def ask_confirm_with_back(message, default=True, instruction=None):
    question = questionary.confirm(
        message,
        default=default,
        instruction=merge_instruction_text(
            instruction,
            "(Enter: confirm | Backspace: previous step / Enter 確認 | Backspace 返回上一階段)",
        ),
    )
    return attach_backspace_binding(question).ask()


def select_system_prompt_variants(existing_prompts=None):
    existing_prompts = existing_prompts or []
    if not existing_prompts:
        default_selection = 0
    elif len(existing_prompts) in (1, 2, 3):
        default_selection = len(existing_prompts)
    else:
        default_selection = "custom"

    while True:
        selection = ask_select_with_back(
            "System prompt variants / 系統提示變體:",
            choices=[
                Choice("N/A | no extra system prompt / 不額外加入 system prompt", value=0),
                Choice("1 variant / 1 種", value=1),
                Choice("2 variants / 2 種", value=2),
                Choice("3 variants / 3 種", value=3),
                Choice("Custom count / 自訂數量", value="custom"),
            ],
            default=default_selection,
        )
        if selection is None:
            return None
        if selection == BACK_ACTION:
            return BACK_ACTION
        if selection == 0:
            return []

        expected_count = selection
        if selection == "custom":
            while True:
                raw_count = ask_text_with_back(
                    "How many system prompt variants? / 要測幾種 system prompt 變體？",
                    default=str(len(existing_prompts) or 1),
                )
                if raw_count is None:
                    return None
                if raw_count == BACK_ACTION:
                    break
                try:
                    expected_count = int((raw_count or "").strip())
                except ValueError:
                    print("System prompt count must be an integer. / system prompt 變體數量必須是整數。")
                    continue
                if expected_count < 0:
                    print("System prompt count cannot be negative. / system prompt 變體數量不能小於 0。")
                    continue
                if expected_count == 0:
                    return []
                break
            if raw_count == BACK_ACTION:
                continue

        prompts = edit_system_prompt_blocks(expected_count)
        if prompts is None:
            print("System prompt editor was cancelled. / system prompt 編輯已取消。")
            return None
        if prompts == BACK_ACTION:
            continue
        return prompts


def print_config_review(config):
    params = config["params"]
    combo_count = estimate_combo_count(params) if params else 1
    system_prompt_variants = build_system_prompt_variants(config.get("system_prompts", []))

    print("\n" + "=" * 62)
    print("Config Review / 設定確認")
    print("=" * 62)
    print(f"- Backend / 後端: {config['backend']}")
    print(f"- Benchmark Mode / 測試模式: {config['capability']}")
    print(f"- Base URL / 基礎 URL: {config['url']}")
    print(f"- Models / 模型: {', '.join(config['models'])}")
    print(f"- Param Count / 參數數量: {len(params)}")
    print(f"- Combination Count / 組合數: {combo_count}")
    print(f"- System Prompt Variants / System Prompt 變體數: {len(system_prompt_variants)}")
    if params:
        print("- Parameter Values / 參數值:")
        for key, values in params.items():
            print(f"  - {key}: {values}")
    else:
        print("- Parameter Values / 參數值: use backend defaults only / 僅使用後端預設值")
    if system_prompt_variants[0]["label"] == "N/A" and len(system_prompt_variants) == 1:
        print("- System Prompts / System Prompt: N/A / 未額外加入")
    else:
        print("- System Prompt Previews / System Prompt 預覽:")
        for variant in system_prompt_variants:
            preview = variant["text"].splitlines()[0] if variant["text"] else ""
            preview = preview[:90] + ("..." if len(preview) > 90 else "")
            print(f"  - {variant['label']}: {preview} ({len(variant['text'])} chars / 字元)")


def build_console_summary_dataframe(results_df):
    def console_column_name(column_name):
        return REPORT_HEADER_BILINGUAL_MAP.get(column_name, column_name).replace("<br>", " / ")

    summary_df = localize_report_dataframe(build_summary_dataframe(results_df))
    summary_df = summary_df.rename(columns=lambda column_name: column_name.replace("<br>", " / "))

    console_columns = [console_column_name("Run"), console_column_name("Status")]
    if console_column_name("Capability") in summary_df.columns:
        console_columns.append(console_column_name("Capability"))
    if console_column_name("System Prompt") in summary_df.columns:
        console_columns.append(console_column_name("System Prompt"))
    console_columns.extend(
        [
            console_column_name("Output Category"),
            console_column_name("Finish Reason"),
            console_column_name("Model"),
            console_column_name("TPS (chunk/s)"),
            console_column_name("Thinking TPS (token/s)"),
            console_column_name("Output TPS (token/s)"),
            console_column_name("Output/Thinking Ratio"),
            console_column_name("TTFT (s)"),
            console_column_name("First Event (s)"),
            console_column_name("Chunks (content/total)"),
            console_column_name("Config"),
        ]
    )
    return summary_df[console_columns]


def interactive_config():
    print("\n" + "=" * 62)
    print("LLM Benchmark / LLM 基準測試")
    print("=" * 62)
    state = {}
    stage_index = 0

    while True:
        if stage_index == 0:
            backend = ask_select_with_back(
                "Select backend / 選擇後端:",
                choices=[
                    Choice("Ollama", value="ollama"),
                    Choice("llama.cpp (llama-server)", value="llama.cpp"),
                ],
                default=state.get("backend"),
            )
            if backend in (None, BACK_ACTION):
                return None
            if backend != state.get("backend"):
                state.pop("url", None)
                state.pop("models", None)
                state.pop("params", None)
            state["backend"] = backend
            stage_index = 1
            continue

        if stage_index == 1:
            capability = ask_select_with_back(
                "Select benchmark mode / 選擇 benchmark 模式:",
                choices=[
                    Choice("Chat | Standard chat response benchmark / 一般文字回覆測試", value="chat"),
                    Choice("Tools | Check whether the model emits tool_calls / 檢查是否輸出 tool_calls", value="tools"),
                ],
                default=state.get("capability"),
            )
            if capability is None:
                return None
            if capability == BACK_ACTION:
                stage_index = 0
                continue
            state["capability"] = capability
            stage_index = 2
            continue

        if stage_index == 2:
            model_result = select_models_and_url(
                state["backend"],
                previous_url=state.get("url"),
                previous_models=state.get("models"),
            )
            if model_result is None:
                return None
            if model_result == BACK_ACTION:
                stage_index = 1
                continue
            url, models = model_result
            if not models:
                print("No models are available. Cancelled. / 沒有可用模型，已取消。")
                return None
            state["url"] = url
            state["models"] = models
            stage_index = 3
            continue

        if stage_index == 3:
            final_params = edit_param_grid(state["backend"], initial_params=state.get("params"))
            if final_params is None:
                print("Parameter grid was cancelled. / 參數表設定已取消。")
                return None
            if final_params == BACK_ACTION:
                stage_index = 2
                continue
            state["params"] = final_params
            stage_index = 4
            continue

        if stage_index == 4:
            prompt = ask_text_with_back(
                "Benchmark prompt / 測試 prompt:",
                default=state.get("prompt", CAPABILITY_DEFAULTS[state["capability"]]),
            )
            if prompt is None:
                return None
            if prompt == BACK_ACTION:
                stage_index = 3
                continue
            state["prompt"] = prompt
            stage_index = 5
            continue

        if stage_index == 5:
            system_prompts = select_system_prompt_variants(state.get("system_prompts"))
            if system_prompts is None:
                return None
            if system_prompts == BACK_ACTION:
                stage_index = 4
                continue
            state["system_prompts"] = system_prompts
            stage_index = 6
            continue

        config = {
            "backend": state["backend"],
            "capability": state["capability"],
            "url": state["url"],
            "models": state["models"],
            "params": state.get("params", {}),
            "prompt": state["prompt"],
            "system_prompts": state.get("system_prompts", []),
        }

        print_config_review(config)
        confirmed = ask_confirm_with_back(
            "Start benchmark with this configuration? / 要用這份設定開始 benchmark 嗎？",
            default=True,
        )
        if confirmed is None:
            return None
        if confirmed == BACK_ACTION:
            stage_index = 5
            continue
        if not confirmed:
            print("Cancelled before benchmark run. / 已在 benchmark 開始前取消。")
            return None
        return config


def main():
    config = interactive_config()
    if not config:
        return

    results_df = run_bench(config)
    if results_df.empty:
        print("No benchmark rows were produced.")
        return

    ok_count = int((results_df["Status"] == "ok").sum())
    warning_count = int((results_df["Status"] == "warning").sum())
    error_count = int((results_df["Status"] == "error").sum())
    capability = config.get("capability", "chat")
    report_dir = ensure_report_output_dir()
    report_stem = report_dir / f"bench_{config['backend']}_{capability}_{time.strftime('%Y%m%d_%H%M%S')}"

    raw_outputs_path = save_raw_outputs(results_df, report_stem)

    print("\n" + "=" * 62)
    console_df = build_console_summary_dataframe(results_df)
    print(dataframe_to_text_table(console_df))
    print(f"\nResult counts / 結果統計: ok={ok_count}, warning={warning_count}, error={error_count}")

    report_path = None
    chart_path = None
    summary_excel_path = None

    try:
        summary_excel_path = save_summary_excel_workbook(results_df, config, report_stem)
    except Exception as exc:
        print(f"Summary Excel export failed / Summary Excel 匯出失敗: {exc}")

    try:
        report_path = save_markdown_report(
            results_df,
            config,
            report_stem,
            summary_excel_path=summary_excel_path,
        )
    except Exception as exc:
        print(f"Report generation failed / 報告產生失敗: {exc}")

    try:
        chart_path = plot_results(results_df, f"{report_stem}.png", capability=capability)
    except Exception as exc:
        print(f"Chart generation failed / 圖表產生失敗: {exc}")

    try:
        export_best_config(results_df, config, output_dir=report_dir)
    except Exception as exc:
        print(f"best_config export failed / best_config 匯出失敗: {exc}")

    print(f"\nSaved artifacts directory / 輸出資料夾: {report_dir}")
    if report_path:
        print(f"\nSaved report / 已儲存報告: {report_path}")
    else:
        print("\nMarkdown report was not saved. / HTML 報告未成功輸出。")
    if summary_excel_path:
        print(f"Saved summary Excel / 已儲存 Summary Excel: {summary_excel_path}")

    print(f"Saved raw outputs / 已儲存原始輸出: {raw_outputs_path}")
    if chart_path:
        print(f"Saved chart / 已儲存圖表: {chart_path}")
    else:
        print("Skipped chart output because there were no eligible successful results. / 因為沒有可繪圖的成功結果，所以略過圖表輸出。")

if __name__ == "__main__":
    exit_code = 0
    try:
        ensure_runtime_ready()
        main()
    except KeyboardInterrupt:
        print("\nExecution cancelled. / 已取消執行。")
    except Exception as exc:
        exit_code = 1
        handle_fatal_error(exc)
    finally:
        pause_before_exit()
    if exit_code:
        sys.exit(exit_code)
