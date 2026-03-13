import unittest
from unittest.mock import patch
from pathlib import Path
import tempfile
import os

import pandas as pd

import ollama_expert_bench as bench


class BenchStandaloneTests(unittest.TestCase):
    def test_build_param_rows_keeps_all_params_visible(self):
        ollama_rows = bench.build_param_rows("ollama")
        ollama_row_by_key = {row.key: row for row in ollama_rows}
        llama_rows = bench.build_param_rows("llama.cpp")
        llama_row_by_key = {row.key: row for row in llama_rows}

        self.assertEqual(set(ollama_row_by_key), set(bench.PARAM_INFO))
        self.assertTrue(ollama_row_by_key["num_ctx"].supported)
        self.assertFalse(llama_row_by_key["num_ctx"].supported)

    def test_build_console_summary_dataframe_includes_new_metrics(self):
        df = pd.DataFrame(
            [
                {
                    "Run_ID": 1,
                    "Status": "ok",
                    "Capability": "chat",
                    "Output_Category": "normal_content",
                    "Diagnosis": "ok",
                    "Finish_Reason": "stop",
                    "Backend": "ollama",
                    "Model": "demo-model",
                    "Params": {"temperature": 0.1},
                    "Applied_Params": {"temperature": 0.1},
                    "Config_Str": "{temperature=0.1}",
                    "TPS": 5.0,
                    "Thinking_TPS": 12.5,
                    "Output_TPS": 8.5,
                    "Output_Thinking_Ratio": 0.68,
                    "Output_Time_s": 3.176,
                    "TTFT": 0.2,
                    "First_Event_s": 0.1,
                    "Stream_Duration_s": 1.0,
                    "Total_Chunks": 3,
                    "Content_Chunks": 2,
                    "Thinking_Chunks": 1,
                    "Non_Content_Chunks": 1,
                    "Non_Content_Types": "reasoning",
                    "VRAM_Base_MiB": 1000,
                    "VRAM_Peak_MiB": 1200,
                    "VRAM_Delta_MiB": 200,
                    "VRAM_Detail": "demo",
                    "Efficiency_Score": 4.267,
                    "Retained_Sections": ["thinking", "dialogue_output"],
                    "Thinking_Chars": 40,
                    "Thinking_Text": "thinking text",
                    "Dialogue_Output_Chars": 27,
                    "Dialogue_Output_Text": "output text",
                    "Output_Chars": 27,
                    "Output_Text": "output text",
                    "Error": "",
                }
            ]
        )

        console_df = bench.build_console_summary_dataframe(df)

        self.assertIn("Thinking TPS (char/s)", console_df.columns)
        self.assertIn("Output TPS (char/s)", console_df.columns)
        self.assertIn("Output/Thinking Ratio", console_df.columns)
        self.assertNotIn("Total Output (chars)", console_df.columns)
        self.assertNotIn("Total Output Time (s)", console_df.columns)
        self.assertEqual(console_df.loc[0, "Thinking TPS (char/s)"], "12.50")
        self.assertEqual(console_df.loc[0, "Output TPS (char/s)"], "8.50")
        self.assertEqual(console_df.loc[0, "Output/Thinking Ratio"], "0.680")

    def test_tools_report_includes_success_stats_and_wrapped_headers(self):
        from openpyxl import load_workbook

        df = pd.DataFrame(
            [
                {
                    "Run_ID": 1,
                    "Status": "ok",
                    "Capability": "tools",
                    "Output_Category": "tool_call",
                    "Diagnosis": "tool ok",
                    "Finish_Reason": "tool_calls",
                    "Backend": "ollama",
                    "Model": "model-a",
                    "System_Prompt_Label": "SP1",
                    "System_Prompt_Text": "Be terse.",
                    "Params": {"temperature": 0.1},
                    "Applied_Params": {"temperature": 0.1},
                    "Config_Str": "{temperature=0.1} | system_prompt=SP1",
                    "TPS": None,
                    "Thinking_TPS": None,
                    "Output_TPS": None,
                    "Output_Thinking_Ratio": None,
                    "Output_Time_s": None,
                    "TTFT": None,
                    "First_Event_s": 0.1,
                    "Stream_Duration_s": 0.4,
                    "Total_Chunks": 2,
                    "Content_Chunks": 0,
                    "Thinking_Chunks": 0,
                    "Non_Content_Chunks": 1,
                    "Non_Content_Types": "tool_calls",
                    "VRAM_Base_MiB": 1000,
                    "VRAM_Peak_MiB": 1200,
                    "VRAM_Delta_MiB": 200,
                    "VRAM_Detail": "demo",
                    "Efficiency_Score": None,
                    "Retained_Sections": [],
                    "Thinking_Chars": 0,
                    "Thinking_Text": "",
                    "Dialogue_Output_Chars": 0,
                    "Dialogue_Output_Text": "",
                    "Output_Chars": 0,
                    "Output_Text": "",
                    "Error": "",
                },
                {
                    "Run_ID": 2,
                    "Status": "warning",
                    "Capability": "tools",
                    "Output_Category": "text_reply_without_tool",
                    "Diagnosis": "no tool",
                    "Finish_Reason": "stop",
                    "Backend": "ollama",
                    "Model": "model-a",
                    "System_Prompt_Label": "SP2",
                    "System_Prompt_Text": "Be verbose.",
                    "Params": {"temperature": 0.8},
                    "Applied_Params": {"temperature": 0.8},
                    "Config_Str": "{temperature=0.8} | system_prompt=SP2",
                    "TPS": 3.0,
                    "Thinking_TPS": 10.0,
                    "Output_TPS": 5.0,
                    "Output_Thinking_Ratio": 0.5,
                    "Output_Time_s": 2.0,
                    "TTFT": 0.3,
                    "First_Event_s": 0.2,
                    "Stream_Duration_s": 0.8,
                    "Total_Chunks": 3,
                    "Content_Chunks": 1,
                    "Thinking_Chunks": 1,
                    "Non_Content_Chunks": 1,
                    "Non_Content_Types": "reasoning",
                    "VRAM_Base_MiB": 1000,
                    "VRAM_Peak_MiB": 1200,
                    "VRAM_Delta_MiB": 200,
                    "VRAM_Detail": "demo",
                    "Efficiency_Score": 2.56,
                    "Retained_Sections": ["thinking", "dialogue_output"],
                    "Thinking_Chars": 20,
                    "Thinking_Text": "thinking",
                    "Dialogue_Output_Chars": 10,
                    "Dialogue_Output_Text": "output",
                    "Output_Chars": 10,
                    "Output_Text": "output",
                    "Error": "",
                },
                {
                    "Run_ID": 3,
                    "Status": "ok",
                    "Capability": "tools",
                    "Output_Category": "tool_call",
                    "Diagnosis": "tool ok",
                    "Finish_Reason": "tool_calls",
                    "Backend": "ollama",
                    "Model": "model-b",
                    "System_Prompt_Label": "SP1",
                    "System_Prompt_Text": "Be terse.",
                    "Params": {"temperature": 0.1},
                    "Applied_Params": {"temperature": 0.1},
                    "Config_Str": "{temperature=0.1} | system_prompt=SP1",
                    "TPS": None,
                    "Thinking_TPS": None,
                    "Output_TPS": None,
                    "Output_Thinking_Ratio": None,
                    "Output_Time_s": None,
                    "TTFT": None,
                    "First_Event_s": 0.1,
                    "Stream_Duration_s": 0.4,
                    "Total_Chunks": 2,
                    "Content_Chunks": 0,
                    "Thinking_Chunks": 0,
                    "Non_Content_Chunks": 1,
                    "Non_Content_Types": "tool_calls",
                    "VRAM_Base_MiB": 1000,
                    "VRAM_Peak_MiB": 1200,
                    "VRAM_Delta_MiB": 200,
                    "VRAM_Detail": "demo",
                    "Efficiency_Score": None,
                    "Retained_Sections": [],
                    "Thinking_Chars": 0,
                    "Thinking_Text": "",
                    "Dialogue_Output_Chars": 0,
                    "Dialogue_Output_Text": "",
                    "Output_Chars": 0,
                    "Output_Text": "",
                    "Error": "",
                },
            ]
        )
        config = {
            "backend": "ollama",
            "capability": "tools",
            "url": "http://localhost:11434/v1",
            "models": ["model-a", "model-b"],
            "prompt": "tool prompt",
            "vram_monitoring": "nvidia-smi",
            "system_prompts": ["Be terse.", "Be verbose."],
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            summary_excel_path = bench.save_summary_excel_workbook(
                df, config, str(Path(temp_dir) / "report")
            )
            report_path = bench.save_markdown_report(
                df,
                config,
                str(Path(temp_dir) / "report"),
                summary_excel_path=summary_excel_path,
            )
            report_text = Path(report_path).read_text(encoding="utf-8")
            workbook = load_workbook(summary_excel_path)
            self.assertTrue(Path(summary_excel_path).exists())

        self.assertEqual(Path(report_path).suffix, ".html")
        self.assertEqual(Path(summary_excel_path).suffix, ".xlsx")
        self.assertEqual(workbook.sheetnames, ["Summary", "Outcome Summary", "Tool Call Success"])
        self.assertIn("Status", str(workbook["Summary"]["B1"].value))
        self.assertIn("狀態", str(workbook["Summary"]["B1"].value))
        self.assertIn("<!DOCTYPE html>", report_text)
        self.assertIn("<table", report_text)
        self.assertIn("Total Output Time", report_text)
        self.assertIn("總輸出時間", report_text)
        self.assertIn("Test Matrix", report_text)
        self.assertIn("測試矩陣", report_text)
        self.assertIn('class="matrix-table"', report_text)
        self.assertIn("Tool Call Success by Model", report_text)
        self.assertIn("各模型工具呼叫成功統計", report_text)
        self.assertIn("Tool Call Success", report_text)
        self.assertIn("Total Output", report_text)
        self.assertIn("Thinking TPS", report_text)
        self.assertIn("Output Category", report_text)
        self.assertIn("System Prompt Variants", report_text)
        self.assertIn("系統提示變體", report_text)
        self.assertIn("Be terse.", report_text)
        self.assertIn("Be verbose.", report_text)
        self.assertIn("model-a", report_text)
        self.assertIn("50.0%", report_text)
        self.assertIn("100.0%", report_text)
        self.assertIn("Download Excel / 下載 Excel 摘要", report_text)
        self.assertIn(Path(summary_excel_path).name, report_text)

    def test_export_best_config_writes_into_report_directory(self):
        df = pd.DataFrame(
            [
                {
                    "Run_ID": 1,
                    "Status": "ok",
                    "Capability": "chat",
                    "Output_Category": "normal_content",
                    "Diagnosis": "ok",
                    "Finish_Reason": "stop",
                    "Backend": "ollama",
                    "Model": "model-a",
                    "Params": {"temperature": 0.1},
                    "Applied_Params": {"temperature": 0.1},
                    "Config_Str": "{temperature=0.1}",
                    "TPS": 8.0,
                    "TTFT": 0.2,
                    "First_Event_s": 0.1,
                    "Stream_Duration_s": 1.0,
                    "Total_Chunks": 3,
                    "Content_Chunks": 2,
                    "Non_Content_Chunks": 1,
                    "Non_Content_Types": "reasoning",
                    "VRAM_Base_MiB": 1000,
                    "VRAM_Peak_MiB": 1200,
                    "VRAM_Delta_MiB": 200,
                    "VRAM_Detail": "demo",
                    "Efficiency_Score": 6.827,
                }
            ]
        )
        config = {
            "backend": "ollama",
            "capability": "chat",
            "prompt": "demo prompt",
        }

        with tempfile.TemporaryDirectory() as temp_dir:
            report_dir = bench.ensure_report_output_dir(temp_dir)
            artifacts = bench.export_best_config(df, config, output_dir=report_dir)

            self.assertEqual(Path(artifacts["best_config_path"]).parent, report_dir)
            self.assertTrue(Path(artifacts["best_config_path"]).exists())
            self.assertEqual(Path(artifacts["modelfile_path"]).parent, report_dir)
            self.assertTrue(Path(artifacts["modelfile_path"]).exists())

    def test_main_writes_runtime_artifacts_into_report_directory(self):
        config = {
            "backend": "ollama",
            "capability": "chat",
            "models": ["demo-model"],
            "prompt": "demo prompt",
            "params": {},
        }
        results_df = pd.DataFrame([{"Status": "ok"}])

        recorded = {}

        def fake_save_raw_outputs(df, report_stem):
            path = Path(f"{report_stem}_outputs.jsonl")
            path.write_text("", encoding="utf-8")
            recorded["raw_outputs_path"] = path.resolve()
            return path

        def fake_save_markdown_report(df, cfg, report_stem, summary_excel_path=None):
            path = Path(f"{report_stem}.html")
            path.write_text("<!DOCTYPE html>", encoding="utf-8")
            recorded["report_path"] = path.resolve()
            recorded["summary_excel_arg"] = None if summary_excel_path is None else Path(summary_excel_path).resolve()
            return path

        def fake_save_summary_excel_workbook(df, cfg, report_stem):
            path = Path(f"{report_stem}_summary.xlsx")
            path.write_text("excel", encoding="utf-8")
            recorded["summary_excel_path"] = path.resolve()
            return path

        def fake_plot_results(df, output_path, capability="chat"):
            path = Path(output_path)
            path.write_text("chart", encoding="utf-8")
            recorded["chart_path"] = path.resolve()
            recorded["plot_capability"] = capability
            return path

        def fake_export_best_config(df, cfg, output_dir="."):
            recorded["best_config_output_dir"] = Path(output_dir).resolve()
            return {
                "best_config_path": str(Path(output_dir) / "best_config.json"),
                "modelfile_path": str(Path(output_dir) / "Ollama_Modelfile_Suggest"),
            }

        with tempfile.TemporaryDirectory() as temp_dir:
            previous_cwd = Path.cwd()
            os.chdir(temp_dir)
            try:
                with patch.object(bench, "interactive_config", return_value=config), patch.object(
                    bench, "run_bench", return_value=results_df
                ), patch.object(
                    bench,
                    "build_console_summary_dataframe",
                    return_value=pd.DataFrame([{"Run": "1"}]),
                ), patch.object(
                    bench,
                    "dataframe_to_text_table",
                    return_value="demo-table",
                ), patch.object(
                    bench,
                    "save_raw_outputs",
                    side_effect=fake_save_raw_outputs,
                ), patch.object(
                    bench,
                    "save_summary_excel_workbook",
                    side_effect=fake_save_summary_excel_workbook,
                ), patch.object(
                    bench,
                    "save_markdown_report",
                    side_effect=fake_save_markdown_report,
                ), patch.object(
                    bench,
                    "plot_results",
                    side_effect=fake_plot_results,
                ), patch.object(
                    bench,
                    "export_best_config",
                    side_effect=fake_export_best_config,
                ):
                    bench.main()
            finally:
                os.chdir(previous_cwd)

            report_dir = (Path(temp_dir) / "Report").resolve()
            self.assertTrue(report_dir.is_dir())
            self.assertEqual(recorded["raw_outputs_path"].parent, report_dir)
            self.assertEqual(recorded["report_path"].parent, report_dir)
            self.assertEqual(recorded["summary_excel_path"].parent, report_dir)
            self.assertEqual(recorded["summary_excel_arg"], recorded["summary_excel_path"])
            self.assertEqual(recorded["chart_path"].parent, report_dir)
            self.assertEqual(recorded["best_config_output_dir"], report_dir)
            self.assertEqual(recorded["plot_capability"], "chat")


if __name__ == "__main__":
    unittest.main()
